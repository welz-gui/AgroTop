"""
AgroTop — Sistema de Gestão de Gado de Corte
PWA responsivo: Streamlit + SQLite + Plotly
"""

import io
import html
import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import date, datetime, timedelta
from typing import Optional  # usado em _decode_qr e _ocr_number
import database as db
from services.constantes import AGE_BANDS
from services.qualidade import avaliar_pesagem
from services.identificadores import REGRAS_PADRAO, validar as validar_formato_id
from services.validacao_regulatoria import validar_animal
from services.recomendacoes import avaliar as avaliar_recomendacoes
import json
from services.regras_regulatorias import simular as simular_regra_pura
from services.geometria import (
    area_hectares as geometria_area_ha,
    centroide as geometria_centroide,
    perimetro_metros as geometria_perimetro_m,
    validar as geometria_validar,
)
from services.lotacao import sobrepostos as lotacao_sobrepostos
from services.sincronizacao import (
    SITUACOES as SITUACOES_SINCRONIZACAO,
    RESOLVIDAS as SITUACOES_RESOLVIDAS,
    SISTEMAS as SISTEMAS_SINCRONIZACAO,
    resumo as sincronizacao_resumo,
    rotulo as sincronizacao_rotulo,
)
from services.estados_dispositivo import (
    ESTADOS as ESTADOS_DISPOSITIVO,
    transicao_permitida as _transicao_dispositivo,
)
from services.previsao_estoque import prever as previsao_estoque_prever
from services.previsao_estoque_adaptador import (
    consumo_diario_planejado,
    montar_insumos as previsao_estoque_montar_insumos,
)
from services.arquivo_dispositivos import (
    ler as arquivo_dispositivos_ler,
    conferir_pareamento as arquivo_dispositivos_conferir_pareamento,
)
from services.reconciliacao_dispositivos import reconciliar as dispositivos_reconciliar
from services.lancamentos import normalizar as lancamentos_normalizar
from services.caixa import resultado_por_competencia, fluxo_de_caixa, em_aberto as caixa_em_aberto
from services.dre import montar_dre
from services.centros_de_custo import consolidar as consolidar_centros_de_custo
from services.rentabilidade_adaptador import montar_ciclos
from services.rentabilidade import ranking_por_raca, por_lote_de_venda
from services.completude_adaptador import normalizar_pesagens, janela_do_mes
from services.completude import avaliar_mes
from services.conformidade_adaptador import montar_rebanho
from services.conformidade import avaliar as conformidade_avaliar
from services.dieta_adaptador import ingredientes_por_cabeca
from services.dieta import custo_por_cabeca_dia, custo_por_arroba_produzida
from services.projecao_adaptador import series_mensais
from services.projecao import correlacao_chuva_gmd
from services.rateio_adaptador import com_dias_no_lote
from services.rateio import ratear
from services.gta_adaptador import montar_contexto as gta_montar_contexto
from services.gta import validar as gta_validar
from ui.tema import cores, plotly_layout, SERIES, ESCALA_RUIM_BOM, ESCALA_BOM_RUIM

# ─── Configuração da página ───────────────────────────────────────────────────
st.set_page_config(
    page_title="AgroTop — Gestão de Gado",
    page_icon="🐄",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={"About": "AgroTop v2.0 — Gestão Completa de Gado de Corte"},
)

db.init_db()
db.refresh_carencia_status()
c = cores()
c = cores()

# ─── CSS ──────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
.main .block-container{padding-top:.8rem;padding-bottom:2rem}
section[data-testid="stSidebar"]{background:#0a1628!important}

/* Botões grandes — uso ao sol */
.stButton>button{min-height:2.75rem;font-size:1rem;font-weight:600;border-radius:10px;transition:all .15s}
.stButton>button:hover{transform:translateY(-1px)}

/* Métricas */
div[data-testid="stMetric"]{background:#1e293b;border:1px solid #334155;border-radius:14px;padding:1rem 1.2rem}
div[data-testid="stMetricValue"]{font-size:1.55rem!important}

/* Título de página */
.page-title{font-size:1.55rem;font-weight:800;color:#4ade80;border-left:4px solid #4ade80;padding-left:.75rem;margin-bottom:1.4rem}

/* Cards */
.card{background:#1e293b;border:1px solid #334155;border-radius:16px;padding:1.2rem 1.5rem;margin-bottom:1rem}
.card-green{background:linear-gradient(135deg,#14532d,#0f172a);border:1px solid #166534;border-radius:16px;padding:1.2rem 1.5rem;margin-bottom:1rem}
.card-yellow{background:linear-gradient(135deg,#422006,#0f172a);border:1px solid #854d0e;border-radius:16px;padding:1.2rem 1.5rem;margin-bottom:1rem}
.card-red{background:linear-gradient(135deg,#450a0a,#0f172a);border:1px solid #7f1d1d;border-radius:16px;padding:1.2rem 1.5rem;margin-bottom:1rem}

/* Badges */
.badge-green {background:#166534;color:#4ade80;padding:2px 10px;border-radius:999px;font-size:.78rem;font-weight:700}
.badge-yellow{background:#713f12;color:#fbbf24;padding:2px 10px;border-radius:999px;font-size:.78rem;font-weight:700}
.badge-red   {background:#7f1d1d;color:#f87171;padding:2px 10px;border-radius:999px;font-size:.78rem;font-weight:700}
.badge-blue  {background:#1e3a5f;color:#60a5fa;padding:2px 10px;border-radius:999px;font-size:.78rem;font-weight:700}
.badge-gray  {background:#1e293b;color:#94a3b8;padding:2px 10px;border-radius:999px;font-size:.78rem;font-weight:700}

/* Linha de histórico */
.hist-item{background:#0f172a;border-left:3px solid #4ade80;border-radius:8px;padding:.45rem .9rem;margin-bottom:.35rem}

/* Teclado numérico */
.keypad-display{font-size:3rem;font-weight:900;color:#4ade80;text-align:center;background:#0f172a;border:2px solid #334155;border-radius:14px;padding:1rem;margin-bottom:.5rem;letter-spacing:.1em}

/* Ocultar elementos padrão (mantendo o botão de abrir o menu lateral) */
#MainMenu,footer{visibility:hidden}
header[data-testid="stHeader"]{background:transparent}
/* Botão para reabrir o menu lateral — sempre visível e destacado */
[data-testid="collapsedControl"]{
    visibility:visible!important;
    display:flex!important;
    opacity:1!important;
    z-index:999999;
}
[data-testid="collapsedControl"] button{
    background:#4ade80!important;
    color:#0f172a!important;
    border-radius:8px;
}

/* Mobile */
@media(max-width:640px){
  .main .block-container{padding-left:.4rem;padding-right:.4rem}
  .stButton>button{min-height:3.2rem;font-size:1.1rem}
  div[data-testid="stMetricValue"]{font-size:1.3rem!important}
}
</style>
""", unsafe_allow_html=True)

# ─── Constantes ───────────────────────────────────────────────────────────────
BREEDS = ["Nelore","Angus","Brahman","Senepol","Brangus","Canchim","Simental",
          "Hereford","Charolês","Tabapuã","Outro"]
ROUTES = ["Subcutânea","Intramuscular","Oral","Intravenosa","Tópica"]
COST_TYPES = ["compra","insumo","operacional","veterinário","outro"]
# Cotação padrão centralizada (usada no Simulador e no Relatório Financeiro)
DEFAULT_PRICE_ARROBA = 320.0   # R$ por arroba (boi gordo)
DEFAULT_PRICE_KG     = 10.0    # R$ por kg de boi vivo
PLOTLY = plotly_layout()

def _layout(**overrides):
    """Mescla o layout padrão PLOTLY com overrides (evita conflito de kwargs)."""
    base = dict(PLOTLY)
    base.update(overrides)
    return base

# ─── Session State ─────────────────────────────────────────────────────────────
_DEFAULTS = dict(
    authenticated=False, user=None, page="dashboard",
    animal_detail=None, campo_id="", keypad_value="",
    unit_pref="kg",   # "kg" ou "@"
)
for k, v in _DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ─── Helpers ──────────────────────────────────────────────────────────────────
def _go(page, animal_id=None):
    st.session_state.page = page
    if animal_id is not None:
        st.session_state.animal_detail = animal_id

def _status_badge(status):
    m = {"ativo":"badge-green ● Ativo","vendido":"badge-blue ✓ Vendido",
         "morto":"badge-red ✕ Morto","carencia":"badge-yellow ⚠ Carência"}
    parts = m.get(status, "badge-gray "+status).split(" ",1)
    return f'<span class="{parts[0]}">{parts[1]}</span>'

def _gmd_badge(gmd):
    if gmd is None: return '<span class="badge-gray">— N/D</span>'
    if gmd > 0:     return f'<span class="badge-green">▲ {gmd:.3f} kg/dia</span>'
    if gmd < 0:     return f'<span class="badge-red">▼ {gmd:.3f} kg/dia</span>'
    return f'<span class="badge-yellow">= {gmd:.3f} kg/dia</span>'

# ─── Helpers de unidade (kg / @) ─────────────────────────────────────────────
def _use_arroba() -> bool:
    return st.session_state.get("unit_pref", "kg") == "@"

def _unit_label() -> str:
    """Retorna o símbolo da unidade configurada."""
    return "@" if _use_arroba() else "kg"

def _prod_weight(kg_gain: float, yield_: float = 0.52) -> float:
    """Converte ganho em kg para a unidade configurada (@ ou kg vivo)."""
    if _use_arroba():
        return db.kg_to_arrobas(kg_gain, yield_)
    return round(kg_gain, 1)

def _live_weight(kg: float, yield_: float = 0.52) -> float:
    """Converte peso vivo para a unidade configurada."""
    if _use_arroba():
        return db.kg_to_arrobas(kg, yield_)
    return round(kg, 1)

def _fmt_live(kg: float, yield_: float = 0.52) -> str:
    """Formata peso vivo na unidade configurada."""
    val = _live_weight(kg, yield_)
    return f"{val:.2f} {_unit_label()}" if _use_arroba() else f"{val:.1f} kg"

def _cost_per_unit_label() -> str:
    # Deixa explícito que é sobre o PESO VIVO ATUAL (não sobre o ganho)
    return "Custo/@ vivo (R$)" if _use_arroba() else "Custo/kg vivo (R$)"

def _cost_per_unit(total_cost: float, kg: float, yield_: float = 0.52) -> float:
    """Custo total dividido pelo peso vivo atual (na unidade configurada)."""
    denom = _live_weight(kg, yield_)
    return round(total_cost / denom, 2) if denom else 0

def _breakeven_label() -> str:
    return "Breakeven (R$/@)" if _use_arroba() else "Breakeven (R$/kg)"

# ─── Formatação PT-BR (plural e números) ─────────────────────────────────────
def _plural(n, singular: str, plural: str = None) -> str:
    """Ex.: _plural(1,'animal','animais') -> '1 animal'; _plural(3,...) -> '3 animais'."""
    plural = plural or (singular + "s")
    return f"{n} {singular if n == 1 else plural}"

def _num_br(v, casas: int = 1) -> str:
    """Número no padrão brasileiro: 10.0 -> '10,0'."""
    try:
        return f"{float(v):.{casas}f}".replace(".", ",")
    except (TypeError, ValueError):
        return str(v)

def _fmt_dose(dose, unit: str) -> str:
    """Dose + unidade formatadas: '10,0 ml', '2,0 doses', '1,0 comprimido'."""
    plurais = {"dose": "doses", "comprimido": "comprimidos"}
    try:
        d = float(dose)
    except (TypeError, ValueError):
        return f"{dose} {unit}"
    u = plurais.get(unit, unit) if d != 1 else unit
    return f"{_num_br(d)} {u}"

# ─── Previsão do tempo (Open-Meteo, gratuito e sem chave) ────────────────────
@st.cache_data(ttl=3600, show_spinner=False)
def _fetch_forecast(lat: float, lon: float):
    """Previsão de 7 dias para a coordenada da fazenda. Cache de 1h."""
    import urllib.request, json
    url = (f"https://api.open-meteo.com/v1/forecast?latitude={lat}&longitude={lon}"
           "&daily=temperature_2m_max,temperature_2m_min,precipitation_sum,"
           "precipitation_probability_max&timezone=America%2FSao_Paulo&forecast_days=7")
    try:
        with urllib.request.urlopen(url, timeout=10) as r:
            return json.loads(r.read().decode())
    except Exception:
        return None

# ─── Câmera: imagem, QR Code e OCR ───────────────────────────────────────────
def _compress_image(raw: bytes, max_side: int = 1000, quality: int = 75) -> bytes:
    """Redimensiona e comprime a foto para JPEG (economiza espaço no banco)."""
    try:
        from PIL import Image, ImageOps
        import io as _io
        img = Image.open(_io.BytesIO(raw))
        img = ImageOps.exif_transpose(img)          # corrige rotação do celular
        img = img.convert("RGB")
        img.thumbnail((max_side, max_side))
        buf = _io.BytesIO()
        img.save(buf, format="JPEG", quality=quality, optimize=True)
        return buf.getvalue()
    except Exception:
        return raw

def _decode_qr(raw: bytes) -> Optional[str]:
    """Tenta decodificar um QR Code na imagem. Retorna o texto ou None."""
    try:
        import cv2, numpy as np, io as _io
        from PIL import Image
        img = Image.open(_io.BytesIO(raw)).convert("RGB")
        arr = cv2.cvtColor(np.array(img), cv2.COLOR_RGB2BGR)
        detector = cv2.QRCodeDetector()
        data, _pts, _ = detector.detectAndDecode(arr)
        return data.strip() if data else None
    except Exception:
        return None

def _ocr_number(raw: bytes) -> Optional[str]:
    """Best-effort: tenta ler dígitos do brinco (OCR). Impreciso — sempre confirmar."""
    try:
        import pytesseract, io as _io
        from PIL import Image, ImageOps
        img = Image.open(_io.BytesIO(raw)).convert("L")
        img = ImageOps.autocontrast(img)
        txt = pytesseract.image_to_string(
            img, config="--psm 7 -c tessedit_char_whitelist=0123456789")
        digitos = "".join(ch for ch in txt if ch.isdigit())
        return digitos or None
    except Exception:
        return None

def _photo_section(animal_id: str, key_prefix: str = ""):
    """Widget de foto do animal: mostra a atual, tira nova (auto-salva) e histórico.
    A câmera só liga após clicar em 'Abrir câmera' (evita pedir permissão sempre)."""
    fotos = db.get_photos(animal_id)
    latest = db.get_latest_photo(animal_id)
    if latest:
        st.image(latest[0], caption=f"Foto mais recente · {fotos[0]['taken_date']}", width=300)
    else:
        st.info("📷 Sem foto ainda. Clique em **Abrir câmera** para tirar a primeira.")

    open_key = f"{key_prefix}foto_cam_on"
    if not st.session_state.get(open_key):
        if st.button("📷 Abrir câmera para tirar foto", key=f"{key_prefix}foto_open",
                     use_container_width=True, type="primary"):
            st.session_state[open_key] = True; st.rerun()
    else:
        nova = st.camera_input("Tirar/atualizar foto", key=f"{key_prefix}cam_foto")
        if nova is not None:
            raw = nova.getvalue()
            sig = f"{animal_id}:{hash(raw)}"
            if st.session_state.get("_last_photo_sig") != sig:
                comp = _compress_image(raw)
                db.add_photo(animal_id, comp, operator=st.session_state.user["name"])
                st.session_state["_last_photo_sig"] = sig
                st.success("📷 Foto salva!")
                st.session_state[open_key] = False
                st.rerun()
        if st.button("✖️ Fechar câmera", key=f"{key_prefix}foto_close",
                     use_container_width=True):
            st.session_state[open_key] = False; st.rerun()

    if len(fotos) > 1:
        with st.expander(f"📸 Histórico de fotos ({len(fotos)})"):
            for f in fotos:
                img = db.get_photo_image(f["id"])
                if img:
                    fc1, fc2 = st.columns([3,1])
                    with fc1:
                        st.image(img[0], caption=f["taken_date"], width=200)
                    with fc2:
                        if st.button("🗑️", key=f"{key_prefix}delph_{f['id']}"):
                            db.delete_photo(f["id"]); st.rerun()

def _df_to_csv(df: pd.DataFrame) -> bytes:
    buf = io.StringIO()
    df.where(pd.notna(df), "").to_csv(buf, index=False, encoding="utf-8-sig")
    return buf.getvalue().encode("utf-8-sig")

def _df_to_xlsx(title: str, df: pd.DataFrame) -> bytes:
    """Exporta o DataFrame para Excel (.xlsx) com cabeçalho estilizado."""
    buf = io.BytesIO()
    sheet = (title[:28] or "Dados").replace("/", "-").replace("\\", "-")
    df = df.where(pd.notna(df), "")   # substitui NaN/None por vazio
    try:
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet, startrow=1)
            wb = writer.book
            ws = writer.sheets[sheet]
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            # Título mesclado no topo
            ncols = max(len(df.columns), 1)
            ws.cell(row=1, column=1, value=f"AgroTop — {title}")
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
            ws.cell(row=1, column=1).font = Font(bold=True, size=13, color="166534")
            ws.cell(row=1, column=1).alignment = Alignment(horizontal="left")

            # Cabeçalho das colunas (linha 2)
            header_fill = PatternFill("solid", fgColor="14532D")
            thin = Side(style="thin", color="D0D0D0")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for col_idx, col_name in enumerate(df.columns, start=1):
                c = ws.cell(row=2, column=col_idx)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = header_fill
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = border

            # Bordas e largura automática das colunas
            for col_idx, col_name in enumerate(df.columns, start=1):
                max_len = len(str(col_name))
                for v in df.iloc[:, col_idx - 1].tolist():
                    max_len = max(max_len, len(str(v)))
                ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = min(max_len + 3, 40)
            ws.freeze_panes = "A3"
        return buf.getvalue()
    except Exception:
        return b""

def _pdf_safe(text) -> str:
    """Sanitiza texto para a fonte Helvetica (Latin-1) do fpdf.
    Substitui caracteres Unicode não suportados; mantém acentos do português."""
    s = str(text)
    repl = {
        "—": "-", "–": "-", "→": "->", "←": "<-", "•": "*", "●": "*",
        "○": "o", "▲": "^", "▼": "v", "⚠️": "!", "⚠": "!",
        "♂": "M", "♀": "F", "≥": ">=", "≤": "<=", "…": "...",
        "“": '"', "”": '"', "‘": "'", "’": "'", " ": " ",
        "🐄": "", "📄": "", "🏷️": "", "✅": "", "❌": "", "⬇️": "",
    }
    for k, v in repl.items():
        s = s.replace(k, v)
    # Remove qualquer emoji/símbolo restante fora do Latin-1
    return s.encode("latin-1", "replace").decode("latin-1")


def _backup_xlsx() -> bytes:
    """Exporta TODAS as tabelas para um único Excel (uma aba por tabela)."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for t in db.ADMIN_TABLES:
            try:
                rows = db.admin_get_rows(t)
                df = pd.DataFrame(rows) if rows else pd.DataFrame({"(vazio)": []})
                df = df.where(pd.notna(df), "")
                df.to_excel(writer, index=False, sheet_name=t[:31])
            except Exception:
                pd.DataFrame({"erro": [f"falha ao ler {t}"]}).to_excel(
                    writer, index=False, sheet_name=t[:31])
    return buf.getvalue()


def _df_to_pdf(title: str, df: pd.DataFrame) -> bytes:
    try:
        from fpdf import FPDF
        df = df.where(pd.notna(df), "")   # evita "nan" no PDF
        pdf = FPDF(orientation="L")   # paisagem — comporta mais colunas
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, _pdf_safe(title), ln=True, align="C")
        pdf.set_font("Helvetica", "", 6)
        cols  = list(df.columns)
        n     = max(len(cols), 1)
        col_w = max(min(277 // n, 45), 12)   # largura útil em paisagem ~277mm
        # Cabeçalho
        pdf.set_fill_color(30, 60, 30)
        pdf.set_text_color(200, 255, 200)
        pdf.set_font("Helvetica", "B", 6)
        for col in cols:
            pdf.cell(col_w, 6, _pdf_safe(col)[:22], border=1, fill=True, align="C")
        pdf.ln()
        pdf.set_font("Helvetica", "", 6)
        pdf.set_text_color(20, 20, 20)
        for i, row in enumerate(df.itertuples(index=False)):
            if i % 2 == 0:
                pdf.set_fill_color(245, 245, 245)
            else:
                pdf.set_fill_color(255, 255, 255)
            for val in row:
                pdf.cell(col_w, 5, _pdf_safe(val)[:22], border=1, fill=True)
            pdf.ln()
        out = pdf.output()
        return bytes(out)
    except ImportError:
        return b""
    except Exception:
        # Falha inesperada não deve derrubar a página de relatórios
        return b""

# ══════════════════════════════════════════════════════════════════════════════
# LOGIN
# ══════════════════════════════════════════════════════════════════════════════
def page_login():
    _, col, _ = st.columns([1,1.6,1])
    with col:
        st.markdown("""
        <div style="text-align:center;padding:2rem 0 1.5rem">
            <div style="font-size:4.5rem;line-height:1">🐄</div>
            <h1 style="color:{c["primaria"]};margin:.4rem 0 0">AgroTop</h1>
            <p style="color:{c["texto_terciario"]};margin:0;font-size:.95rem">Sistema de Gestão de Gado de Corte</p>
        </div>""", unsafe_allow_html=True)
        with st.form("login"):
            user = st.text_input("👤 Usuário", placeholder="seu.usuario")
            pwd  = st.text_input("🔑 Senha", type="password", placeholder="••••••••")
            lembrar = st.checkbox("Manter conectado neste dispositivo", value=True)
            if st.form_submit_button("🔓  Entrar no Sistema", use_container_width=True, type="primary"):
                u = db.verify_login(user.strip(), pwd)
                if u:
                    st.session_state.authenticated = True
                    st.session_state.user = u
                    # Página inicial conforme o perfil
                    st.session_state.page = "dashboard" if u["role"]=="admin" else "campo"
                    # Login persistente: token em COOKIE (nunca na URL)
                    if lembrar:
                        cm = _cookie_manager()
                        if cm is not None:
                            try:
                                token = db.create_session(u["id"])
                                cm.set(_COOKIE_NAME, token,
                                       expires_at=datetime.now()+timedelta(days=7),
                                       key="set_sid")
                                import time as _t; _t.sleep(0.6)  # tempo p/ o navegador gravar
                            except Exception:
                                pass
                    st.rerun()
                else:
                    st.error("Usuário ou senha inválidos.")

# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
def _sidebar():
    alerts  = db.get_alert_animals()
    low_stk  = db.check_low_stock()
    n_low_perf = len(db.get_low_performance())
    # Badge conta TODOS os itens exibidos na página de Alertas (mesma regra)
    n_alerts = (len(alerts["sumidos"]) + len(alerts["carencia"])
                + len(alerts["prontos"]) + len(low_stk) + n_low_perf)
    user     = st.session_state.user

    with st.sidebar:
        st.markdown(f"""
        <div style="text-align:center;padding:.8rem 0 .5rem">
            <div style="font-size:2.5rem">🐄</div>
            <h2 style="color:{c["primaria"]};margin:0">AgroTop</h2>
            <div style="color:{c["texto_secundario"]};font-size:.8rem;margin-top:.25rem">
                {html.escape(user['name'])}<br>
                <span style="color:{c["primaria"]}">●</span>&nbsp;
                {"Administrador" if user['role']=='admin' else "Operador"}
            </div>
        </div>""", unsafe_allow_html=True)
        st.markdown("---")

        if user["role"] == "admin":
            pages = [
                ("📊","Dashboard","dashboard",""),
                ("📱","Modo Campo","campo",""),
                ("📋","Rebanho","rebanho",""),
                ("🌿","Lotes / Pastagem","lotes",""),
                ("📈","Desempenho","desempenho",""),
                ("💰","Financeiro","financeiro",""),
                ("📦","Estoque","estoque",f" 🔴{len(low_stk)}" if low_stk else ""),
                ("🏷️","Brincos","brincos",""),
                ("🚚","Movimentação","movimentacao",""),
                ("🏞️","Propriedades","propriedades",""),
                ("📜","Regras","regras",""),
                ("📡","Sincronização","sincronizacao",""),
                ("🌾","Nutrição","nutricao",""),
                ("💉","Sanitário","sanitario",""),
                ("🌧️","Clima & Chuva","clima",""),
                ("🔔","Alertas","alertas",f" 🔴{n_alerts}" if n_alerts else ""),
                ("📄","Relatórios","relatorios",""),
                ("➕","Cadastrar Animal","cadastrar",""),
                ("⚙️","Admin","admin",""),
            ]
        else:
            # Operador: apenas manejo de campo, cadastro e estoque
            pages = [
                ("📱","Modo Campo","campo",""),
                ("➕","Cadastrar Animal","cadastrar",""),
                ("📦","Estoque","estoque",f" 🔴{len(low_stk)}" if low_stk else ""),
                ("🏷️","Brincos","brincos",""),
            ]

        for icon, label, key, badge in pages:
            active = st.session_state.page == key
            if st.button(f"{icon}  {label}{badge}", key=f"nav_{key}",
                         use_container_width=True,
                         type="primary" if active else "secondary"):
                _go(key); st.rerun()

        st.markdown("---")

        # ── Configuração de unidade (rádio ligado direto ao session_state) ────
        # Fonte única de verdade: a chave 'unit_pref'. Sem rerun manual, evita
        # a barra lateral e a página ficarem dessincronizadas.
        st.radio("⚖️ Unidade de Produção", ["kg", "@"],
                 key="unit_pref", horizontal=True,
                 help="Define a unidade usada em custos, ganhos e relatórios")

        st.markdown("---")
        stats = db.get_rebanho_stats()
        if stats.total > 0:
            # Usa a MESMA fonte da página (_use_arroba) — sempre consistente
            if _use_arroba():
                prod_str = f"🏷️ <b style='color:{c['atencao']}'>{stats.arrobas_prod:.1f} @</b> ganhas"
            else:
                total_gain_kg = sum(a["current_weight"]-a["entry_weight"]
                                    for a in db.get_all_animals())
                prod_str = f"📦 <b style='color:{c['atencao']}'>{total_gain_kg:.0f} kg</b> ganhos"

            st.markdown(f"""
            <div style="font-size:.78rem;color:{c["texto_terciario"]};text-align:center;line-height:2">
                🐄 <b style="color:{c["texto"]}">{stats.total}</b> animais ativos &nbsp;
                ⚖️ <b style="color:{c["texto"]}">{stats.avg_weight:.0f} kg</b> médio<br>
                📈 GMD <b style="color:{c["primaria"]}">{stats.avg_gmd:.3f} kg/dia</b> &nbsp;
                {prod_str}
            </div>""", unsafe_allow_html=True)
        st.markdown("---")
        if st.button("🚪  Sair", use_container_width=True, type="secondary"):
            cm = _cookie_manager()
            if cm is not None:
                try:
                    tok = cm.get(_COOKIE_NAME)
                    if tok:
                        db.delete_session(tok)
                    cm.delete(_COOKIE_NAME, key="del_sid")
                except Exception:
                    pass
            # Limpa também qualquer resquício de token antigo na URL
            try:
                st.query_params.clear()
            except Exception:
                pass
            for k,v in _DEFAULTS.items(): st.session_state[k] = v
            st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
def page_dashboard():
    st.markdown('<div class="page-title">📊 Dashboard — Visão Geral</div>', unsafe_allow_html=True)
    stats   = db.get_rebanho_stats()
    animals = db.get_all_animals()
    alerts  = db.get_alert_animals()
    if not animals:
        st.info("Nenhum animal cadastrado. Use **Cadastrar Animal** para começar."); return

    # KPIs
    # Produção na unidade configurada
    if _use_arroba():
        prod_label = "🏷️ @ Ganhas"
        prod_value = f"{stats.arrobas_prod:.1f} @"
    else:
        total_gain_kg = sum(a["current_weight"]-a["entry_weight"] for a in animals)
        prod_label = "📦 Ganho Total"
        prod_value = f"{total_gain_kg:.0f} kg"

    k = st.columns(7)
    k[0].metric("🐄 Animais",    stats.total)
    k[1].metric("⚖️ Peso Médio", f"{stats.avg_weight:.1f} kg")
    k[2].metric("📈 GMD Médio",  f"{stats.avg_gmd:.3f} kg/dia")
    k[3].metric(prod_label,      prod_value)
    k[4].metric("🌿 Lotação",    f"{stats.lotacao_ua_ha:.2f} UA/ha")
    k[5].metric("♂ Machos",      stats.males)
    k[6].metric("♀ Fêmeas",      stats.females)

    st.markdown("---")

    # Alertas resumidos
    n_sum = len(alerts["sumidos"]); n_car = len(alerts["carencia"]); n_pro = len(alerts["prontos"])
    if n_sum or n_car or n_pro:
        ac1, ac2, ac3 = st.columns(3)
        with ac1:
            st.markdown(f"""<div class="card-red">
                <b style="color:{c["perigo"]}">🔴 {n_sum} Sumidos</b><br>
                <span style="color:{c["texto_secundario"]};font-size:.85rem">Sem pesagem há +30 dias</span>
            </div>""", unsafe_allow_html=True)
        with ac2:
            st.markdown(f"""<div class="card-yellow">
                <b style="color:{c["atencao"]}">🟡 {n_car} Em Carência</b><br>
                <span style="color:{c["texto_secundario"]};font-size:.85rem">Não podem ser abatidos</span>
            </div>""", unsafe_allow_html=True)
        with ac3:
            st.markdown(f"""<div class="card-green">
                <b style="color:{c["primaria"]}">🟢 {n_pro} Prontos para Abate</b><br>
                <span style="color:{c["texto_secundario"]};font-size:.85rem">Peso-alvo atingido</span>
            </div>""", unsafe_allow_html=True)
        st.markdown("---")

    col_main, col_side = st.columns([3,2])

    with col_main:
        st.subheader("📈 Evolução de Peso do Rebanho")
        raw = db.get_all_weighings()
        if raw:
            df_all = pd.DataFrame(raw)
            df_all["weigh_date"] = pd.to_datetime(df_all["weigh_date"])
            df_avg = (df_all.groupby("weigh_date")["weight"]
                      .mean().reset_index()
                      .rename(columns={"weigh_date":"Data","weight":"Peso Médio (kg)"}))
            df_all = df_all.sort_values(["animal_id", "weigh_date"])
            fig = px.line(
                df_all, x="weigh_date", y="weight", color="animal_id", markers=True,
                color_discrete_sequence=[c["borda"]]
            )
            fig.update_traces(
                showlegend=False, opacity=0.35, line=dict(width=1), marker=dict(size=3),
                hovertemplate="<b>%{fullData.name}</b><br>%{x|%d/%m/%Y}<br>%{y:.1f} kg<extra></extra>"
            )
            fig.add_trace(go.Scatter(x=df_avg["Data"],y=df_avg["Peso Médio (kg)"],
                mode="lines+markers",name="Média do Rebanho",
                line=dict(width=3,color=c["primaria"]),
                marker=dict(size=9,color=c["primaria"],line=dict(width=2,color=c["fundo"])),
                hovertemplate="<b>Média</b><br>%{x|%d/%m/%Y}<br>%{y:.1f} kg<extra></extra>"))
            fig.update_layout(**PLOTLY,height=350,
                legend=dict(orientation="h",yanchor="bottom",y=1.02,xanchor="right",x=1),
                xaxis=dict(gridcolor=c["superficie"],title="Data"),
                yaxis=dict(gridcolor=c["superficie"],title="Peso (kg)"))
            st.plotly_chart(fig, use_container_width=True)

    with col_side:
        st.subheader("🥧 Por Raça")
        df_br = pd.Series([a["breed"] for a in animals]).value_counts().reset_index()
        df_br.columns=["Raça","Qtd"]
        fig_p=px.pie(df_br,names="Raça",values="Qtd",hole=0.45,
            color_discrete_sequence=SERIES)
        fig_p.update_layout(**_layout(height=240,margin=dict(l=0,r=0,t=10,b=10),
            legend=dict(orientation="h",yanchor="bottom",y=-0.2)))
        fig_p.update_traces(textposition="inside",textinfo="percent+label")
        st.plotly_chart(fig_p, use_container_width=True)

        st.subheader("📊 GMD por Animal")
        a_ids = [a["id"] for a in animals]
        gmd_batch = db.calculate_gmd_bulk(a_ids)
        gmd_data=[{"ID":a["id"],"GMD":gmd_batch.get(a["id"])} for a in animals]
        df_g=pd.DataFrame([r for r in gmd_data if r["GMD"] is not None]).sort_values("GMD")
        if not df_g.empty:
            fig_g=px.bar(df_g,x="GMD",y="ID",orientation="h",color="GMD",
                color_continuous_scale=ESCALA_RUIM_BOM,
                labels={"GMD":"kg/dia"})
            fig_g.add_vline(x=0,line_dash="dash",line_color=c["borda_suave"])
            fig_g.update_layout(**PLOTLY,height=max(180,len(df_g)*27),
                coloraxis_showscale=False,
                xaxis=dict(gridcolor=c["superficie"]),yaxis=dict(gridcolor=c["superficie"],title=""))
            st.plotly_chart(fig_g, use_container_width=True)

    # Tabela resumo
    st.markdown("---"); st.subheader("📋 Resumo Rápido")
    rows=[]
    ul = _unit_label()
    a_ids = [a["id"] for a in animals]
    gmd_batch = db.calculate_gmd_bulk(a_ids)
    wd_batch = db.get_withdrawal_end_batch(a_ids)
    for a in animals:
        gmd  = gmd_batch.get(a["id"])
        wd   = wd_batch.get(a["id"])
        gain = round(a["current_weight"]-a["entry_weight"], 1)
        rows.append({"ID":a["id"],"Raça":a["breed"],
            "Sexo":"♂" if a["sex"]=="M" else "♀",
            "Categoria":db.get_age_category(a.get("birth_date")),
            "Lote":a.get("lote_id") or "—",
            "Peso Atual (kg)":a["current_weight"],
            f"Ganho ({ul})":_prod_weight(gain),
            "GMD (kg/dia)":gmd,"Status":a["status"],
            "Carência até":wd.isoformat() if wd else "—"})
    df_sum=pd.DataFrame(rows)
    gain_col = f"Ganho ({ul})"
    fmt_gain = "%.2f" if _use_arroba() else "%.1f"
    st.dataframe(df_sum,use_container_width=True,hide_index=True,height=300,
        column_config={"Peso Atual (kg)":st.column_config.NumberColumn(format="%.1f"),
            gain_col:st.column_config.NumberColumn(format=fmt_gain),
            "GMD (kg/dia)":st.column_config.NumberColumn(format="%.3f")})

    _dash_conformidade()
    _dash_completude()


_FAIXA_CONFORMIDADE = {
    "completo": ("🟢", "Completo"), "bom": ("🟢", "Bom"),
    "atencao": ("🟡", "Atenção"), "critico": ("🔴", "Crítico"),
}


def _dash_conformidade():
    """Escore de conformidade PNIB (spec 0029 + adaptador 0036).

    Indicador de GESTÃO — não substitui avaliação legal (o próprio
    `services.conformidade.avaliar` repete isso em cada dimensão e em cada
    mensagem). Fica aberto por padrão quando a faixa não é "completo"/"bom":
    é o tipo de pendência que precisa aparecer, não esperar alguém abrir.
    """
    animais = db.get_all_animals(status=None)
    identificadores_ativos = [
        item for itens in db.identificadores._por_animal().values()
        for item in itens if item.get("status") == "ativo"
    ]
    dispositivos = db.dispositivos.com_divergencia()
    eventos_pendentes = db.eventos.contar_pendentes()
    movimentacoes_abertas = db.movimentacoes.abertas()
    referencia = date.today().isoformat()

    rebanho = montar_rebanho(
        animais=animais, identificadores_ativos=identificadores_ativos,
        dispositivos=dispositivos, eventos_pendentes=eventos_pendentes,
        movimentacoes_abertas=movimentacoes_abertas, referencia=referencia)
    resultado = conformidade_avaliar(rebanho, referencia)

    emoji, rotulo = _FAIXA_CONFORMIDADE.get(resultado["faixa"], ("⚪", resultado["faixa"]))
    aberto_por_padrao = resultado["faixa"] not in ("completo", "bom")

    with st.expander(
        f"🛡️ Conformidade PNIB — {emoji} {rotulo} ({resultado['escore']:.1f}/100)",
        expanded=aberto_por_padrao):
        st.caption("Indicador de gestão a partir dos dados cadastrados — não substitui "
                   "avaliação de conformidade legal nem certificação oficial.")
        if resultado["prazo_relevante"]:
            st.info(f"ℹ️ Identificação oficial exigível no trânsito a partir de "
                   f"**{resultado['prazo_relevante']}** — animais em preparo contam "
                   "à parte até lá.")

        df_dim = pd.DataFrame([{
            "Dimensão": d["nome"], "Peso (%)": d["peso"], "Nota": d["nota"],
            "Faltam": d["faltam"], "Situação": d["mensagem"],
        } for d in resultado["dimensoes"]])
        st.dataframe(df_dim, use_container_width=True, hide_index=True,
            column_config={
                "Peso (%)": st.column_config.NumberColumn(format="%.0f%%"),
                "Nota": st.column_config.NumberColumn(format="%.1f")})

        fig = px.bar(df_dim, x="Dimensão", y="Nota", color="Nota",
                     color_continuous_scale=ESCALA_RUIM_BOM, range_color=[0, 100],
                     text="Nota")
        fig.update_traces(texttemplate="%{text:.0f}", textposition="outside")
        fig.add_hline(y=70, line_dash="dash", line_color=c["borda_suave"])
        fig.update_layout(**PLOTLY, height=300, coloraxis_showscale=False,
            xaxis=dict(gridcolor=c["superficie"]),
            yaxis=dict(gridcolor=c["superficie"], range=[0, 105]))
        st.plotly_chart(fig, use_container_width=True)

        if resultado["pendencias_criticas"]:
            st.warning("**Pendências:**\n" + "\n".join(
                f"- {p}" for p in resultado["pendencias_criticas"]))
        if resultado["pendencias_informativas"]:
            st.caption("**Também vale saber:** " + " · ".join(
                resultado["pendencias_informativas"]))
        if not resultado["pendencias_criticas"] and not resultado["pendencias_informativas"]:
            st.success("✅ Nenhuma pendência identificada.")


def _dash_completude():
    """Indicador de completude de dados (spec 0016 + adaptador 0035).

    ROADMAP Trilha 4: "mostra, mês a mês, se a base está ficando treinável" —
    é diagnóstico de qualidade de coleta, não operação do dia a dia, por
    isso fica num expander fechado por padrão.
    """
    with st.expander("📋 Completude dos Dados (últimos 3 meses)", expanded=False):
        st.caption("Cinco indicadores de qualidade de coleta: pesagem em dia, "
                   "intervalo útil pro GMD, contexto da pesagem (lote+método), "
                   "execução do trato planejado e cobertura de leitura de chuva. "
                   "Abaixo do mínimo gera alerta — é o que torna a base "
                   "confiável para os modelos preditivos da Trilha 4.")

        hoje = date.today()
        meses = []
        for i in range(2, -1, -1):
            m, a = hoje.month - i, hoje.year
            while m <= 0:
                m += 12; a -= 1
            meses.append((a, m))

        animais_ativos = len(db.get_all_animals(status="ativo"))
        pesagens = normalizar_pesagens(db.get_all_weighings())

        linhas = []
        alertas_ultimo_mes = []
        for ano, mes in meses:
            inicio = date(ano, mes, 1)
            prox_mes = date(ano + (1 if mes == 12 else 0),
                            1 if mes == 12 else mes + 1, 1)
            fim = prox_mes - timedelta(days=1)
            checagens = db.get_feeding_checks(
                start_date=inicio.isoformat(), end_date=fim.isoformat())
            chuvas = db.get_rain(
                start_date=inicio.isoformat(), end_date=fim.isoformat())
            janela = janela_do_mes(ano, mes, checagens_de_trato=checagens,
                                   leituras_de_chuva=chuvas)
            r = avaliar_mes(ano, mes, animais_ativos, pesagens, **janela)
            linhas.append({
                "Mês": f"{mes:02d}/{ano}",
                "Pesagem em dia": round(r["animais_com_pesagem_em_dia"] * 100, 1),
                "Intervalo útil GMD": round(r["intervalos_uteis_gmd"] * 100, 1),
                "Contexto da pesagem": round(r["contexto_da_pesagem"] * 100, 1),
                "Execução nutricional": round(r["execucao_nutricional"] * 100, 1),
                "Cobertura ambiental": round(r["cobertura_ambiental"] * 100, 1),
            })
            alertas_ultimo_mes = r["alertas"]

        df_c = pd.DataFrame(linhas)
        df_melt = df_c.melt(id_vars="Mês", var_name="Indicador",
                            value_name="Completude (%)")
        fig = px.line(df_melt, x="Mês", y="Completude (%)", color="Indicador",
                     markers=True)
        fig.update_layout(**PLOTLY, height=320,
            xaxis=dict(gridcolor=c["superficie"]),
            yaxis=dict(gridcolor=c["superficie"], range=[0, 105]))
        st.plotly_chart(fig, use_container_width=True)

        st.dataframe(df_c, use_container_width=True, hide_index=True,
            column_config={col: st.column_config.NumberColumn(format="%.1f%%")
                           for col in df_c.columns if col != "Mês"})

        mes_label = linhas[-1]["Mês"]
        if alertas_ultimo_mes:
            for al in alertas_ultimo_mes:
                st.warning(f"⚠️ **{al['indicador']}** em "
                          f"{al['valor']*100:.0f}% (mínimo {al['minimo']*100:.0f}%) "
                          f"em {mes_label} — {al['mensagem']}")
        else:
            st.success(f"✅ Todos os indicadores dentro do mínimo em {mes_label}.")

# ══════════════════════════════════════════════════════════════════════════════
# MODO CAMPO  (Mobile-first, máx. 3 cliques)
# ══════════════════════════════════════════════════════════════════════════════
def _campo_trato():
    """Checagem de trato/nutrição por piquete — primeira coisa no Modo Campo."""
    hoje = date.today()
    pend = db.get_pending_feedings(hoje)
    if not pend:
        st.info("Nenhum plano de nutrição definido pelo administrador. "
                "Quando houver, os piquetes aparecerão aqui para checagem.")
        return

    # Agrupa por piquete; só mostra piquetes com plano
    lotes_ids = sorted({p["lote_id"] for p in pend})
    pendentes_total = sum(1 for p in pend if not p["done_this_period"])

    if pendentes_total == 0:
        st.success("✅ Todos os tratos do período já foram confirmados. Bom trabalho!")
    else:
        st.markdown(f"**🌾 Trato do dia — {hoje.strftime('%d/%m/%Y')}** · "
                    f"{pendentes_total} item(ns) pendente(s)")

    for lid in lotes_ids:
        itens = [p for p in pend if p["lote_id"]==lid]
        lote_nome = itens[0].get("lote_name") or lid
        pend_lote = [p for p in itens if not p["done_this_period"]]

        # Cabeçalho do piquete
        st.markdown(f'<div class="card" style="margin-bottom:.4rem">'
                    f'<b style="font-size:1.05rem;color:{c["primaria"]}">🌿 {lid} — {lote_nome}</b>'
                    f'</div>', unsafe_allow_html=True)

        for p in itens:
            freq = db.FEEDING_FREQUENCIES.get(p["frequency"], p["frequency"])
            if p["done_this_period"]:
                st.markdown(
                    f'<div class="hist-item" style="border-left-color:{c["sucesso_escuro"]};opacity:.7">'
                    f'✅ <b>{p["product_name"]}</b> — {p["quantity"]:.0f} {p["unit"]} '
                    f'· {freq} · <span style="color:{c["primaria"]}">confirmado</span> '
                    f'(último: {p["last_check"] or "—"})</div>', unsafe_allow_html=True)
                continue

            with st.form(f"trato_{p['id']}", clear_on_submit=True):
                st.markdown(f'**{p["product_name"]}** — aplicar **{p["quantity"]:.0f} {p["unit"]}** · '
                            f'{freq}')
                fc1, fc2, fc3 = st.columns([2,2,2])
                with fc1:
                    status = st.selectbox("Situação", list(db.FEEDING_CHECK_STATUS.keys()),
                        format_func=lambda s: db.FEEDING_CHECK_STATUS[s], key=f"st_{p['id']}")
                with fc2:
                    qtd_real = st.number_input(f"Qtd aplicada ({p['unit']})",
                        min_value=0.0, value=float(p["quantity"]), step=1.0,
                        key=f"q_{p['id']}")
                with fc3:
                    baixar = st.checkbox("Baixar do estoque",
                        value=bool(p.get("insumo_id")),
                        disabled=not p.get("insumo_id"),
                        help="Disponível se o item estiver vinculado a um insumo",
                        key=f"bx_{p['id']}")
                if st.form_submit_button("✅ Confirmar aplicação", type="primary",
                                         use_container_width=True):
                    db.add_feeding_check(
                        p["id"], lid, hoje.isoformat(), status,
                        actual_quantity=qtd_real,
                        operator=st.session_state.user["name"],
                        deduct_stock=baixar,
                        insumo_id=p.get("insumo_id"),
                        quantity_unit=p["unit"],
                    )
                    st.success(f"✅ {p['product_name']} confirmado para {lote_nome}")
                    st.rerun()


@st.fragment
def _teclado_numerico():
    """Teclado numérico isolado em fragmento: digitar não re-roda a página toda."""
    st.caption("Teclado grande para uso ao sol / com luvas.")
    disp = st.session_state.keypad_value or "——"
    st.markdown(f'<div class="keypad-display">BR {disp}</div>', unsafe_allow_html=True)
    rows_kbd = [["7","8","9"],["4","5","6"],["1","2","3"],["C","0","✓"]]
    for row in rows_kbd:
        kc = st.columns(3)
        for i, k in enumerate(row):
            with kc[i]:
                if st.button(k, key=f"kp_{k}_{row[0]}", use_container_width=True):
                    if k == "C":
                        st.session_state.keypad_value = ""
                        st.rerun(scope="fragment")
                    elif k == "✓":
                        st.session_state.campo_id = f"BR{st.session_state.keypad_value.zfill(4)}"
                        st.session_state.keypad_value = ""
                        st.rerun(scope="app")   # recarrega a página p/ localizar o animal
                    else:
                        if len(st.session_state.keypad_value) < 4:
                            st.session_state.keypad_value += k
                        st.rerun(scope="fragment")


def _tab_pesagem(animal):
    # Comparação com estimativa anterior pendente
    pend = db.get_last_estimate(animal["id"])
    if pend:
        met_lbl = db.WEIGH_METHODS.get(pend.get("method"),"estimativa")
        st.info(f"📋 Última pesagem foi **{met_lbl.lower()}**: "
                f"**{pend['weight']:.1f} kg** em {pend['weigh_date']}. "
                f"Se pesar agora na balança, o app mostra a diferença.")

    # Método fora do form para reagir à escolha
    metodo_peso = st.radio("Método da pesagem",
        list(db.WEIGH_METHODS.keys()),
        format_func=lambda m: db.WEIGH_METHODS[m],
        horizontal=True, key=f"peso_metodo_{animal['id']}")

    nw = float(animal["current_weight"])
    if metodo_peso == "medicao":
        st.caption("Informe as medidas do animal — o peso é estimado pela fórmula "
                   "de Schaeffer (perímetro torácico e comprimento corporal).")
        mm1, mm2 = st.columns(2)
        with mm1:
            pt = st.number_input("Perímetro torácico (cm)", min_value=0.0,
                max_value=350.0, value=180.0, step=1.0, key=f"pt_{animal['id']}")
        with mm2:
            comp = st.number_input("Comprimento corporal (cm)", min_value=0.0,
                max_value=350.0, value=150.0, step=1.0, key=f"comp_{animal['id']}")
        nw = db.estimate_weight_by_measurement(pt, comp)
        st.success(f"⚖️ Peso estimado por medição: **{nw:.1f} kg**")
        medida_nota = f"PT={pt:.0f}cm Comp={comp:.0f}cm"
    else:
        medida_nota = ""

    _pend_alerta = st.session_state.get(f"alerta_peso_{animal['id']}")
    if _pend_alerta:
        st.error("⚠️ **Confira antes de salvar** — o peso informado parece fora do padrão:")
        for _a in _pend_alerta["alertas"]:
            _ic = "🔴" if _a["severidade"] == "alta" else "🟡"
            st.markdown(f"{_ic} {_a['mensagem']}")
        _cc1, _cc2 = st.columns(2)
        if _cc1.button("✅ Está correto, salvar", key=f"okpeso_{animal['id']}",
                       use_container_width=True):
            db.add_weighing(animal["id"], _pend_alerta["peso"], _pend_alerta["data"],
                st.session_state.user["name"], _pend_alerta["notas"],
                method=_pend_alerta["metodo"])
            st.session_state.pop(f"alerta_peso_{animal['id']}", None)
            st.success(f"✅ {_pend_alerta['peso']:.1f} kg salvo."); st.rerun()
        if _cc2.button("↩️ Corrigir", key=f"nopeso_{animal['id']}",
                       use_container_width=True):
            st.session_state.pop(f"alerta_peso_{animal['id']}", None); st.rerun()

    with st.form("f_peso",clear_on_submit=True):
        pc1,pc2=st.columns(2)
        with pc1:
            if metodo_peso == "medicao":
                st.number_input("Peso estimado (kg)", value=float(nw),
                    disabled=True, key=f"pesomed_{animal['id']}")
                nw_final = nw
            else:
                lbl = "Peso (kg) — balança" if metodo_peso=="pesado" else "Peso estimado (kg)"
                nw_final = st.number_input(lbl, min_value=1.0, max_value=2000.0,
                    value=float(animal["current_weight"]), step=0.5, format="%.1f")
        with pc2:
            wd_=st.date_input("Data",value=date.today())
        notes_p=st.text_area("Obs.",height=60,placeholder="Opcional",
            value=medida_nota)
        if st.form_submit_button("✅ Salvar Pesagem",type="primary",use_container_width=True):
            # Confere indícios de erro ANTES de gravar. Não bloqueia: se houver
            # alerta de severidade alta, pede uma confirmação — peso errado
            # contamina GMD, projeção de abate, custo por arroba e ranking.
            _hist = [{"peso": w["weight"], "data": w["weigh_date"]}
                     for w in db.get_weighings(animal["id"])]
            _alertas = avaliar_pesagem(nw_final, wd_.strftime("%Y-%m-%d"), _hist)
            _graves = [a for a in _alertas if a["severidade"] == "alta"]
            _ja_confirmado = st.session_state.pop(f"conf_peso_{animal['id']}", False)

            if _graves and not _ja_confirmado:
                st.session_state[f"alerta_peso_{animal['id']}"] = {
                    "alertas": _alertas, "peso": nw_final,
                    "data": wd_.strftime("%Y-%m-%d"), "notas": notes_p,
                    "metodo": metodo_peso}
                st.rerun()

            db.add_weighing(animal["id"], nw_final, wd_.strftime("%Y-%m-%d"),
                st.session_state.user["name"], notes_p, method=metodo_peso)
            for _a in _alertas:
                st.warning(f"⚠️ {_a['mensagem']}")
            msg = f"✅ {nw_final:.1f} kg salvo ({db.WEIGH_METHODS[metodo_peso]})"
            # Comparação estimativa × pesagem real
            if metodo_peso == "pesado" and pend:
                err = nw_final - pend["weight"]
                pct = (err/pend["weight"]*100) if pend["weight"] else 0
                msg += (f" · Diferença para a estimativa anterior "
                        f"({pend['weight']:.1f} kg): {err:+.1f} kg ({pct:+.1f}%)")
            st.success(msg)
            st.rerun()


def _tab_medicamento(animal):
    ROUTES = ["Subcutânea (SC)","Intramuscular (IM)","Intravenosa (IV)","Oral (PO)","Tópica (Pour-on)","Intramamária"]
    insumos=[i for i in db.get_all_insumos() if i["category"] in ("medicamento","vacina")]
    with st.form("f_med",clear_on_submit=True):
        use_stock=st.toggle("Usar do Estoque",value=bool(insumos))
        if use_stock and insumos:
            ins_sel=st.selectbox("Insumo",insumos,format_func=lambda x:f"{x['name']} ({x['current_stock']:.0f} {x['unit']} em estoque)")
            med_name=ins_sel["name"]; unit_def=ins_sel["unit"]; insumo_id=ins_sel["id"]
        else:
            med_name=st.text_input("Medicamento *",placeholder="Ex: Ivermectina 1%")
            unit_def="ml"; insumo_id=None
            ins_sel=None
        mc1,mc2,mc3=st.columns(3)
        with mc1: dose=st.number_input("Dose",min_value=0.0,step=0.5,format="%.1f")
        with mc2: unit=st.selectbox("Unidade",["ml","mg","g","dose","comprimido"],
                        index=["ml","mg","g","dose","comprimido"].index(unit_def) if unit_def in ["ml","mg","g","dose","comprimido"] else 0)
        with mc3: route=st.selectbox("Via",ROUTES)
        wd_c=st.number_input("Carência (dias)",min_value=0,max_value=180,value=0,step=1)
        md_=st.date_input("Data Aplicação",value=date.today())
        notes_m=st.text_area("Obs.",height=60,placeholder="Opcional")
        if st.form_submit_button("✅ Salvar Medicamento",type="primary",use_container_width=True):
            if not med_name:
                st.error("Informe o medicamento.")
            else:
                db.add_medication(animal["id"],med_name,dose,unit,route,
                    int(wd_c),md_.strftime("%Y-%m-%d"),
                    st.session_state.user["name"],insumo_id,notes_m)
                st.success(f"✅ {med_name} registrado!" + (f" Carência: {wd_c} dias" if wd_c else ""))
                st.rerun()


def _tab_movimentacao(animal):
    lotes=db.get_all_lotes()
    with st.form("f_mov",clear_on_submit=True):
        dest=st.selectbox("Destino (Lote)",lotes,
            format_func=lambda x:f"{x['id']} — {x['name']} ({_plural(x['animal_count'],'animal','animais')} | {x['area_ha']} ha)")
        mv_date=st.date_input("Data",value=date.today())
        reason=st.selectbox("Motivo",["manejo","pesagem","tratamento","separação","venda","óbito"])
        notes_mv=st.text_area("Obs.",height=60,placeholder="Opcional")
        if st.form_submit_button("✅ Mover Animal",type="primary",use_container_width=True):
            if dest:
                db.move_animal(animal["id"],dest["id"],mv_date.strftime("%Y-%m-%d"),
                    reason,st.session_state.user["name"],notes_mv)
                st.success(f"✅ {animal['id']} movido para {dest['name']}")
                st.rerun()


def _tab_obito(animal):
    if animal["status"] == "morto":
        st.info("Este animal já está registrado como morto.")
    else:
        st.warning("Registrar óbito é **irreversível** e muda o status do animal para 'morto'.")
        with st.form("f_obito", clear_on_submit=True):
            causa = st.selectbox("Causa do óbito *", db.DEATH_CAUSES)
            od1, od2 = st.columns(2)
            with od1:
                obito_data = st.date_input("Data do óbito", value=date.today())
            with od2:
                st.metric("Perda estimada", f"R$ {db.get_total_cost(animal['id']):,.2f}",
                          help="Custo investido no animal até agora")
            obs_ob = st.text_area("Observações", height=60,
                placeholder="Ex: encontrado no piquete norte, suspeita de cobra")
            confirmar = st.checkbox("Confirmo o registro do óbito deste animal")
            if st.form_submit_button("☠️ Registrar Óbito", type="primary", use_container_width=True):
                if not confirmar:
                    st.error("Marque a confirmação para registrar.")
                else:
                    r = db.register_death(animal["id"], obito_data.strftime("%Y-%m-%d"),
                        causa, operator=st.session_state.user["name"], notes=obs_ob)
                    st.success(f"Óbito registrado. Perda contabilizada: R$ {r.get('perda',0):,.2f}")
                    st.rerun()


def _tab_historico(animal):
    c = cores()
    h1,h2=st.columns(2)
    with h1:
        st.markdown("**⚖️ Pesagens**")
        ws=db.get_weighings(animal["id"])
        if len(ws)>=2:
            df_hw=pd.DataFrame(ws)[["weigh_date","weight"]].sort_values("weigh_date")
            df_hw.columns=["Data","Peso (kg)"]; df_hw["Data"]=pd.to_datetime(df_hw["Data"])
            fig_hw=px.line(df_hw,x="Data",y="Peso (kg)",markers=True,
                color_discrete_sequence=[c["primaria"]])
            fig_hw.update_layout(**PLOTLY,height=150,xaxis=dict(gridcolor=c["superficie"]),
                yaxis=dict(gridcolor=c["superficie"]))
            st.plotly_chart(fig_hw,use_container_width=True)
        for w in ws[:5]:
            met = w.get("method") or "pesado"
            mbadge = {"pesado":'<span class="badge-green">balança</span>',
                      "estimado":'<span class="badge-yellow">estimado</span>',
                      "medicao":'<span class="badge-blue">medição</span>'}.get(met,"")
            st.markdown(f'<div class="hist-item"><b>{w["weight"]:.1f} kg</b> {mbadge}'
                f'<span style="color:{c["texto_terciario"]};font-size:.8rem;float:right">{w["weigh_date"]}</span><br>'
                f'<span style="color:{c["texto_secundario"]};font-size:.78rem">{w["operator"] or "—"}</span></div>',
                unsafe_allow_html=True)
    with h2:
        st.markdown("**💉 Medicamentos**")
        for m in db.get_medications(animal["id"], limit=5):
            end_=datetime.strptime(m["med_date"],"%Y-%m-%d").date()+timedelta(days=m["withdrawal_days"] or 0)
            badge='<span class="badge-yellow">Carência</span>' if m["withdrawal_days"] and end_>=date.today() else ""
            st.markdown(f'<div class="hist-item" style="border-left-color:{c["info"]}">'
                f'<b>{html.escape(str(m["medication_name"]))}</b> {badge}<br>'
                f'<span style="color:{c["texto_terciario"]};font-size:.78rem">'
                f'{_fmt_dose(m["dose"], m["unit"])} · {m["application_route"]} · {m["med_date"]}'
                f'{"  ·  carência "+str(m["withdrawal_days"])+"d" if m["withdrawal_days"] else ""}'
                f'</span></div>',unsafe_allow_html=True)
        st.markdown("**🚚 Movimentações**")
        for mv in db.get_movements(animal["id"], limit=4):
            st.markdown(f'<div class="hist-item" style="border-left-color:{c["destaque"]}">'
                f'<b>{mv.get("from_name") or "—"} → {mv.get("to_name","?")}</b><br>'
                f'<span style="color:{c["texto_terciario"]};font-size:.78rem">{mv["movement_date"]} · {mv["reason"]}</span>'
                f'</div>',unsafe_allow_html=True)


def _campo_animal():
    # ── Passo 1: Localizar animal ─────────────────────────────────────────────
    tab_dig, tab_cam, tab_kbd = st.tabs(["⌨️ Digitar ID","📷 Câmera (brinco)","🔢 Teclado Numérico"])

    with tab_dig:
        c1,c2=st.columns([3,1])
        with c1:
            typed=st.text_input("🏷️ ID do Animal",value=st.session_state.campo_id,
                placeholder="Ex: BR0001",key="campo_text").strip().upper()
        with c2:
            st.markdown("<br>",unsafe_allow_html=True)
            if st.button("🔍 Buscar",type="primary",use_container_width=True):
                st.session_state.campo_id=typed; st.rerun()

    with tab_cam:
        st.caption("Aponte a câmera para o brinco. Se tiver **QR Code**, é lido automaticamente. "
                   "Senão, o app tenta ler o **número** — confira sempre antes de buscar.")
        if not st.session_state.get("cam_brinco_on"):
            if st.button("📷 Abrir câmera para ler o brinco", type="primary",
                         use_container_width=True, key="brinco_open"):
                st.session_state["cam_brinco_on"] = True; st.rerun()
        else:
            foto = st.camera_input("Foto do brinco", key="cam_brinco")
            if foto is not None:
                raw = foto.getvalue()
                qr = _decode_qr(raw)
                if qr:
                    cand = qr.strip().upper()
                    st.success(f"✅ QR Code lido: **{cand}**")
                    if st.button(f"🔍 Buscar {cand}", type="primary", use_container_width=True):
                        st.session_state.campo_id = cand
                        st.session_state["cam_brinco_on"] = False; st.rerun()
                else:
                    ocr = _ocr_number(raw)
                    sug = f"BR{ocr.zfill(4)}" if ocr else ""
                    st.info("QR não encontrado. Leitura automática do número (confira!):")
                    lido = st.text_input("Número lido / digite o ID", value=sug,
                        key="cam_ocr_id").strip().upper()
                    if st.button("🔍 Buscar", type="primary", use_container_width=True, key="cam_busca"):
                        if lido:
                            st.session_state.campo_id = lido
                            st.session_state["cam_brinco_on"] = False; st.rerun()
            if st.button("✖️ Fechar câmera", use_container_width=True, key="brinco_close"):
                st.session_state["cam_brinco_on"] = False; st.rerun()

    with tab_kbd:
        _teclado_numerico()

    # ── Passo 2: Exibir animal ────────────────────────────────────────────────
    eid=st.session_state.campo_id
    if not eid: st.info("Selecione ou busque um animal para começar."); return

    animal=db.get_animal(eid)
    if not animal:
        st.error(f"Animal **{eid}** não encontrado."); return

    gmd=db.calculate_gmd(animal["id"])
    wd =db.get_withdrawal_end(animal["id"])
    gc =c["primaria"] if (gmd and gmd>0) else c["perigo"] if (gmd and gmd<0) else c["texto_secundario"]
    cat=db.get_age_category(animal.get("birth_date"))
    idade=db.get_age_display(animal)

    carencia_html = (f'<div style="color:{c["atencao"]};font-size:.82rem;margin-top:.3rem">'
                     f'⚠️ Carência até {wd.isoformat()}</div>') if wd else ''
    sex_sym = "♂" if animal['sex']=='M' else "♀"
    gmd_txt = f'{gmd:+.3f} kg/dia' if gmd is not None else '— sem dados'
    st.markdown(
        f'<div class="card-green">'
        f'<div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:.5rem">'
        f'<div>'
        f'<div style="font-size:2rem;font-weight:900;color:{c["primaria"]};line-height:1.1">🐄 {animal["id"]}</div>'
        f'<div style="color:{c["texto_secundario"]};font-size:.92rem;margin-top:.2rem">'
        f'{animal["breed"]} · {sex_sym} · {cat} ({idade})<br>'
        f'Lote: <b style="color:{c["texto"]}">{animal.get("lote_name") or "—"}</b> &nbsp;'
        f'{_status_badge(animal["status"])}'
        f'</div>{carencia_html}</div>'
        f'<div style="text-align:right">'
        f'<div style="font-size:2.4rem;font-weight:900;color:{c["texto"]};line-height:1">'
        f'{animal["current_weight"]:.1f}<span style="font-size:1rem;color:{c["texto_terciario"]}"> kg</span></div>'
        f'<div style="color:{c["atencao"]};font-size:.85rem">{_fmt_live(animal["current_weight"])}</div>'
        f'<div style="color:{gc};font-size:.88rem;font-weight:600">GMD: {gmd_txt}</div>'
        f'</div></div></div>',
        unsafe_allow_html=True)

    # ── Passo 3: Ação ─────────────────────────────────────────────────────────
    t1,t2,t3,t6,t5,t4=st.tabs(["⚖️ Pesagem","💉 Medicamento","🚚 Movimentação",
                               "📷 Foto","☠️ Óbito","📜 Histórico"])

    with t6:  # FOTO
        _photo_section(animal["id"], key_prefix="campo_")

    with t5:  # ÓBITO
        _tab_obito(animal)

    with t1:  # PESAGEM
        _tab_pesagem(animal)

    with t2:  # MEDICAMENTO
        _tab_medicamento(animal)

    with t3:  # MOVIMENTAÇÃO
        _tab_movimentacao(animal)

    with t4:  # HISTÓRICO
        _tab_historico(animal)


def _campo_importar():
    """Importa pesagens de um CSV exportado pelo indicador da balança."""
    st.subheader("📥 Importar pesagens de CSV")
    st.caption("Três colunas: **brinco, peso, data**. Aceita `;` ou `,` como separador, "
               "data em `AAAA-MM-DD` ou `DD/MM/AAAA`, e peso com vírgula decimal. "
               "Cabeçalho é opcional.")

    arquivo = st.file_uploader("Arquivo do indicador", type=["csv", "txt"])
    if arquivo is None:
        return

    bruto = arquivo.getvalue()
    try:
        texto = bruto.decode("utf-8-sig")
    except UnicodeDecodeError:
        # Indicadores de balança costumam exportar em cp1252/latin-1.
        texto = bruto.decode("latin-1")

    ativos = {a["id"] for a in db.get_all_animals(status="ativo")}
    resultado = db.parse_pesagens(texto, ids_conhecidos=ativos)
    aceitas, rejeitadas = resultado["aceitas"], resultado["rejeitadas"]

    m1, m2, m3 = st.columns(3)
    m1.metric("Linhas lidas", resultado["total_linhas"])
    m2.metric("Aceitas", len(aceitas))
    m3.metric("Rejeitadas", len(rejeitadas))

    if rejeitadas:
        st.error(f"{len(rejeitadas)} linha(s) não serão importadas:")
        st.dataframe(pd.DataFrame([
            {"Linha": r["linha"], "Motivo": r["motivo"], "Conteúdo": r["conteudo"]}
            for r in rejeitadas
        ]), use_container_width=True, hide_index=True)

    if not aceitas:
        st.info("Nada a importar.")
        return

    # Qualidade: o histórico acumula as próprias linhas do arquivo, senão duas
    # pesagens do mesmo animal no mesmo CSV não se enxergariam.
    animal_ids = {linha["animal_id"] for linha in aceitas}
    all_weighings = db.get_weighings_batch(animal_ids)
    hist = {}
    for a_id in animal_ids:
        hist[a_id] = [{"peso": w["weight"], "data": w["weigh_date"]}
                      for w in all_weighings.get(a_id, [])]

    previa, graves = [], 0
    for linha in sorted(aceitas, key=lambda x: (x["animal_id"], x["data"])):
        alertas = avaliar_pesagem(linha["peso"], linha["data"], hist[linha["animal_id"]])
        altos = [a["mensagem"] for a in alertas if a["severidade"] == "alta"]
        graves += 1 if altos else 0
        previa.append({"Animal": linha["animal_id"], "Peso (kg)": linha["peso"],
                       "Data": linha["data"],
                       "⚠️": " · ".join(altos) if altos else ""})
        hist[linha["animal_id"]].insert(
            0, {"peso": linha["peso"], "data": linha["data"]})

    st.markdown("**Prévia do que será gravado**")
    st.dataframe(pd.DataFrame(previa), use_container_width=True, hide_index=True)

    if graves:
        st.warning(f"⚠️ {graves} pesagem(ns) com indício de erro — confira antes de gravar. "
                   "Peso errado contamina GMD, projeção de abate e custo por arroba.")

    if st.button(f"💾 Gravar {len(aceitas)} pesagem(ns)", type="primary"):
        gravadas = 0
        for linha in aceitas:
            db.add_weighing(linha["animal_id"], linha["peso"], linha["data"],
                            operator=st.session_state.user["name"],
                            notes=f"importado de {arquivo.name}")
            gravadas += 1
        st.success(f"✅ {gravadas} pesagem(ns) importada(s).")
        st.rerun()


def page_campo():
    st.markdown('<div class="page-title">📱 Modo Campo</div>', unsafe_allow_html=True)
    # Badge com nº de tratos pendentes na aba
    pend = db.get_pending_feedings()
    n_pend = sum(1 for p in pend if not p["done_this_period"])
    trato_label = f"🌾 Trato do Dia{' 🔴'+str(n_pend) if n_pend else ''}"
    tab_trato, tab_animal, tab_import = st.tabs(
        [trato_label, "🐄 Manejo do Animal", "📥 Importar CSV"])
    with tab_trato:
        _campo_trato()
    with tab_animal:
        _campo_animal()
    with tab_import:
        _campo_importar()

# ══════════════════════════════════════════════════════════════════════════════
# REBANHO
# ══════════════════════════════════════════════════════════════════════════════
def page_rebanho():
    st.markdown('<div class="page-title">📋 Rebanho</div>', unsafe_allow_html=True)
    animals_all=db.get_all_animals(status=None)
    if not animals_all:
        st.info("Nenhum animal cadastrado."); return

    f1,f2,f3,f4,f5=st.columns([2,1,1,1,1])
    with f1: busca=st.text_input("🔍 ID / Raça / Lote",placeholder="Ex: BR0003").upper()
    with f2:
        races=["Todas"]+sorted({a["breed"] for a in animals_all})
        fr=st.selectbox("Raça",races)
    with f3:
        fcat=st.selectbox("Categoria",["Todas"]+AGE_BANDS)
    with f4:
        statuses=["Todos","ativo","vendido","morto","carencia"]
        fs=st.selectbox("Status",statuses)
    with f5:
        lotes_opts=["Todos"]+sorted({a.get("lote_id","") or "—" for a in animals_all})
        fl=st.selectbox("Lote",lotes_opts)

    ul = _unit_label()
    rows=[]
    a_ids_all = [a["id"] for a in animals_all]
    gmd_batch = db.calculate_gmd_bulk(a_ids_all)
    wd_batch = db.get_withdrawal_end_batch(a_ids_all)
    for a in animals_all:
        gmd=gmd_batch.get(a["id"])
        wd =wd_batch.get(a["id"])
        rows.append({"ID":a["id"],"Raça":a["breed"],"Sexo":"♂" if a["sex"]=="M" else "♀",
            "Categoria":db.get_age_category(a.get("birth_date")),
            "Idade":db.get_age_display(a),
            "Lote":a.get("lote_id") or "—","Status":a["status"],
            "Peso Atual (kg)":a["current_weight"],
            f"Ganho ({ul})":_prod_weight(a["current_weight"]-a["entry_weight"]),
            "GMD (kg/dia)":gmd,
            "Carência até":wd.isoformat() if wd else "—",
            "Fornecedor":a.get("fornecedor_name") or "—"})
    df=pd.DataFrame(rows)

    if busca:
        b = busca.lower()
        df = df[
            df["ID"].str.contains(busca, na=False, regex=False) |
            df["Raça"].str.lower().str.contains(b, na=False, regex=False) |
            df["Lote"].str.contains(busca, na=False, regex=False)
        ]
    if fr!="Todas": df=df[df["Raça"]==fr]
    if fcat!="Todas": df=df[df["Categoria"]==fcat]
    if fs!="Todos": df=df[df["Status"]==fs]
    if fl!="Todos": df=df[df["Lote"]==fl]

    st.markdown(f"**{len(df)}** registro(s)")
    fmt_gain = "%.2f" if _use_arroba() else "%.1f"
    st.dataframe(df,use_container_width=True,hide_index=True,height=460,
        column_config={"Peso Atual (kg)":st.column_config.NumberColumn(format="%.1f"),
            f"Ganho ({ul})":st.column_config.NumberColumn(format=fmt_gain),
            "GMD (kg/dia)":st.column_config.NumberColumn(format="%.3f")})

    st.markdown("---")
    r1,r2=st.columns([2,1])
    with r1:
        sel=st.selectbox("Animal para detalhar",[a["id"] for a in animals_all])
    with r2:
        st.markdown("<br>",unsafe_allow_html=True)
        if st.button("📂 Abrir Ficha",type="primary",use_container_width=True):
            _go("animal",sel); st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# FICHA DO ANIMAL  (Linha do Tempo)
# ══════════════════════════════════════════════════════════════════════════════
_ICONE_GRAVIDADE = {"bloqueio": "🔴", "alerta": "🟡", "informativo": "🔵"}

# Validações que hoje disparam em TODO animal porque cobram campo que o schema
# ainda não tem: `propriedade_id` chega na etapa B4 e genealogia na B3 do ADR
# 0004. Não são erro de cadastro — são funcionalidade ausente, e mostrá-las em
# cada ficha ensina o usuário a ignorar o aviso inteiro. Voltam quando os campos
# existirem e o pecuarista puder de fato preenchê-los.
_CODIGOS_AGUARDANDO_SCHEMA = {"animal_sem_origem", "nascimento_sem_mae"}


def _consistencia_regulatoria(animal: dict, movs: list):
    """Problemas de consistência do cadastro (PNIB §17.3).

    A checagem é `services/validacao_regulatoria.py`, função pura: tudo o que ela
    precisa entra por parâmetro. Chave ausente no contexto faz a validação
    correspondente ser **pulada**, não falhar — por isso genealogia (que só chega
    na etapa B3) simplesmente não é avaliada hoje.
    """
    uuid = animal.get("uuid")
    ids = db.identificadores.get_identificadores(uuid) if uuid else []

    morte = None
    for d in db.get_deaths():
        if d.get("animal_id") == animal["id"]:
            morte = d.get("death_date")
            break

    eventos = [{"tipo": "movimentacao", "data": m.get("movement_date"),
                "propriedade_id": None} for m in movs]
    if morte:
        eventos.append({"tipo": "morte", "data": morte, "propriedade_id": None})

    problemas = validar_animal(
        {"id": animal["id"],
         "sexo": animal.get("sex"),
         "nascimento": animal.get("birth_date"),
         "nascimento_estimado": bool(animal.get("birth_estimated")),
         "morte": morte},
        {"eventos": eventos,
         "identificadores": [{"tipo": i["tipo"], "valor": i["valor"],
                              "ativo": i["status"] == "ativo"} for i in ids],
         "hoje": date.today().isoformat()},
    )
    problemas = [p for p in problemas
                 if p["codigo"] not in _CODIGOS_AGUARDANDO_SCHEMA]
    if not problemas:
        return

    graves = [p for p in problemas if p["gravidade"] == "bloqueio"]
    rotulo = f"⚠️ {len(problemas)} inconsistência(s) no cadastro"
    with st.expander(rotulo, expanded=bool(graves)):
        for p in problemas:
            icone = _ICONE_GRAVIDADE.get(p["gravidade"], "•")
            st.markdown(f"{icone} **{p['gravidade'].capitalize()}** — {p['mensagem']}")
        st.caption("Regras do §17.3 do PNIB. Genealogia e propriedade de origem ainda "
                   "não são avaliadas: dependem das etapas B3 e B4.")


def _identificadores_do_animal(animal: dict):
    """Brincos e demais identificadores, com histórico (PNIB §4.1 e §4.2).

    Trocar brinco **não** apaga o anterior: encerra a vigência e abre outra. É o
    que permite reconstruir qual brinco o animal usava em cada data (§4.2.3), e é
    a razão de a identidade ter passado para o uuid no ADR 0004.
    """
    uuid = animal.get("uuid")
    if not uuid:
        st.warning("Este animal ainda não tem identificador interno (uuid). "
                   "Rode a migração do ADR 0004 antes de gerenciar brincos.")
        return

    itens = db.identificadores.get_identificadores(uuid)
    ativos = [i for i in itens if i["status"] == "ativo"]

    if itens:
        st.dataframe(pd.DataFrame([{
            "Tipo": i["tipo"], "Valor": i["valor"],
            "Situação": "✅ vigente" if i["status"] == "ativo" else "histórico",
            "Aplicado em": i.get("aplicado_em") or "—",
            "Removido em": i.get("removido_em") or "—",
            "Motivo da remoção": i.get("motivo_remocao") or "—",
        } for i in itens]), use_container_width=True, hide_index=True)
    else:
        st.info("Nenhum identificador registrado para este animal.")

    st.markdown("---")
    c_novo, c_troca = st.columns(2)

    with c_novo:
        st.markdown("**Aplicar identificador**")
        tipo = st.selectbox("Tipo", db.identificadores.TIPOS, key=f"id_tipo_{uuid}")
        valor = st.text_input("Valor", key=f"id_valor_{uuid}").strip()

        erros = []
        if valor:
            erros = validar_formato_id(valor, REGRAS_PADRAO.get(tipo, {})).get("erros", [])
            for e in erros:
                st.error(f"🚫 {e}")
            if any(i["tipo"] == tipo for i in ativos):
                st.warning(f"Já existe um `{tipo}` vigente. Use a troca ao lado para "
                           "preservar o histórico.")

        if st.button("➕ Aplicar", disabled=not valor or bool(erros),
                     key=f"id_aplicar_{uuid}"):
            r = db.identificadores.aplicar(
                uuid, tipo, valor, aplicado_por=st.session_state.user["name"])
            if r.get("ok"):
                st.success("Já estava aplicado." if r.get("ja_existia")
                           else f"✅ {tipo} {valor} aplicado.")
                st.rerun()
            else:
                st.error(f"🚫 {r.get('erro')}")

    with c_troca:
        st.markdown("**Trocar / remover**")
        if not ativos:
            st.caption("Nenhum identificador vigente para trocar.")
            return
        # Opções são texto, não dicionário: opção-dicionário quebra quando a lista
        # é reconstruída no rerun, porque o valor guardado é outro objeto.
        rotulos = {f"{i['tipo']} — {i['valor']}": i for i in ativos}
        alvo = rotulos[st.selectbox("Vigente", list(rotulos), key=f"id_alvo_{uuid}")]
        motivo = st.text_input("Motivo (obrigatório)", key=f"id_motivo_{uuid}").strip()
        novo = st.text_input("Novo valor (deixe vazio para só remover)",
                             key=f"id_novo_{uuid}").strip()

        erros_novo = []
        if novo:
            erros_novo = validar_formato_id(
                novo, REGRAS_PADRAO.get(alvo["tipo"], {})).get("erros", [])
            for e in erros_novo:
                st.error(f"🚫 {e}")

        rotulo = "🔄 Substituir" if novo else "🗑️ Encerrar vigência"
        if st.button(rotulo, disabled=not motivo or bool(erros_novo),
                     key=f"id_trocar_{uuid}"):
            if novo:
                r = db.identificadores.substituir(
                    uuid, alvo["tipo"], novo, motivo,
                    aplicado_por=st.session_state.user["name"])
                if not r.get("ok"):
                    st.error(f"🚫 {r.get('erro')}"); return
            else:
                db.identificadores.remover(uuid, alvo["tipo"], motivo)
            st.success("✅ Registrado — o valor anterior fica no histórico.")
            st.rerun()


_TIPO_EVENTO_ROTULO = {
    "nascimento": "Nascimento", "cadastro_inicial": "Cadastro inicial",
    "identificacao_interna": "Identificação interna",
    "identificacao_oficial": "Identificação oficial",
    "aplicacao_dispositivo": "Aplicação de dispositivo",
    "leitura_conferencia": "Leitura / conferência", "perda_brinco": "Perda de brinco",
    "dano_dispositivo": "Dano em dispositivo", "substituicao": "Substituição",
    "retirada_autorizada": "Retirada autorizada",
    "entrada_propriedade": "Entrada na propriedade",
    "saida_propriedade": "Saída da propriedade", "venda": "Venda", "compra": "Compra",
    "transferencia_sem_titularidade": "Transferência (sem titularidade)",
    "mudanca_titularidade": "Mudança de titularidade",
    "emissao_gta": "Emissão de GTA", "cancelamento_gta": "Cancelamento de GTA",
    "chegada_confirmada": "Chegada confirmada", "recusa_recepcao": "Recusa de recepção",
    "manejo_sanitario": "Manejo sanitário", "vacinacao": "Vacinação",
    "vacinacao_brucelose": "Vacinação (brucelose)", "teste_sanitario": "Teste sanitário",
    "tratamento": "Tratamento", "pesagem": "Pesagem", "mudanca_lote": "Mudança de lote",
    "mudanca_categoria": "Mudança de categoria", "morte": "Morte", "abate": "Abate",
    "correcao": "Correção", "estorno": "Estorno",
}


def _linha_do_tempo_do_animal(animal: dict):
    """Histórico completo de eventos do animal (PNIB §6).

    Não se confunde com as abas de peso/sanidade/movimentação acima: aquelas
    mostram cada operação sob a ótica de quem opera a fazenda. Esta mostra a
    trilha regulatória — o que foi *registrado como fato*, append-only, com a
    diferença entre quando aconteceu e quando foi lançado (§6.2).

    Duas decisões que o §6 impõe e que o repositório não pode tomar sozinho:

    - **Não existe editar.** Só "registrar correção", que cria outro evento
      apontando para o original (§6.3). O original nunca some da tela.
    - **A diferença entre `ocorrido_em` e `registrado_em` é auditável**, então
      aparece sempre — inclusive quando é zero, porque zero também é informação.
    """
    uuid = animal.get("uuid")
    if not uuid:
        st.warning("Este animal ainda não tem identificador interno (uuid).")
        return

    filtro = st.selectbox(
        "Filtrar por tipo", ["Todos"] + list(db.eventos.TIPOS),
        format_func=lambda t: "Todos" if t == "Todos"
        else _TIPO_EVENTO_ROTULO.get(t, t),
        key=f"ev_filtro_{uuid}")

    eventos = db.eventos.do_animal(uuid, tipo=None if filtro == "Todos" else filtro)
    if not eventos:
        st.info("Nenhum evento registrado" +
                ("." if filtro == "Todos" else f" do tipo '{filtro}'."))
        return

    # Quem corrige quem: uma correção aponta pra trás via evento_anterior_id.
    # Sem isto, o original pareceria a última palavra quando não é.
    correcoes_de = {}
    for e in eventos:
        alvo = e.get("evento_anterior_id")
        if alvo:
            correcoes_de.setdefault(alvo, []).append(e)

    for e in eventos:
        _cartao_de_evento(e, correcoes_de.get(e["id"], []))


def _cartao_de_evento(e: dict, correcoes: list[dict]):
    ocorrido = (e.get("ocorrido_em") or "")[:16].replace("T", " ")
    registrado = (e.get("registrado_em") or "")[:16].replace("T", " ")
    atraso = ""
    try:
        d1 = datetime.fromisoformat(e["ocorrido_em"])
        d2 = datetime.fromisoformat(e["registrado_em"])
        delta = d2 - d1
        if delta.total_seconds() > 3600:
            dias = delta.days
            atraso = (f" · registrado {dias} dia(s) depois" if dias >= 1
                      else f" · registrado {int(delta.total_seconds()//3600)}h depois")
    except (ValueError, TypeError, KeyError):
        pass

    cor = c["atencao"] if e["tipo"] in ("correcao", "estorno") else c["primaria"]
    rotulo = _TIPO_EVENTO_ROTULO.get(e["tipo"], e["tipo"])

    with st.container():
        st.markdown(
            f'<div class="hist-item" style="border-left-color:{cor}">'
            f'<b>{rotulo}</b> {"🔺 corrigido depois" if correcoes else ""}<br>'
            f'<span style="color:{c["texto_secundario"]};font-size:.82rem">'
            f'ocorreu: {ocorrido} · registrado: {registrado}{atraso}'
            f'{"  ·  por: " + e["usuario_registro"] if e.get("usuario_registro") else ""}'
            f'</span>'
            + (f'<br><span style="font-size:.85rem">{e["observacoes"]}</span>'
               if e.get("observacoes") else "")
            + (f'<br><span style="color:{c["atencao"]};font-size:.82rem">'
               f'justificativa: {e["justificativa"]}</span>'
               if e.get("justificativa") else "")
            + '</div>', unsafe_allow_html=True)

        for corr in correcoes:
            st.caption(f"↳ corrigido por evento #{corr['id']} ({corr['ocorrido_em'][:10]}): "
                       f"{corr.get('justificativa', '')}")

        if e["tipo"] not in ("correcao", "estorno"):
            with st.expander(f"🖊️ Registrar correção do evento #{e['id']}"):
                st.caption("§6.3: isto NÃO altera o evento acima — ele permanece "
                           "como está. Cria um evento novo, apontando para este, "
                           "com o que deveria ter sido registrado.")
                tipo_corr = st.selectbox(
                    "Tipo", ["correcao", "estorno"], key=f"corr_tipo_{e['id']}",
                    format_func=lambda t: "Correção (ajusta o registro)" if t == "correcao"
                    else "Estorno (desfaz o efeito)")
                justificativa = st.text_input(
                    "Justificativa *", key=f"corr_just_{e['id']}").strip()
                if st.button("Registrar", disabled=not justificativa,
                             key=f"corr_salvar_{e['id']}"):
                    r = db.eventos.corrigir(
                        e["id"], justificativa,
                        usuario_registro=st.session_state.user["name"],
                        tipo=tipo_corr)
                    if r.get("ok"):
                        db.clear_cache()
                        st.success("✅ Correção registrada.")
                        st.rerun()
                    else:
                        st.error(f"🚫 {r.get('erro')}")


def _render_tab_peso(ws):
    if len(ws)>=2:
        import numpy as np
        df_w=pd.DataFrame(ws)[["weigh_date","weight"]].sort_values("weigh_date")
        df_w.columns=["Data","Peso (kg)"]; df_w["Data"]=pd.to_datetime(df_w["Data"])
        x_num=(df_w["Data"]-df_w["Data"].min()).dt.days
        coef=np.polyfit(x_num,df_w["Peso (kg)"],1)
        fig=go.Figure()
        fig.add_trace(go.Scatter(x=df_w["Data"],y=np.polyval(coef,x_num),
            mode="lines",name="Tendência",line=dict(dash="dot",color=c["atencao"],width=2),hoverinfo="skip"))
        fig.add_trace(go.Scatter(x=df_w["Data"],y=df_w["Peso (kg)"],
            mode="lines+markers",name="Pesagens",
            line=dict(color=c["primaria"],width=2.5),
            marker=dict(size=10,color=c["primaria"],line=dict(width=2,color=c["fundo"])),
            hovertemplate="%{x|%d/%m/%Y}<br><b>%{y:.1f} kg</b><extra></extra>"))
        fig.update_layout(**PLOTLY,height=300,
            xaxis=dict(gridcolor=c["superficie"],title="Data"),
            yaxis=dict(gridcolor=c["superficie"],title="Peso (kg)"),
            legend=dict(orientation="h",y=1.08))
        st.plotly_chart(fig,use_container_width=True)
    else:
        st.info("São necessárias ao menos 2 pesagens para exibir o gráfico.")
    st.subheader("Tabela de Pesagens")
    if ws:
        df_wt=pd.DataFrame(ws)[["weigh_date","weight","method","lote_id","operator","notes"]].copy()
        df_wt["method"]=df_wt["method"].fillna("pesado").map(
            lambda m: db.WEIGH_METHODS.get(m,m))
        df_wt.columns=["Data","Peso (kg)","Método","Lote","Operador","Obs"]
        st.dataframe(df_wt,use_container_width=True,hide_index=True,
            column_config={"Peso (kg)":st.column_config.NumberColumn(format="%.1f")})

def _render_tab_med(meds):
    if meds:
        for m_ in meds:
            end_=datetime.strptime(m_["med_date"],"%Y-%m-%d").date()+timedelta(days=m_["withdrawal_days"] or 0)
            active=m_["withdrawal_days"] and end_>=date.today()
            bc=f"border-left-color:{c['perigo']}" if active else f"border-left-color:{c['info']}"
            st.markdown(f'<div class="hist-item" style="{bc}">'
                f'<b style="font-size:1rem">{html.escape(str(m_["medication_name"]))}</b>'
                f'{"  "+_gmd_badge(None).replace("badge-gray","badge-yellow").replace("N/D","Carência ativa") if active else ""}<br>'
                f'<span style="color:{c["texto_secundario"]};font-size:.82rem">'
                f'{m_["med_date"]} · {_fmt_dose(m_["dose"], m_["unit"])} · {m_["application_route"]}'
                f'{"  ·  carência "+str(m_["withdrawal_days"])+" dias (até "+end_.isoformat()+")" if m_["withdrawal_days"] else ""}'
                f'{"  ·  por: "+m_["applied_by"] if m_["applied_by"] else ""}'
                f'</span></div>',unsafe_allow_html=True)
    else:
        st.info("Nenhum medicamento registrado.")

def _render_tab_mov(movs):
    if movs:
        for mv in movs:
            st.markdown(f'<div class="hist-item" style="border-left-color:{c["destaque"]}">'
                f'<b>{mv.get("from_name") or "Entrada"} → {mv.get("to_name","?")}</b><br>'
                f'<span style="color:{c["texto_secundario"]};font-size:.82rem">'
                f'{mv["movement_date"]} · {mv["reason"]} · {mv.get("operator") or "—"}'
                f'</span></div>',unsafe_allow_html=True)
    else:
        st.info("Nenhuma movimentação registrada.")

def _render_tab_fin(aid, animal, gain, yield_, cost_total, ul):
    costs=db.get_animal_costs(aid)
    prod_gain   = _prod_weight(gain, yield_) if gain > 0 else 0
    cpu_val     = _cost_per_unit(cost_total, animal["current_weight"], yield_)
    cpu_gain = round(cost_total / prod_gain, 2) if prod_gain > 0 else 0
    cp1,cp2,cp3 = st.columns(3)
    cp1.metric("Custo Total",          f"R$ {cost_total:,.2f}",
               help="Compra + insumos + custeio operacional")
    cp2.metric(_cost_per_unit_label(), f"R$ {cpu_val:,.2f}" if cpu_val else "—",
               help="Custo total ÷ peso vivo atual")
    cp3.metric(f"Custo da Produção (R$/{ul})",
               f"R$ {cpu_gain:,.2f}" if cpu_gain else "—",
               help=f"Custo total ÷ {ul} ganhos (produzidos) desde a entrada")
    if costs:
        df_c=pd.DataFrame(costs)[["cost_date","cost_type","description","amount"]]
        df_c.columns=["Data","Tipo","Descrição","Valor (R$)"]
        st.dataframe(df_c,use_container_width=True,hide_index=True,
            column_config={"Valor (R$)":st.column_config.NumberColumn(format="R$ %.2f")})
    # Adicionar custo
    with st.expander("➕ Adicionar Custo"):
        with st.form("f_cost",clear_on_submit=True):
            cc1,cc2,cc3=st.columns(3)
            with cc1: ct=st.selectbox("Tipo",COST_TYPES)
            with cc2: val=st.number_input("Valor (R$)",min_value=0.0,step=0.01,format="%.2f")
            with cc3: cd=st.date_input("Data",value=date.today())
            desc=st.text_input("Descrição")
            if st.form_submit_button("Salvar",type="primary",use_container_width=True):
                db.add_animal_cost(aid,ct,desc,val,cd.strftime("%Y-%m-%d"))
                st.success("Custo registrado!"); st.rerun()


def page_animal():
    aid=st.session_state.animal_detail
    if not aid: st.warning("Nenhum animal selecionado."); return
    animal=db.get_animal(aid)
    if not animal: st.error(f"Animal {aid} não encontrado."); return

    if st.button("← Voltar",type="secondary"): _go("rebanho"); st.rerun()

    gmd =db.calculate_gmd(aid)
    ws  =db.get_weighings(aid)
    meds=db.get_medications(aid)
    movs=db.get_movements(aid)
    cost_total=db.get_total_cost(aid)
    yield_     =animal.get("carcass_yield") or 0.52
    arrobas    =db.kg_to_arrobas(animal["current_weight"], yield_)
    gain       =round(animal["current_weight"]-animal["entry_weight"],1)
    wd =db.get_withdrawal_end(aid)
    cat=db.get_age_category(animal.get("birth_date"))
    ul =_unit_label()

    st.markdown(f"## 🐄 Ficha — {animal['id']}  {_status_badge(animal['status'])}",
        unsafe_allow_html=True)

    m=st.columns(6)
    m[0].metric("Raça",       animal["breed"])
    m[1].metric("Categoria",  cat)
    m[2].metric("Peso Atual", f"{animal['current_weight']:.1f} kg")
    m[3].metric("Ganho",      f"{gain:+.1f} kg",
                help="Peso atual menos o peso de entrada")
    m[4].metric("@ Atuais",   f"{arrobas:.2f} @",
                help="Arrobas equivalentes ao peso vivo atual (rendimento de carcaça)")
    gmd_total = db.calculate_gmd_total(animal)
    gmd_txt = f"{gmd:.3f} kg/dia" if gmd else "N/A"
    tot_txt = f"{gmd_total:.3f} kg/dia" if gmd_total else "N/A"
    m[5].metric("GMD recente", gmd_txt,
                delta=f"{gmd:.3f}" if gmd else None,
                help=f"Entre as duas últimas pesagens (atual). GMD total de vida: {tot_txt}")
    st.caption(f"📈 **GMD recente** (entre pesagens): {gmd_txt}  ·  "
               f"**GMD total** (de vida = peso atual − entrada ÷ dias): {tot_txt}")

    src_label = db.AGE_SOURCES.get(animal.get("age_source","propriedade"),"—")
    doc_parts = []
    if animal.get("nf_number"):  doc_parts.append(f"NF: **{animal['nf_number']}**")
    if animal.get("gta_number"): doc_parts.append(f"GTA: **{animal['gta_number']}**")
    st.caption(f"📆 Origem da idade: **{src_label}**"
               + (f" · nascimento: {animal['birth_date']}" if animal.get("birth_date") else "")
               + (f"  |  📄 {' · '.join(doc_parts)}" if doc_parts else ""))

    # Editor de idade
    with st.expander("✏️ Corrigir / redefinir idade"):
        bd2, est2, src2, err2 = _age_inputs(date.today(), f"edit_{aid}_")
        if st.button("💾 Salvar nova idade", key=f"save_age_{aid}", type="primary"):
            if err2:
                st.error(f"❌ {err2}")
            else:
                db.update_animal_age(aid, bd2, est2, src2)
                st.success(f"✅ Idade atualizada · Categoria: **{db.get_age_category(bd2)}**")
                st.rerun()

    if wd:
        st.warning(f"⚠️ Animal em carência até **{wd.isoformat()}** "
                   f"({(wd-date.today()).days} dias restantes). Não pode ser abatido.")

    st.markdown("---")
    _consistencia_regulatoria(animal, movs)

    tl_peso,tl_med,tl_mov,tl_fin,tl_foto,tl_id,tl_ev=st.tabs(
        ["📈 Curva de Peso","💉 Sanidade","🚚 Movimentações","💰 Financeiro","📷 Foto",
         "🏷️ Identificadores","🕒 Linha do Tempo"])

    with tl_ev:
        _linha_do_tempo_do_animal(animal)

    with tl_id:
        _identificadores_do_animal(animal)

    with tl_foto:
        _photo_section(aid, key_prefix="ficha_")

    with tl_peso:
        _render_tab_peso(ws)

    with tl_med:
        _render_tab_med(meds)

    with tl_mov:
        _render_tab_mov(movs)

    with tl_fin:
        _render_tab_fin(aid, animal, gain, yield_, cost_total, ul)

    st.markdown("---")
    qa1,qa2,qa3=st.columns(3)
    with qa1:
        if st.button("📱 Abrir no Campo",use_container_width=True):
            st.session_state.campo_id=aid; _go("campo"); st.rerun()
    with qa2:
        if st.button("📊 Dashboard",use_container_width=True):
            _go("dashboard"); st.rerun()
    with qa3:
        if st.button("📋 Rebanho",use_container_width=True):
            _go("rebanho"); st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# LOTES / PASTAGEM
# ══════════════════════════════════════════════════════════════════════════════
def _sobreposicoes_dos_lotes(lotes: list[dict]) -> list[dict]:
    """Piquetes com polígono desenhado que se cruzam (§ROADMAP, migration 0015).

    Só entram na checagem lotes que **têm** `poligono` — a maioria não vai ter
    tão cedo, e é isso mesmo: o alerta só aparece quando há dado real para
    comparar, nunca por omissão.
    """
    piquetes = []
    for l in lotes:
        if not l.get("poligono"):
            continue
        try:
            anel = _ler_poligono(_poligono_para_texto(l["poligono"]))
        except (ValueError, TypeError):
            continue
        if anel and not geometria_validar(anel):
            piquetes.append({"id": l["id"], "anel": anel})
    if len(piquetes) < 2:
        return []
    return lotacao_sobrepostos(piquetes)


def page_lotes():
    st.markdown('<div class="page-title">🌿 Lotes / Pastagem</div>', unsafe_allow_html=True)
    lotes=db.get_all_lotes()

    sobrepostos = _sobreposicoes_dos_lotes(lotes)
    if sobrepostos:
        pares = ", ".join(f"{s['a']}×{s['b']} ({_num_br(s['pct_do_menor'], 0)}% do menor)"
                          for s in sobrepostos)
        st.warning(f"⚠️ Piquetes com perímetro sobreposto: {pares}. Confira o desenho — "
                   "pode ser piquete redesenhado sem apagar o anterior, ou área "
                   "dividida por engano.")

    lt1,lt2,lt3=st.tabs(["📋 Visão Geral","➕ Novo Lote","🔀 Transferir Animais"])

    with lt1:
        for l in lotes:
            ua  = l["total_ua"] or 0
            cap = l["capacity_ua"] or 0
            has_cap = cap > 0
            pct = min(ua/cap*100, 100) if has_cap else 0
            bar_col=c["primaria"] if pct<75 else c["atencao"] if pct<95 else c["perigo"]
            status_badge={"ativo":'<span class="badge-green">Ativo</span>',
                "descanso":'<span class="badge-yellow">Descanso</span>',
                "reforma":'<span class="badge-red">Reforma</span>'}.get(l["status"],'')
            dias_ocup=""
            if l.get("last_entry_date") and l.get("last_exit_date"):
                d0=datetime.strptime(l["last_entry_date"],"%Y-%m-%d").date()
                d1=datetime.strptime(l["last_exit_date"],"%Y-%m-%d").date()
                dias_ocup=f"Última ocupação: {abs((d1-d0).days)} dias"
            elif l.get("last_entry_date"):
                d0=datetime.strptime(l["last_entry_date"],"%Y-%m-%d").date()
                dias_ocup=f"Em ocupação há {(date.today()-d0).days} dias"

            # Ocupação: só mostra % quando há capacidade definida (> 0)
            if has_cap:
                ocup_txt = f"{ua:.1f} / {cap:.0f} UA ({pct:.0f}%)"
                cap_txt  = f"Cap. {cap:.0f} UA"
                barra = (f'<div style="background:{c["fundo"]};border-radius:6px;height:8px;margin-top:.6rem;overflow:hidden">'
                         f'<div style="background:{bar_col};width:{pct:.0f}%;height:100%;border-radius:6px;transition:width .4s"></div></div>')
            else:
                ocup_txt = f"{ua:.1f} UA · sem capacidade definida"
                cap_txt  = "Sem capacidade de pasto (curral/manejo)"
                barra = ""

            st.markdown(f"""
            <div class="card">
              <div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap">
                <div>
                  <b style="font-size:1.1rem;color:{c["texto"]}">{l['id']} — {l['name']}</b>&nbsp;{status_badge}
                  <div style="color:{c["texto_secundario"]};font-size:.82rem;margin-top:.2rem">
                    {l['area_ha']} ha · {cap_txt} · {dias_ocup}
                  </div>
                </div>
                <div style="text-align:right">
                  <div style="font-size:1.5rem;font-weight:800;color:{c["primaria"]}">{_plural(l['animal_count'],'animal','animais')}</div>
                  <div style="color:{c["texto_secundario"]};font-size:.82rem">{ocup_txt}</div>
                </div>
              </div>
              {barra}
            </div>""",unsafe_allow_html=True)

            # Lista de animais do lote
            with st.expander(f"Ver animais do {l['name']}"):
                anilist=db.get_all_animals(lote_id=l["id"])
                if anilist:
                    a_ids = [a["id"] for a in anilist]
                    gmd_batch = db.calculate_gmd_bulk(a_ids)
                    rows_l=[{"ID":a["id"],"Raça":a["breed"],"Sexo":"♂" if a["sex"]=="M" else "♀",
                        "Peso (kg)":a["current_weight"],"GMD":gmd_batch.get(a["id"])} for a in anilist]
                    st.dataframe(pd.DataFrame(rows_l),use_container_width=True,hide_index=True,
                        column_config={"Peso (kg)":st.column_config.NumberColumn(format="%.1f"),
                            "GMD":st.column_config.NumberColumn(format="%.3f")})
                else:
                    st.caption("Nenhum animal neste lote.")

            # Perímetro do piquete (migration 0015)
            with st.expander(f"🗺️ Perímetro do {l['name']}"):
                st.caption("Um vértice por linha, `longitude, latitude` — a ordem do "
                           "GeoJSON, a mesma da tela de Propriedades. A área é "
                           "**calculada** do desenho, não digitada — salvar o perímetro "
                           "atualiza o campo Área do piquete automaticamente.")
                texto_l = st.text_area(
                    "Vértices", value=_poligono_para_texto(l.get("poligono")),
                    height=120, key=f"lote_poligono_{l['id']}",
                    placeholder="-51.2300, -30.0300\n-51.2280, -30.0300\n"
                                "-51.2280, -30.0320")

                anel_l, erro_l, problemas_l = [], "", []
                if texto_l.strip():
                    try:
                        anel_l = _ler_poligono(texto_l)
                    except ValueError as e:
                        erro_l = str(e)
                    else:
                        problemas_l = geometria_validar(anel_l)

                if erro_l:
                    st.error(f"🚫 {erro_l}")
                for prob in problemas_l:
                    st.error(f"🚫 {prob}")

                if anel_l and not problemas_l:
                    area_desenhada = geometria_area_ha(anel_l)
                    diverge = round(area_desenhada, 2) != round(l["area_ha"] or 0, 2)
                    gl1, gl2 = st.columns(2)
                    gl1.metric("Área do desenho", f"{_num_br(area_desenhada, 2)} ha",
                               help="É o que vai gravar no campo Área ao salvar.")
                    gl2.metric("Área cadastrada hoje", f"{_num_br(l['area_ha'], 2)} ha",
                               delta=(f"{_num_br(area_desenhada - l['area_ha'], 2)} ha ao salvar"
                                     if diverge else None))
                elif not texto_l.strip() and l.get("poligono"):
                    st.caption("Apagar o perímetro **não muda** a Área — ela vira o "
                               "último valor calculado, editável à mão (Trilha 2: "
                               "piquete sem geometria continua funcionando).")

                pode_l = not erro_l and not problemas_l
                if st.button("💾 Salvar perímetro", disabled=not pode_l,
                             key=f"lote_poligono_salvar_{l['id']}"):
                    novo = (json.dumps({"type": "Polygon",
                                        "coordinates": [[list(v) for v in anel_l]]})
                           if anel_l else None)
                    db.set_lote_poligono(l["id"], novo)
                    if novo:
                        st.success(f"✅ Perímetro salvo — Área atualizada para "
                                  f"{_num_br(geometria_area_ha(anel_l), 2)} ha.")
                    else:
                        st.success("✅ Perímetro removido.")
                    st.rerun()

        # Gráfico UA por Lote
        if lotes:
            df_lot=pd.DataFrame([{"Lote":f"{l['id']}·{l['name'][:8]}",
                "UA Atual":l["total_ua"] or 0,"Cap. UA":l["capacity_ua"]} for l in lotes])
            fig_l=go.Figure()
            fig_l.add_bar(x=df_lot["Lote"],y=df_lot["Cap. UA"],name="Capacidade",
                marker_color=c["borda"])
            fig_l.add_bar(x=df_lot["Lote"],y=df_lot["UA Atual"],name="UA Atual",
                marker_color=c["primaria"])
            fig_l.update_layout(**PLOTLY,height=280,barmode="overlay",
                legend=dict(orientation="h",y=1.1),
                xaxis=dict(gridcolor=c["superficie"]),yaxis=dict(gridcolor=c["superficie"],title="UA"))
            st.plotly_chart(fig_l,use_container_width=True)

    with lt2:
        with st.form("f_lote",clear_on_submit=True):
            nl1,nl2=st.columns(2)
            with nl1:
                lid=st.text_input("ID do Lote *",placeholder="Ex: P06").strip().upper()
                name=st.text_input("Nome *",placeholder="Ex: Piquete Sul 2")
            with nl2:
                area=st.number_input("Área (ha)",min_value=0.0,step=0.5,format="%.1f")
                cap=st.number_input("Capacidade (UA)",min_value=0.0,step=1.0,format="%.0f")
            notes_l=st.text_area("Obs.",height=60)
            if st.form_submit_button("✅ Criar Lote",type="primary",use_container_width=True):
                if not lid or not name:
                    st.error("ID e Nome são obrigatórios.")
                elif db.get_lote(lid):
                    st.error(f"Lote {lid} já existe.")
                else:
                    db.add_lote(lid,name,area,cap,notes_l)
                    st.success(f"✅ Lote {lid} criado!"); st.rerun()

    with lt3:
        _lotes_transferir_animais(lotes)


def _lotes_transferir_animais(lotes):
    """Transferência de animais entre piquetes, em lote (§5, Trilha 3).

    `db.move_animals_bulk` já existia como transferência 1 a 1
    (`move_animal`, na ficha do animal); mover um piquete inteiro exigia
    abrir a ficha de cada animal, um de cada vez. Aqui é a mesma operação
    para vários animais de uma vez, na mesma transação.
    """
    st.subheader("🔀 Transferir Animais entre Piquetes")
    st.caption("Move várias cabeças de um piquete para outro numa operação só. Para mover "
               "um animal isolado, a ficha dele (aba Movimentação) continua funcionando "
               "normalmente.")

    if len(lotes) < 2:
        st.info("Cadastre ao menos 2 piquetes para transferir animais entre eles.")
        return

    origem = st.selectbox("Piquete de origem", lotes,
        format_func=lambda l: f"{l['id']} — {l['name']} ({_plural(l['animal_count'],'animal','animais')})",
        key="transf_origem")
    animais_origem = db.get_all_animals(status="ativo", lote_id=origem["id"])
    if not animais_origem:
        st.info(f"Nenhum animal ativo em {origem['name']} para transferir.")
        return

    opts = {f"{a['id']} · {a['breed']} · {a['current_weight']:.0f}kg": a["id"]
            for a in animais_origem}
    selecionados = st.multiselect("Animais a transferir", list(opts.keys()),
        key="transf_animais")
    sel_ids = [opts[s] for s in selecionados]

    destinos_possiveis = [l for l in lotes if l["id"] != origem["id"]]
    with st.form("f_transferencia", clear_on_submit=True):
        tc1, tc2 = st.columns(2)
        with tc1:
            destino = st.selectbox("Piquete de destino", destinos_possiveis,
                format_func=lambda l: f"{l['id']} — {l['name']}")
            mv_date = st.date_input("Data", value=date.today())
        with tc2:
            reason = st.selectbox("Motivo",
                ["manejo", "pesagem", "tratamento", "separação"])
        notes_t = st.text_area("Obs.", height=60, placeholder="Opcional")

        if st.form_submit_button("✅ Transferir", type="primary", use_container_width=True):
            if not sel_ids:
                st.error("Selecione ao menos um animal.")
            else:
                r = db.move_animals_bulk(sel_ids, destino["id"], mv_date.strftime("%Y-%m-%d"),
                    reason, st.session_state.user["name"], notes_t)
                if r["movidos"]:
                    st.success(f"✅ {_plural(len(r['movidos']),'animal transferido','animais transferidos')} "
                              f"para {destino['name']}.")
                if r["erros"]:
                    st.error(f"❌ {_plural(len(r['erros']),'animal não encontrado','animais não encontrados')}: "
                            f"{', '.join(r['erros'])}")
                if r["movidos"] or r["erros"]:
                    st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# FINANCEIRO
# ══════════════════════════════════════════════════════════════════════════════
def _fin_precos():
    """Tabela de valor esperado de venda por categoria (R$/kg)."""
    st.subheader("🏷️ Valor Esperado por Categoria (R$/kg)")
    st.caption("Informe quanto você espera receber por kg em cada categoria (idade × sexo). "
               "Serve de referência e preenche automaticamente a venda por kg.")
    precos = db.get_category_prices()
    with st.form("f_precos"):
        st.markdown("**Preço esperado por kg (R$)**")
        novos = {}
        for band in AGE_BANDS:
            c1, c2, c3 = st.columns([2,1,1])
            with c1:
                st.markdown(f"<div style='padding-top:.55rem'>{band}</div>", unsafe_allow_html=True)
            with c2:
                novos[(band,"M")] = st.number_input(f"♂ Macho", min_value=0.0, step=0.10,
                    value=float(precos.get((band,"M"),0.0)), key=f"pk_{band}_M", format="%.2f")
            with c3:
                novos[(band,"F")] = st.number_input(f"♀ Fêmea", min_value=0.0, step=0.10,
                    value=float(precos.get((band,"F"),0.0)), key=f"pk_{band}_F", format="%.2f")
        if st.form_submit_button("💾 Salvar Preços", type="primary", use_container_width=True):
            for (band,sex), val in novos.items():
                if val != precos.get((band,sex), 0.0):
                    db.set_category_price(band, sex, val)
            st.success("✅ Preços por categoria atualizados!")
            st.rerun()


def _fin_venda(animals):
    """Registro de venda (por kg / cabeça / lote), com lucro real."""
    st.subheader("💵 Registrar Venda")
    if not animals:
        st.info("Não há animais ativos para vender.");
    else:
        modo = st.radio("Modo de precificação", list(db.PRICING_MODES.keys()),
            format_func=lambda m: db.PRICING_MODES[m], horizontal=True, key="venda_modo")
        tipo = st.radio("Tipo de venda", list(db.SALE_TYPES.keys()),
            format_func=lambda t: db.SALE_TYPES[t], horizontal=True, key="venda_tipo")

        prazo = st.checkbox("Venda a prazo?", key="venda_a_prazo",
            help="Desmarcado (padrão) = à vista, comportamento de sempre. Marcado gera "
                 "as parcelas em Contas a Receber (aba em Financeiro).")
        num_parcelas_v, primeira_parcela_v = 1, None
        if prazo:
            pvc1, pvc2 = st.columns(2)
            with pvc1:
                num_parcelas_v = st.number_input("Número de parcelas", min_value=1,
                    max_value=36, value=1, step=1, key="venda_num_parcelas")
            with pvc2:
                primeira_parcela_v = st.date_input("Vencimento da 1ª parcela",
                    value=date.today()+timedelta(days=30), key="venda_primeira_parcela")

        # Seleção de animais
        opts = {f"{a['id']} · {a['breed']} · {a['current_weight']:.0f}kg · {db.get_age_category(a.get('birth_date'))}": a['id']
                for a in animals}
        multi = modo in ("lote",) or True   # sempre permite múltiplos
        sel = st.multiselect("Animais a vender", list(opts.keys()),
            help="Selecione um ou mais animais")
        sel_ids = [opts[s] for s in sel]
        sel_animals = [a for a in animals if a["id"] in sel_ids]
        peso_total = sum(a["current_weight"] for a in sel_animals)

        with st.form("f_venda", clear_on_submit=True):
            if modo == "kg":
                # sugere preço da categoria do 1º animal
                sug = 0.0
                if sel_animals:
                    sug = db.get_expected_price_kg(
                        db.get_age_category(sel_animals[0].get("birth_date")), sel_animals[0]["sex"])
                valor = st.number_input("Preço por kg (R$)", min_value=0.0, step=0.10,
                    value=float(sug), format="%.2f",
                    help="Sugerido pela tabela de categoria; ajuste se necessário")
                if peso_total:
                    st.caption(f"Peso total selecionado: **{peso_total:.0f} kg** → "
                               f"receita estimada: **R$ {peso_total*valor:,.2f}**")
            elif modo == "cabeca":
                valor = st.number_input("Valor por cabeça (R$)", min_value=0.0, step=50.0, format="%.2f",
                    help="Mesmo valor para cada animal selecionado")
                if sel_animals:
                    st.caption(f"{len(sel_animals)} animais → receita: **R$ {valor*len(sel_animals):,.2f}**")
            else:  # lote
                valor = st.number_input("Valor TOTAL do lote (R$)", min_value=0.0, step=100.0, format="%.2f",
                    help="Valor fechado do grupo; será rateado proporcionalmente ao peso")
                if peso_total:
                    st.caption(f"{len(sel_animals)} animais, {peso_total:.0f} kg → "
                               "rateio proporcional ao peso")
            c1, c2 = st.columns(2)
            with c1: sale_date = st.date_input("Data da venda", value=date.today())
            with c2: buyer = st.text_input("Comprador", placeholder="Ex: Frigorífico / Fazenda X")
            notes = st.text_input("Observações", placeholder="Opcional")

            if st.form_submit_button("✅ Confirmar Venda", type="primary", use_container_width=True):
                if not sel_ids:
                    st.error("Selecione ao menos um animal.")
                elif valor <= 0:
                    st.error("Informe um valor maior que zero.")
                elif prazo and not primeira_parcela_v:
                    st.error("Informe o vencimento da 1ª parcela.")
                else:
                    r = db.register_sale(sel_ids, sale_date.strftime("%Y-%m-%d"), tipo, modo,
                        valor, buyer=buyer, operator=st.session_state.user["name"], notes=notes,
                        a_prazo=prazo, num_parcelas=int(num_parcelas_v),
                        primeiro_vencimento=primeira_parcela_v.isoformat() if primeira_parcela_v else None)
                    cor = c["primaria"] if r["lucro"] >= 0 else c["perigo"]
                    st.success(f"✅ {r['n']} animal(is) vendido(s)!")
                    st.markdown(
                        f"<div class='card'>Receita: <b>R$ {r['receita']:,.2f}</b> · "
                        f"Custo: <b>R$ {r['custo']:,.2f}</b> · "
                        f"<b style='color:{cor}'>{'Lucro' if r['lucro']>=0 else 'Prejuízo'}: "
                        f"R$ {r['lucro']:,.2f}</b></div>", unsafe_allow_html=True)
                    if r.get("parcelas_a_receber"):
                        st.caption(f"📥 {r['parcelas_a_receber']} parcela(s) geradas em "
                                  "Contas a Receber (aba em Financeiro).")
                    st.rerun()

    # Histórico de vendas
    st.markdown("---")
    st.markdown("**📜 Vendas Registradas**")
    vendas = db.get_sales()
    if vendas:
        df_v = pd.DataFrame(vendas)[["sale_date","animal_id","breed","sale_type","pricing_mode",
                                     "weight_kg","total_value","cost_at_sale","profit","buyer"]].copy()
        df_v["sale_type"]=df_v["sale_type"].map(lambda t: db.SALE_TYPES.get(t,t))
        df_v["pricing_mode"]=df_v["pricing_mode"].map(lambda m: {"kg":"kg","cabeca":"cabeça","lote":"lote"}.get(m,m))
        df_v.columns=["Data","Animal","Raça","Tipo","Modo","Peso (kg)","Receita (R$)","Custo (R$)","Lucro (R$)","Comprador"]
        st.dataframe(df_v, use_container_width=True, hide_index=True,
            column_config={col: st.column_config.NumberColumn(format="R$ %.2f")
                           for col in ["Receita (R$)","Custo (R$)","Lucro (R$)"]})
        tot_luc = sum(v["profit"] for v in vendas)
        st.metric("Lucro/Prejuízo acumulado nas vendas", f"R$ {tot_luc:,.2f}")
    else:
        st.info("Nenhuma venda registrada ainda.")

    # Custo por lote de venda (ROADMAP §5, Trilha 3 — último item da trilha:
    # já existia custo/kg e custo/@ por animal (aba "Custos por Animal") e
    # por piquete (Nutrição, `_nutricao_custo_por_piquete`); faltava por
    # LOTE DE VENDA, o agrupamento que `register_sale` já grava em `lot_ref`
    # sempre que a venda sai com mais de um animal.
    st.markdown("---")
    st.markdown("**📦 Custo por Lote de Venda**")
    st.caption("Cada lote agrupa as vendas que saíram juntas (`lot_ref`) — venda "
               "avulsa de um único animal vira seu próprio lote de 1 cabeça, nunca "
               "se mistura com outra venda avulsa.")
    if vendas:
        lotes_venda = por_lote_de_venda(vendas)
        df_l = pd.DataFrame([{
            "Lote": lv["lot_ref"] or "(venda avulsa)",
            "Data": lv["sale_date"],
            "Cabeças": lv["animais"],
            "Peso Total (kg)": lv["peso_total_kg"],
            "Custo Total (R$)": lv["custo_total"],
            "Custo/kg (R$)": lv["custo_por_kg"],
            "Custo/@ (R$)": lv["custo_por_arroba"],
            "Lucro (R$)": lv["lucro_total"],
        } for lv in lotes_venda])
        st.dataframe(df_l, use_container_width=True, hide_index=True,
            column_config={
                "Peso Total (kg)": st.column_config.NumberColumn(format="%.1f"),
                "Custo Total (R$)": st.column_config.NumberColumn(format="R$ %.2f"),
                "Custo/kg (R$)": st.column_config.NumberColumn(format="R$ %.2f"),
                "Custo/@ (R$)": st.column_config.NumberColumn(format="R$ %.2f"),
                "Lucro (R$)": st.column_config.NumberColumn(format="R$ %.2f")})
    else:
        st.info("Nenhuma venda registrada ainda.")


def _fin_lancamentos(start_iso=None, end_iso=None) -> list[dict]:
    """Lista única de lançamentos (spec 0034) — vendas, custos fixos, custos
    por animal e compras de insumo, no formato que `services.caixa` espera.

    Compra de insumo hoje é gravada como `type='entrada'` + `reason='compra'`
    (nunca `type='compra'`, que é o que a spec 0034 supunha existir na
    tabela) — `db.get_insumo_compras` já traduz isso; aqui só falta encaixar
    o preço unitário no formato aninhado que `normalizar()` espera.
    """
    vendas = db.get_sales(start_iso, end_iso)
    custos_fixos = db.get_fixed_costs(start_iso, end_iso)
    custos_animal = db.get_all_animal_costs(start_iso, end_iso)
    compras_insumo = [{
        "transaction_date": c["transaction_date"],
        "quantity": c["quantity"],
        "type": "compra",
        "insumo": {"name": c["insumo_nome"], "cost_per_unit": c["insumo_cost_per_unit"]},
    } for c in db.get_insumo_compras(start_iso, end_iso)]
    return lancamentos_normalizar(vendas=vendas, custos_fixos=custos_fixos,
                                  custos_animal=custos_animal,
                                  compras_insumo=compras_insumo)


def _fin_lancamentos_caixa() -> list[dict]:
    """Lançamentos para o fluxo de caixa (§5, Trilha 3) — diferente de
    `_fin_lancamentos` (competência), que conta TODA venda/compra no dia do
    fato: aqui o que tem parcela (`contas_pagar`/`contas_receber`) entra só
    pela parcela, nunca pelo lançamento de origem, senão duplica.

    Por isso venda a prazo é excluída de `vendas` (o caixa dela mora em
    `contas_receber`) e compra com nota fiscal é excluída de
    `compras_insumo` (o caixa dela mora em `contas_pagar` — identificada por
    `compra_id`, que só existe em compra feita via `repositories.compras.
    registrar`; entrada avulsa sem nota continua contando direto).

    Sem filtro de período aqui de propósito: `fluxo_de_caixa`/`em_aberto` já
    filtram por vencimento/pagamento, e `em_aberto` precisa enxergar TODA
    conta ainda aberta, não só as criadas numa janela.
    """
    vendas_a_vista = [v for v in db.get_sales() if not v.get("a_prazo")]
    custos_fixos = db.get_fixed_costs()
    custos_animal = db.get_all_animal_costs()
    compras_insumo = [{
        "transaction_date": c["transaction_date"],
        "quantity": c["quantity"],
        "type": "compra",
        "insumo": {"name": c["insumo_nome"], "cost_per_unit": c["insumo_cost_per_unit"]},
    } for c in db.get_insumo_compras() if not c.get("compra_id")]
    contas_pagar = [c for c in db.compras.listar_contas_pagar() if c["status"] != "cancelado"]
    contas_receber = [c for c in db.listar_contas_receber() if c["status"] != "cancelado"]
    return lancamentos_normalizar(
        vendas=vendas_a_vista, custos_fixos=custos_fixos, custos_animal=custos_animal,
        compras_insumo=compras_insumo, contas_pagar=contas_pagar,
        contas_receber=contas_receber)


def _fin_fluxo_de_caixa():
    """Fluxo de caixa realizado e projetado (§5, Trilha 3) —
    `services.caixa.fluxo_de_caixa`/`em_aberto` (spec 0021), órfãos até
    aqui: só faziam sentido depois de existir vencimento/pagamento de
    verdade (contas_pagar/contas_receber, PRs anteriores desta trilha).

    `fluxo_de_caixa` soma `valor` sem olhar o sinal de `tipo` (contrato já
    testado em `tests/test_caixa.py`, não mexido aqui) — por isso é chamada
    duas vezes, uma por tipo, e o saldo com sinal certo é montado aqui.
    """
    st.subheader("💵 Fluxo de Caixa")
    st.caption("Realizado = já pago/recebido na janela. Projetado = vencimento na "
               "janela, ainda em aberto. Só o que tem parcela (contas a pagar/"
               "receber) pode aparecer como projetado — o resto é sempre à vista.")

    c1, c2 = st.columns(2)
    with c1: start = st.date_input("De", value=date.today().replace(day=1), key="fx_start")
    with c2: end = st.date_input("Até", value=date.today()+timedelta(days=60), key="fx_end",
                                 help="Padrão: hoje + 60 dias, para o projetado aparecer sem "
                                      "precisar ajustar a data — realizado não passa de hoje "
                                      "de qualquer forma.")

    lancamentos = _fin_lancamentos_caixa()
    receitas_l = [l for l in lancamentos if l["tipo"] == "receita"]
    despesas_l = [l for l in lancamentos if l["tipo"] == "despesa"]

    fx_r = fluxo_de_caixa(receitas_l, start.isoformat(), end.isoformat())
    fx_d = fluxo_de_caixa(despesas_l, start.isoformat(), end.isoformat())

    realizado = round(fx_r["realizado"] - fx_d["realizado"], 2)
    projetado = round(fx_r["projetado"] - fx_d["projetado"], 2)
    saldo_projetado = round(realizado + projetado, 2)

    k1, k2, k3 = st.columns(3)
    k1.metric("Realizado no período", f"R$ {realizado:,.2f}",
             help=f"Entrou R$ {fx_r['realizado']:,.2f} · Saiu R$ {fx_d['realizado']:,.2f}")
    k2.metric("Projetado no período", f"R$ {projetado:,.2f}",
             help=f"A entrar R$ {fx_r['projetado']:,.2f} · A sair R$ {fx_d['projetado']:,.2f}")
    k3.metric("Saldo (realizado + projetado)", f"R$ {saldo_projetado:,.2f}",
             delta=f"{saldo_projetado:+,.2f}")

    st.markdown("---")
    st.markdown("**📋 Em Aberto**")
    hoje = date.today().isoformat()
    abertas = caixa_em_aberto(lancamentos, hoje)
    if not abertas:
        st.success("✅ Nenhuma conta em aberto.")
        return

    vencidas = [a for a in abertas if a["dias_atraso"] > 0]
    if vencidas:
        st.warning(f"⚠️ **{_plural(len(vencidas),'conta vencida','contas vencidas')}** "
                  "entre as em aberto.")

    TIPO_LABEL = {"receita": "📥 A Receber", "despesa": "📤 A Pagar"}
    rows = [{
        "Tipo": TIPO_LABEL.get(a["tipo"], a["tipo"]),
        "Descrição": a["categoria"],
        "Vencimento": a["vencimento"],
        "Valor (R$)": a["valor"],
        "Situação": f"🔴 Vencida há {a['dias_atraso']}d" if a["dias_atraso"] > 0 else "🟡 Em dia",
    } for a in abertas]
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True,
        column_config={"Valor (R$)": st.column_config.NumberColumn(format="R$ %.2f")})


def _fin_competencia():
    """Resultado por competência (spec 0034 + `services.caixa`), últimos 6 meses.

    Diferente do "Resultado Consolidado" (um total só para o intervalo
    escolhido), aqui cada lançamento conta no mês do fato — o que revela
    tendência entre meses que um total único esconde.
    """
    st.subheader("📅 Resultado por Competência")
    st.caption("Receita e despesa contadas no mês do fato — venda, custo fixo, custo "
               "de animal ou compra de insumo —, não somadas num total só. "
               "Últimos 6 meses.")

    hoje = date.today()
    meses = []
    for i in range(5, -1, -1):
        m, a = hoje.month - i, hoje.year
        while m <= 0:
            m += 12; a -= 1
        meses.append((a, m))

    inicio_janela = date(meses[0][0], meses[0][1], 1)
    lancamentos = _fin_lancamentos(inicio_janela.isoformat(), hoje.isoformat())

    linhas = []
    for ano, mes in meses:
        r = resultado_por_competencia(lancamentos, ano, mes)
        linhas.append({"Mês": f"{mes:02d}/{ano}", "Receitas": r["receitas"],
                       "Despesas": r["despesas"], "Resultado": r["resultado"],
                       "_por_categoria": r["por_categoria"]})

    ultimo = linhas[-1]
    k1, k2, k3 = st.columns(3)
    k1.metric(f"Receitas — {ultimo['Mês']}", f"R$ {ultimo['Receitas']:,.2f}")
    k2.metric(f"Despesas — {ultimo['Mês']}", f"R$ {ultimo['Despesas']:,.2f}")
    k3.metric(f"Resultado — {ultimo['Mês']}", f"R$ {ultimo['Resultado']:,.2f}",
             delta=f"{ultimo['Resultado']:+,.2f}")

    df_comp = pd.DataFrame(linhas)[["Mês", "Receitas", "Despesas", "Resultado"]]
    df_melt = df_comp.melt(id_vars="Mês", value_vars=["Receitas", "Despesas"],
                           var_name="Tipo", value_name="Valor")
    fig = px.bar(df_melt, x="Mês", y="Valor", color="Tipo", barmode="group",
        color_discrete_map={"Receitas": c["primaria"], "Despesas": c["perigo"]})
    fig.update_layout(**PLOTLY, height=300, xaxis=dict(gridcolor=c["superficie"]),
                      yaxis=dict(gridcolor=c["superficie"]))
    st.plotly_chart(fig, use_container_width=True)

    st.dataframe(df_comp, use_container_width=True, hide_index=True,
        column_config={"Receitas": st.column_config.NumberColumn(format="R$ %.2f"),
                       "Despesas": st.column_config.NumberColumn(format="R$ %.2f"),
                       "Resultado": st.column_config.NumberColumn(format="R$ %.2f")})

    if ultimo["_por_categoria"]:
        st.markdown(f"**Por categoria — {ultimo['Mês']}**")
        st.dataframe(pd.DataFrame(ultimo["_por_categoria"]).rename(
            columns={"categoria": "Categoria", "valor": "Valor (R$)"}),
            use_container_width=True, hide_index=True,
            column_config={"Valor (R$)": st.column_config.NumberColumn(format="R$ %.2f")})


def _fin_centros_de_custo(lotes):
    """Custo por centro de custo — o piquete (§5, Trilha 3).

    `services.centros_de_custo.consolidar` junta custo fixo alocado
    (`fixed_costs.lote_id`, aba 🏢 Custos Fixos) com custo por animal
    (`animal_costs`, pelo piquete ATUAL do animal — mesma limitação já
    conhecida de `_nutricao_custo_por_piquete`: animal que mudou de piquete
    no meio do período carrega o custo histórico inteiro para o piquete de
    hoje).
    """
    st.subheader("🏭 Centros de Custo")
    st.caption("Custo fixo alocado a um piquete (aba 🏢 Custos Fixos) + custo por animal, "
               "somado pelo piquete onde cada animal está **hoje**. 'Geral da Fazenda' é o "
               "custo fixo que não foi alocado a nenhum piquete específico.")

    c1, c2 = st.columns(2)
    with c1: start = st.date_input("De", value=date(date.today().year,1,1), key="cc_start")
    with c2: end = st.date_input("Até", value=date.today(), key="cc_end")
    s_iso, e_iso = start.isoformat(), end.isoformat()

    fixos_por_lote = db.get_fixed_costs_by_lote(s_iso, e_iso)
    animal_por_lote = db.get_animal_costs_by_lote(s_iso, e_iso)
    if not fixos_por_lote and not animal_por_lote:
        st.info("Nenhum custo lançado no período selecionado.")
        return

    cabecas_por_lote = {l["id"]: l["animal_count"] for l in lotes}
    nomes_lotes = {l["id"]: l["name"] for l in lotes}
    linhas = consolidar_centros_de_custo(fixos_por_lote, animal_por_lote,
                                         cabecas_por_lote, nomes_lotes)

    total_geral = round(sum(l["total"] for l in linhas), 2)
    k1, k2 = st.columns(2)
    k1.metric("Total no período", f"R$ {total_geral:,.2f}")
    k2.metric("Centros de custo com lançamento", len(linhas))

    df = pd.DataFrame(linhas)[["nome", "cabecas", "custos_fixos", "custos_animal", "total"]]
    df.columns = ["Centro de Custo", "Cabeças", "Custos Fixos (R$)", "Custos de Animal (R$)",
                  "Total (R$)"]
    st.dataframe(df, use_container_width=True, hide_index=True,
        column_config={col: st.column_config.NumberColumn(format="R$ %.2f")
                       for col in ["Custos Fixos (R$)", "Custos de Animal (R$)", "Total (R$)"]})

    fig = px.bar(df, x="Centro de Custo",
                y=["Custos Fixos (R$)", "Custos de Animal (R$)"],
                barmode="stack", color_discrete_sequence=[c["perigo"], c["atencao"]])
    fig.update_layout(**PLOTLY, height=300, xaxis=dict(gridcolor=c["superficie"]),
                      yaxis=dict(gridcolor=c["superficie"], title="R$"))
    st.plotly_chart(fig, use_container_width=True)


def _render_account_metrics(contas: list) -> str:
    """Renderiza métricas de contas em aberto e vencidas, retornando a data atual em ISO."""
    hoje = date.today().isoformat()
    abertas = [c for c in contas if c["status"] == "aberto"]
    vencidas = [c for c in abertas if c["vencimento"] < hoje]

    kk = st.columns(3)
    kk[0].metric("Em Aberto", len(abertas))
    kk[1].metric("Vencidas", len(vencidas))
    kk[2].metric("Total em Aberto", f"R$ {sum(c['valor'] for c in abertas):,.2f}")

    if vencidas:
        st.warning(f"⚠️ **{_plural(len(vencidas),'conta vencida','contas vencidas')}:** " +
            ", ".join(f"**{c['descricao']}** ({c['parcela_numero']}/{c['parcela_total']})"
                     for c in vencidas))
    return hoje

def _fin_contas_a_pagar():
    """Parcelas geradas por compra de insumo com nota fiscal (§5, Trilha 3).

    Diferente de Custos Fixos e Custos por Animal (gasto já ocorrido, sem
    prazo), aqui a conta nasce **antes** do pagamento — tem vencimento e pode
    estar vencida. `db.compras.registrar()` (aba 🛒 Compra com Nota Fiscal,
    em Estoque) é quem gera; aqui só se acompanha e se marca como paga.
    """
    st.caption("Parcelas das compras de insumo com nota fiscal (aba 🛒 Compra com "
               "Nota Fiscal, em Estoque). Vencidas aparecem destacadas.")

    contas = db.compras.listar_contas_pagar()
    if not contas:
        st.info("Nenhuma conta a pagar registrada ainda.")
        return

    hoje = _render_account_metrics(contas)

    STATUS_LABEL = {"aberto": "🟡 Aberto", "pago": "🟢 Pago", "cancelado": "⚪ Cancelado"}
    f_status = st.selectbox("Filtrar por situação", ["Todas", "aberto", "pago", "cancelado"],
        format_func=lambda s: s if s == "Todas" else STATUS_LABEL.get(s, s),
        key="pag_filtro_status")
    filtradas = contas if f_status == "Todas" else [c for c in contas if c["status"] == f_status]

    rows = [{"Descrição": c["descricao"], "Parcela": f"{c['parcela_numero']}/{c['parcela_total']}",
        "Vencimento": c["vencimento"], "Valor (R$)": c["valor"],
        "Situação": ("🔴 Vencida" if c["status"] == "aberto" and c["vencimento"] < hoje
                     else STATUS_LABEL.get(c["status"], c["status"]))}
        for c in filtradas]
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True,
        column_config={"Valor (R$)": st.column_config.NumberColumn(format="R$ %.2f")})

    pendentes = [c for c in filtradas if c["status"] == "aberto"]
    if pendentes:
        with st.expander("💳 Marcar como paga"):
            conta_sel = st.selectbox("Conta", pendentes,
                format_func=lambda c: f"{c['descricao']} ({c['parcela_numero']}/{c['parcela_total']}) — "
                    f"R$ {c['valor']:.2f}, vence {c['vencimento']}",
                key="pag_conta_selecionada")
            pg1, pg2 = st.columns(2)
            with pg1:
                data_pg = st.date_input("Data do pagamento", value=date.today(), key="pag_data")
            with pg2:
                forma_pg = st.selectbox("Forma de pagamento",
                    ["pix", "boleto", "transferência", "dinheiro", "cartão", "outro"],
                    key="pag_forma")
            if st.button("✅ Confirmar Pagamento", type="primary", key="pag_confirmar_btn"):
                ok = db.compras.marcar_pago(conta_sel["id"], data_pg.isoformat(), forma_pg)
                if ok:
                    st.success("✅ Conta marcada como paga.")
                    st.rerun()
                else:
                    st.error("Não foi possível marcar como paga.")


def _fin_contas_a_receber():
    """Parcelas geradas por venda a prazo (§5, Trilha 3) — espelha `_fin_contas_a_pagar`.

    `db.register_sale(..., a_prazo=True, ...)` (aba 💵 Registrar Venda) é quem
    gera; venda à vista (o padrão) nunca aparece aqui.
    """
    st.caption("Parcelas das vendas a prazo (aba 💵 Registrar Venda, marcando \"Venda a "
               "prazo?\"). Venda à vista não gera nada aqui. Vencidas aparecem destacadas.")

    contas = db.listar_contas_receber()
    if not contas:
        st.info("Nenhuma conta a receber registrada ainda.")
        return

    hoje = _render_account_metrics(contas)

    STATUS_LABEL = {"aberto": "🟡 Aberto", "recebido": "🟢 Recebido", "cancelado": "⚪ Cancelado"}
    f_status = st.selectbox("Filtrar por situação", ["Todas", "aberto", "recebido", "cancelado"],
        format_func=lambda s: s if s == "Todas" else STATUS_LABEL.get(s, s),
        key="rec_filtro_status")
    filtradas = contas if f_status == "Todas" else [c for c in contas if c["status"] == f_status]

    rows = [{"Descrição": c["descricao"], "Parcela": f"{c['parcela_numero']}/{c['parcela_total']}",
        "Vencimento": c["vencimento"], "Valor (R$)": c["valor"],
        "Situação": ("🔴 Vencida" if c["status"] == "aberto" and c["vencimento"] < hoje
                     else STATUS_LABEL.get(c["status"], c["status"]))}
        for c in filtradas]
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True,
        column_config={"Valor (R$)": st.column_config.NumberColumn(format="R$ %.2f")})

    pendentes = [c for c in filtradas if c["status"] == "aberto"]
    if pendentes:
        with st.expander("💳 Marcar como recebida"):
            conta_sel = st.selectbox("Conta", pendentes,
                format_func=lambda c: f"{c['descricao']} ({c['parcela_numero']}/{c['parcela_total']}) — "
                    f"R$ {c['valor']:.2f}, vence {c['vencimento']}",
                key="rec_conta_selecionada")
            rg1, rg2 = st.columns(2)
            with rg1:
                data_rec = st.date_input("Data do recebimento", value=date.today(), key="rec_data")
            with rg2:
                forma_rec = st.selectbox("Forma de recebimento",
                    ["pix", "boleto", "transferência", "dinheiro", "cartão", "outro"],
                    key="rec_forma")
            if st.button("✅ Confirmar Recebimento", type="primary", key="rec_confirmar_btn"):
                ok = db.marcar_recebido(conta_sel["id"], data_rec.isoformat(), forma_rec)
                if ok:
                    st.success("✅ Conta marcada como recebida.")
                    st.rerun()
                else:
                    st.error("Não foi possível marcar como recebida.")


def _fin_rentabilidade_por_raca():
    """Rentabilidade de ciclos encerrados por raça (spec 0042).

    Diferente de "🏆 Origem" (agrupa por fornecedor, todo o histórico), aqui
    agrupa por raça e olha só para ciclos que já venderam — um ciclo por
    venda, não por animal. Por isso usa `get_all_animals(status=None)`, não
    o rebanho ativo: o animal vendido já não está mais ativo.
    """
    st.subheader("🐄 Rentabilidade por Raça")
    st.caption("Lucro por cabeça, por arroba produzida e GMD médio de cada raça — "
               "só ciclos **encerrados** (já vendidos), um ciclo por venda.")

    vendas = db.get_sales()
    if not vendas:
        st.info("Nenhuma venda registrada ainda.")
        return

    animais_por_uuid = {a["uuid"]: a for a in db.get_all_animals(status=None)
                        if a.get("uuid")}
    custos_por_id = db._costs_by_animal()
    custo_total_por_uuid = {uuid: custos_por_id.get(a["id"], 0.0)
                            for uuid, a in animais_por_uuid.items()}

    ciclos = montar_ciclos(vendas, animais_por_uuid, custo_total_por_uuid)
    ranking = ranking_por_raca(ciclos)
    if not ranking:
        st.info("Nenhum ciclo com receita informada ainda.")
        return

    df_r = pd.DataFrame([{
        "Raça": r["raca"], "Animais": r["animais"],
        "Lucro/Cabeça (R$)": r["lucro_por_cabeca"],
        "Lucro/@ produzida (R$)": r["lucro_por_arroba_produzida"],
        "GMD Médio (kg/dia)": r["gmd_medio"],
        "Margem (%)": r["margem"] * 100,
    } for r in ranking])

    st.dataframe(df_r, use_container_width=True, hide_index=True,
        column_config={
            "Lucro/Cabeça (R$)": st.column_config.NumberColumn(format="R$ %.2f"),
            "Lucro/@ produzida (R$)": st.column_config.NumberColumn(format="R$ %.2f"),
            "GMD Médio (kg/dia)": st.column_config.NumberColumn(format="%.3f"),
            "Margem (%)": st.column_config.NumberColumn(format="%.1f%%")})

    c1, c2 = st.columns(2)
    with c1:
        fig_l = px.bar(df_r, x="Raça", y="Lucro/Cabeça (R$)",
            color="Lucro/Cabeça (R$)", color_continuous_scale=ESCALA_RUIM_BOM,
            text="Lucro/Cabeça (R$)")
        fig_l.update_traces(texttemplate="R$ %{text:.0f}", textposition="outside")
        fig_l.update_layout(**PLOTLY, height=300, coloraxis_showscale=False,
            title="Lucro por cabeça, por raça",
            xaxis=dict(gridcolor=c["superficie"]), yaxis=dict(gridcolor=c["superficie"]))
        st.plotly_chart(fig_l, use_container_width=True)
    with c2:
        fig_m = px.bar(df_r, x="Raça", y="Margem (%)",
            color="Margem (%)", color_continuous_scale=ESCALA_RUIM_BOM,
            text="Margem (%)")
        fig_m.update_traces(texttemplate="%{text:.1f}%", textposition="outside")
        fig_m.update_layout(**PLOTLY, height=300, coloraxis_showscale=False,
            title="Margem por raça",
            xaxis=dict(gridcolor=c["superficie"]), yaxis=dict(gridcolor=c["superficie"]))
        st.plotly_chart(fig_m, use_container_width=True)

    melhor = ranking[0]
    st.success(f"🥇 Melhor lucro por cabeça: **{melhor['raca']}** "
               f"(R$ {melhor['lucro_por_cabeca']:,.2f}). Raça com margem negativa "
               "teve custo maior que receita no período — não é erro de exibição.")


def _fin_resultado():
    """Planilha financeira consolidada (todas as entradas e saídas)."""
    st.subheader("📒 Resultado Consolidado")
    c1, c2 = st.columns(2)
    with c1: start = st.date_input("De", value=date(date.today().year,1,1), key="res_start")
    with c2: end = st.date_input("Até", value=date.today(), key="res_end")
    fin = db.get_financial_summary(start.isoformat(), end.isoformat())

    st.markdown("**💸 Saídas (custos)**")
    s = st.columns(5)
    s[0].metric("Compra de animais", f"R$ {fin['compra_animais']:,.0f}")
    s[1].metric("Medicamentos",      f"R$ {fin['medicamentos']:,.0f}")
    s[2].metric("Nutrição/Trato",    f"R$ {fin['nutricao']:,.0f}")
    s[3].metric("Operacional",       f"R$ {fin['operacional']:,.0f}")
    s[4].metric("Custos fixos",      f"R$ {fin['custos_fixos']:,.0f}")

    st.markdown("**💰 Entradas (vendas)**")
    vendas = fin["vendas"]
    e = st.columns(3)
    abate = vendas.get("abate", {"receita":0,"lucro":0,"n":0})
    criacao = vendas.get("criacao", {"receita":0,"lucro":0,"n":0})
    e[0].metric("Vendas p/ abate",   f"R$ {abate['receita']:,.0f}", help=f"{abate['n']} animais")
    e[1].metric("Vendas p/ criação", f"R$ {criacao['receita']:,.0f}", help=f"{criacao['n']} animais")
    e[2].metric("Receita total",     f"R$ {fin['receita_total']:,.0f}")

    if fin.get("perda_mortalidade", 0) > 0:
        st.markdown(
            f"<div class='card-red'>☠️ <b>Perda por mortalidade no período:</b> "
            f"R$ {fin['perda_mortalidade']:,.2f} — investimento em animais que morreram "
            f"e não vão gerar receita (já incluído nas saídas acima).</div>",
            unsafe_allow_html=True)

    st.markdown("---")
    res = fin["resultado"]
    cor = c["primaria"] if res >= 0 else c["perigo"]
    rk = st.columns(3)
    rk[0].metric("Total de Saídas",  f"R$ {fin['saidas_total']:,.2f}")
    rk[1].metric("Total de Entradas",f"R$ {fin['receita_total']:,.2f}")
    rk[2].metric("Resultado (Caixa)", f"R$ {res:,.2f}",
                 delta=f"{res:+,.2f}", delta_color="normal")
    st.caption("ℹ️ 'Resultado (Caixa)' é o fluxo do período: tudo que saiu (incluindo a compra de "
               "animais que ainda estão no pasto) menos tudo que entrou. O lucro *realizado* por "
               "venda aparece na aba **Registrar Venda**.")

    # Gráfico saídas x entradas
    df_fluxo = pd.DataFrame({
        "Categoria": ["Compra","Medicamentos","Nutrição","Operacional","Fixos","Vendas"],
        "Valor": [fin['compra_animais'],fin['medicamentos'],fin['nutricao'],
                  fin['operacional'],fin['custos_fixos'],fin['receita_total']],
        "Tipo": ["Saída"]*5+["Entrada"]})
    fig=px.bar(df_fluxo,x="Categoria",y="Valor",color="Tipo",
        color_discrete_map={"Saída":c["perigo"],"Entrada":c["primaria"]})
    fig.update_layout(**PLOTLY,height=300,xaxis=dict(gridcolor=c["superficie"]),yaxis=dict(gridcolor=c["superficie"]))
    st.plotly_chart(fig,use_container_width=True)


def _fin_dre():
    """DRE gerencial (§5, Trilha 3) — `services.dre.montar_dre` sobre o mesmo
    resumo do período que "📒 Resultado" já usa.

    Diferente do Resultado (Caixa): o CPV aqui é o custo do animal **casado
    com a venda** (`cost_at_sale`), não o que foi gasto comprando/mantendo
    animais no período. Comprar um lote em janeiro e vendê-lo só em julho não
    vira despesa da DRE de janeiro — o dinheiro saiu, mas o valor virou
    rebanho (patrimônio), não custo do período.
    """
    st.subheader("📈 DRE Gerencial")
    st.caption("Receita casada com o custo do animal na hora da venda (competência), "
               "não com o que foi gasto comprando no período — essa é a diferença para "
               "a aba **Resultado**, que é caixa. Ver o porquê no código "
               "(`services/dre.py::montar_dre`).")
    c1, c2 = st.columns(2)
    with c1: start = st.date_input("De", value=date(date.today().year,1,1), key="dre_start")
    with c2: end = st.date_input("Até", value=date.today(), key="dre_end")

    fin = db.get_financial_summary(start.isoformat(), end.isoformat())
    dre = montar_dre(fin)

    def _linha(rotulo, valor, negrito=False, indent=False):
        cor_valor = c["perigo"] if valor < 0 else c["texto"]
        peso = "700" if negrito else "400"
        rec = f"padding-left:{'1.5rem' if indent else '0'}"
        st.markdown(
            f"<div style='display:flex;justify-content:space-between;{rec};"
            f"font-weight:{peso}'><span>{rotulo}</span>"
            f"<span style='color:{cor_valor}'>R$ {valor:,.2f}</span></div>",
            unsafe_allow_html=True)

    st.markdown("<div class='card'>", unsafe_allow_html=True)
    _linha("Receita Bruta de Vendas", dre["receita_bruta"], negrito=True)
    _linha("(–) CPV — Custo dos Animais Vendidos", -dre["cpv"], indent=True)
    st.markdown("<hr style='margin:.3rem 0'>", unsafe_allow_html=True)
    _linha("= Lucro Bruto", dre["lucro_bruto"], negrito=True)
    if dre["margem_bruta_pct"] is not None:
        st.caption(f"Margem bruta: {dre['margem_bruta_pct']:.1f}%")
    for rotulo, valor in dre["despesas_operacionais"].items():
        _linha(f"(–) {rotulo}", -valor, indent=True)
    st.markdown("<hr style='margin:.3rem 0'>", unsafe_allow_html=True)
    _linha("= Resultado Operacional", dre["resultado_operacional"], negrito=True)
    if dre["perda_mortalidade"] > 0:
        _linha("(–) Perda por Mortalidade", -dre["perda_mortalidade"], indent=True)
    st.markdown("<hr style='margin:.3rem 0'>", unsafe_allow_html=True)
    _linha("= Resultado Líquido do Período", dre["resultado_liquido"], negrito=True)
    st.markdown("</div>", unsafe_allow_html=True)

    if dre["margem_liquida_pct"] is not None:
        st.metric("Margem Líquida", f"{dre['margem_liquida_pct']:.1f}%")

    with st.expander("Por que o CPV não é igual à 'Compra de animais' do Resultado (Caixa)?"):
        st.markdown(
            "Um animal comprado neste período mas ainda ativo no rebanho **não** entra "
            "na DRE como despesa — o valor pago por ele continua no patrimônio (é um "
            "boi no pasto, não um custo do período). Só quando ele é **vendido** (o "
            "custo acumulado dele vira CPV, casado com a receita da venda) ou **morre** "
            "(vira perda) é que o valor sai do 'estoque' e afeta o resultado. "
            "Medicamentos, nutrição e custos fixos continuam como despesa operacional "
            "do período — o sistema ainda não aloca esses custos por animal individual.")


def _fin_mortalidade():
    """Taxas de mortalidade: geral, por causa e por piquete."""
    st.subheader("☠️ Mortalidade")
    c1, c2 = st.columns(2)
    with c1: start = st.date_input("De", value=date(date.today().year,1,1), key="mort_start")
    with c2: end = st.date_input("Até", value=date.today(), key="mort_end")
    m = db.get_mortality_stats(start.isoformat(), end.isoformat())

    k = st.columns(4)
    k[0].metric("Óbitos no período", m["n_deaths"])
    k[1].metric("Taxa de mortalidade", f"{m['taxa_geral']:.1f}%",
                help=f"{m['n_deaths']} mortes ÷ {m['expostos']} animais que entraram até a data")
    k[2].metric("Animais expostos", m["expostos"])
    k[3].metric("Perda financeira", f"R$ {m['perda_total']:,.2f}")

    if m["n_deaths"] == 0:
        st.success("✅ Nenhum óbito registrado no período."); return

    cc1, cc2 = st.columns(2)
    with cc1:
        st.markdown("**Por causa**")
        dfc = pd.DataFrame([{"Causa":k2,"Mortes":v} for k2,v in m["por_causa"].items()]).sort_values("Mortes",ascending=False)
        figc = px.pie(dfc, names="Causa", values="Mortes", hole=0.45,
            color_discrete_sequence=[c["perigo"], c["atencao_secundario"], c["atencao_brilhante"], c["destaque"], c["info"], c["sucesso"], c["destaque_secundario"]])
        figc.update_layout(**_layout(height=260, margin=dict(l=0,r=0,t=10,b=10),
            legend=dict(orientation="h",yanchor="bottom",y=-0.35)))
        st.plotly_chart(figc, use_container_width=True)
    with cc2:
        st.markdown("**Por piquete**")
        dfl = pd.DataFrame([{"Piquete":k2,"Mortes":v} for k2,v in m["por_lote"].items()]).sort_values("Mortes",ascending=True)
        figl = px.bar(dfl, x="Mortes", y="Piquete", orientation="h",
            color="Mortes", color_continuous_scale=[c["atencao"], c["perigo"]])
        figl.update_layout(**_layout(height=260, coloraxis_showscale=False,
            xaxis=dict(gridcolor=c["superficie"]), yaxis=dict(gridcolor=c["superficie"],title="")))
        st.plotly_chart(figl, use_container_width=True)

    st.markdown("**📜 Registro de óbitos**")
    deaths = db.get_deaths(start.isoformat(), end.isoformat())
    df_d = pd.DataFrame(deaths)[["death_date","animal_id","breed","cause","lote_name",
                                 "weight_at_death","cost_at_death","operator"]].copy()
    df_d.columns=["Data","Animal","Raça","Causa","Piquete","Peso (kg)","Perda (R$)","Registrado por"]
    st.dataframe(df_d, use_container_width=True, hide_index=True,
        column_config={"Peso (kg)":st.column_config.NumberColumn(format="%.1f"),
                       "Perda (R$)":st.column_config.NumberColumn(format="R$ %.2f")})


def page_financeiro():
    st.markdown('<div class="page-title">💰 Financeiro & Mercado</div>', unsafe_allow_html=True)
    animals=db.get_all_animals()
    lotes=db.get_all_lotes()

    (t_res,t_dre,t_comp,t_fx,t_cc,t_pag,t_ven,t_rec,t_pre,t_mort,ft1,ft_fix,ft2,ft3,ft4,ft5,
     ft6)=st.tabs(
        ["📒 Resultado","📈 DRE Gerencial","📅 Competência","💵 Fluxo de Caixa",
         "🏭 Centros de Custo","📋 Contas a Pagar","💵 Registrar Venda","📥 Contas a Receber",
         "🏷️ Preços/Categoria","☠️ Mortalidade","📊 Custos por Animal","🏢 Custos Fixos",
         "💹 Simulador","⚖️ Breakeven","🏆 Origem","🐄 Por Raça","➗ Rateio de Lote"])

    with t_res: _fin_resultado()
    with t_dre: _fin_dre()
    with t_comp: _fin_competencia()
    with t_fx: _fin_fluxo_de_caixa()
    with t_cc: _fin_centros_de_custo(lotes)
    with t_pag: _fin_contas_a_pagar()
    with t_ven: _fin_venda(animals)
    with t_rec: _fin_contas_a_receber()
    with t_pre: _fin_precos()
    with t_mort: _fin_mortalidade()

    # "Por raça" olha para ciclos ENCERRADOS (vendas), não para o rebanho
    # ativo — por isso fica fora do guard abaixo: rebanho ativo zerado não
    # significa que não há histórico de venda para ranquear.
    with ft5: _fin_rentabilidade_por_raca()

    if not animals:
        for t in (ft1,ft_fix,ft2,ft3,ft4,ft6):
            with t: st.info("Sem animais ativos para esta análise.")
        return

    with ft1: _fin_custos_por_animal(animals)
    with ft_fix: _fin_custos_fixos(animals, lotes)
    with ft2: _fin_simulador(animals)
    with ft3: _fin_breakeven(animals)
    with ft4: _fin_desempenho_origem()

    with ft6:
        _fin_rateio_de_lote(animals)



def _fin_custos_por_animal(animals):
    ul   = _unit_label()
    rows_f=[]
    costs = db._costs_by_animal()
    for a in animals:
        tc    = costs.get(a["id"], 0.0)
        yield_= a.get("carcass_yield") or 0.52
        prod  = _live_weight(a["current_weight"], yield_)
        gain  = a["current_weight"] - a["entry_weight"]
        prod_g= _prod_weight(gain, yield_) if gain > 0 else 0
        cpu   = round(tc/prod, 2) if prod else 0
        cpu_g = round(tc/prod_g, 2) if prod_g > 0 else 0
        rows_f.append({"ID":a["id"],"Raça":a["breed"],
            "Peso (kg)":a["current_weight"],
            f"Prod. ({ul})":prod,
            "Custo Total (R$)":tc,
            _cost_per_unit_label():cpu,
            f"Ganho ({ul})":prod_g,
            f"Custo Produção/{ul}":cpu_g})
    df_f=pd.DataFrame(rows_f)

    tot_tc  = df_f["Custo Total (R$)"].sum()
    tot_prod= df_f[f"Prod. ({ul})"].sum()
    tot_gnh = df_f[f"Ganho ({ul})"].sum()
    kk=st.columns(4)
    kk[0].metric("Custo Total do Rebanho", f"R$ {tot_tc:,.2f}")
    kk[1].metric(f"Total {ul} no Rebanho", f"{tot_prod:.1f} {ul}")
    kk[2].metric(f"Total {ul} Ganhos",     f"{tot_gnh:.1f} {ul}")
    kk[3].metric(_cost_per_unit_label(),    f"R$ {tot_tc/tot_prod:.2f}" if tot_prod else "—")

    prod_col = f"Prod. ({ul})"
    cpu_col  = _cost_per_unit_label()
    fmt_prod = "%.2f" if _use_arroba() else "%.1f"
    st.dataframe(df_f,use_container_width=True,hide_index=True,
        column_config={
            "Peso (kg)":st.column_config.NumberColumn(format="%.1f"),
            prod_col:st.column_config.NumberColumn(format=fmt_prod),
            "Custo Total (R$)":st.column_config.NumberColumn(format="R$ %.2f"),
            cpu_col:st.column_config.NumberColumn(format="R$ %.2f"),
            f"Ganho ({ul})":st.column_config.NumberColumn(format=fmt_prod),
            f"Custo Produção/{ul}":st.column_config.NumberColumn(format="R$ %.2f")})

    fig_c=px.scatter(df_f,x=prod_col,y="Custo Total (R$)",
        color=cpu_col,text="ID",
        color_continuous_scale=ESCALA_BOM_RUIM,
        labels={prod_col:f"Produção ({ul})","Custo Total (R$)":"Custo Total (R$)"})
    fig_c.update_traces(textposition="top center",marker=dict(size=12))
    fig_c.update_layout(**PLOTLY,height=320,coloraxis_colorbar=dict(title=f"R$/{ul}"))
    st.plotly_chart(fig_c,use_container_width=True)



def _fin_custos_fixos(animals, lotes):
    st.subheader("🏢 Custos Fixos da Fazenda")
    st.caption("Aluguel de pastagem, salários, bonificações, impostos, taxas e outros "
               "custos que não são atribuídos a um animal específico.")

    # Filtro por período
    pc1,pc2=st.columns(2)
    with pc1:
        start_f=st.date_input("De", value=date(date.today().year,1,1), key="fix_start")
    with pc2:
        end_f=st.date_input("Até", value=date.today(), key="fix_end")
    s_iso, e_iso = start_f.isoformat(), end_f.isoformat()

    fixed=db.get_fixed_costs(s_iso, e_iso)
    total_fix=db.get_total_fixed_costs(s_iso, e_iso)
    by_cat=db.get_fixed_costs_by_category(s_iso, e_iso)

    mk=st.columns(3)
    mk[0].metric("Total de Custos Fixos", f"R$ {total_fix:,.2f}")
    mk[1].metric("Lançamentos", len(fixed))
    n_animals=len(animals)
    mk[2].metric("Rateio por Animal Ativo",
                 f"R$ {total_fix/n_animals:,.2f}" if n_animals else "—",
                 help="Custo fixo dividido igualmente pelos animais ativos")

    # Formulário de lançamento
    with st.expander("➕ Lançar Custo Fixo", expanded=not fixed):
        with st.form("f_fixed",clear_on_submit=True):
            fx1,fx2=st.columns(2)
            with fx1:
                fx_cat=st.selectbox("Categoria *", db.FIXED_COST_CATEGORIES)
                fx_amount=st.number_input("Valor (R$) *", min_value=0.0, step=50.0, format="%.2f")
            with fx2:
                fx_date=st.date_input("Data *", value=date.today())
                fx_recur=st.checkbox("Custo recorrente (mensal)")
            fx_cc=st.selectbox("Centro de Custo", [None]+[l["id"] for l in lotes],
                format_func=lambda lid: "🏭 Geral da Fazenda" if lid is None
                    else next((f"🌿 {l['id']} — {l['name']}" for l in lotes if l["id"]==lid), lid),
                help="Piquete a que este custo pertence. 'Geral da Fazenda' para o que "
                     "não é de um piquete específico (salário, contabilidade).")
            fx_desc=st.text_input("Descrição", placeholder="Ex: Aluguel piquete Norte / Salário João")
            if st.form_submit_button("✅ Lançar Custo Fixo", type="primary", use_container_width=True):
                if fx_amount<=0:
                    st.error("O valor deve ser maior que zero.")
                else:
                    db.add_fixed_cost(fx_cat, fx_desc, fx_amount,
                                      fx_date.strftime("%Y-%m-%d"), fx_recur, "",
                                      lote_id=fx_cc)
                    st.success(f"✅ {fx_cat}: R$ {fx_amount:,.2f} lançado!")
                    st.rerun()

    if fixed:
        # Gráfico por categoria
        cga,cgb=st.columns([2,3])
        with cga:
            df_cat=pd.DataFrame(by_cat)
            df_cat.columns=["Categoria","Total"]
            fig_fx=px.pie(df_cat,names="Categoria",values="Total",hole=0.45,
                color_discrete_sequence=SERIES + [c["perigo"]])
            fig_fx.update_layout(**_layout(height=260,margin=dict(l=0,r=0,t=10,b=10),
                legend=dict(orientation="h",yanchor="bottom",y=-0.25)))
            fig_fx.update_traces(textposition="inside",textinfo="percent")
            st.plotly_chart(fig_fx,use_container_width=True)
        with cgb:
            nomes_lote={l["id"]:l["name"] for l in lotes}
            df_fx=pd.DataFrame(fixed)[["cost_date","category","description","amount",
                                       "recurring","lote_id"]].copy()
            df_fx["recurring"]=df_fx["recurring"].map({1:"Mensal",0:"Único"})
            df_fx["lote_id"]=df_fx["lote_id"].map(
                lambda lid: "Geral" if not lid else nomes_lote.get(lid, lid))
            df_fx.columns=["Data","Categoria","Descrição","Valor (R$)","Tipo","Centro de Custo"]
            st.dataframe(df_fx,use_container_width=True,hide_index=True,height=260,
                column_config={"Valor (R$)":st.column_config.NumberColumn(format="R$ %.2f")})

        # Excluir lançamento
        with st.expander("🗑️ Excluir um lançamento"):
            opt={f"#{f['id']} · {f['cost_date']} · {f['category']} · R$ {f['amount']:,.2f}":f["id"] for f in fixed}
            sel_del=st.selectbox("Lançamento", list(opt.keys()), key="del_fix")
            if st.button("Excluir", type="secondary"):
                db.delete_fixed_cost(opt[sel_del])
                st.success("Lançamento excluído."); st.rerun()
    else:
        st.info("Nenhum custo fixo lançado no período selecionado.")



def _fin_simulador(animals):
    ul = _unit_label()
    arroba_mode = _use_arroba()
    st.subheader("💵 Simulador de Venda")

    precos_cat = db.get_category_prices()
    base = st.radio("Base de preço",
        ["categoria","manual"],
        format_func=lambda b: "🏷️ Tabela de preços por categoria" if b=="categoria"
                              else "✏️ Preço único manual",
        horizontal=True, key="sim_base")

    cotacao = 0.0
    rendimento = 52
    ajuste_pct = 0
    if base == "manual":
        sc1, sc2 = st.columns(2)
        with sc1:
            price_label = "Cotação por @ (R$)" if arroba_mode else "Cotação por kg de boi vivo (R$)"
            default_price = DEFAULT_PRICE_ARROBA if arroba_mode else DEFAULT_PRICE_KG
            cotacao=st.number_input(price_label, min_value=0.01, max_value=5000.0,
                value=default_price, step=(5.0 if arroba_mode else 0.10), format="%.2f")
            if arroba_mode:
                rendimento=st.slider("Rendimento de Carcaça (%)",40,65,52)
        with sc2:
            sub = ("Rendimento: "+str(rendimento)+"%") if arroba_mode else "Peso vivo (sem desconto de carcaça)"
            st.markdown(
                f'<div class="card"><div style="color:{c["texto_secundario"]};font-size:.85rem">Cotação única</div>'
                f'<div style="font-size:2rem;font-weight:800;color:{c["primaria"]}">R$ {cotacao:.2f}/{ul}</div>'
                f'<div style="color:{c["texto_secundario"]};font-size:.85rem;margin-top:.5rem">{sub}</div></div>',
                unsafe_allow_html=True)
    else:  # categoria
        st.caption("Cada animal é avaliado pelo **R$/kg da sua categoria** (definido em "
                   "**Preços/Categoria**). Use o ajuste abaixo para simular alta/baixa de mercado.")
        if not precos_cat or all(v <= 0 for v in precos_cat.values()):
            st.warning("⚠️ Nenhum preço por categoria definido. Vá em **Preços/Categoria** "
                       "e informe os valores por kg de cada categoria.")
        ajuste_pct = st.slider("Ajuste global de preço (%)", -30, 30, 0,
            help="Ex: mercado subiu 5% → +5. Aplica sobre todos os preços da tabela.")
        # Mostra os preços em uso
        linhas = []
        for band in AGE_BANDS:
            for sex in ("M","F"):
                p = precos_cat.get((band,sex),0.0) * (1+ajuste_pct/100)
                if p > 0:
                    linhas.append(f"{band} · {'♂' if sex=='M' else '♀'}: R$ {p:.2f}/kg")
        if linhas:
            st.caption("Preços aplicados: " + "  |  ".join(linhas))

    incluir_fixos=st.checkbox("Incluir rateio de custos fixos no cálculo",
        help="Divide os custos fixos do ano igualmente entre os animais ativos")

    rateio_fixo = 0.0
    if incluir_fixos and animals:
        total_fix_ano = db.get_total_fixed_costs(
            date(date.today().year,1,1).isoformat(), date.today().isoformat())
        rateio_fixo = total_fix_ano / len(animals)
        st.info(f"Rateio de custos fixos: **R\\$ {rateio_fixo:,.2f}** por animal "
                f"(total R\\$ {total_fix_ano:,.2f} ÷ {len(animals)} animais ativos).")

    sim_rows=[]
    sem_preco=[]
    costs = db._costs_by_animal()
    for a in animals:
        tc  = costs.get(a["id"], 0.0) + rateio_fixo
        band = db.get_age_category(a.get("birth_date"))
        if base == "categoria":
            price_kg = precos_cat.get((band, a["sex"]), 0.0) * (1+ajuste_pct/100)
            receita  = round(a["current_weight"] * price_kg, 2)
            preco_aplicado = round(price_kg, 2)
            if price_kg <= 0: sem_preco.append(a["id"])
        else:
            prod    = _live_weight(a["current_weight"], rendimento/100)
            receita = round(prod * cotacao, 2)
            preco_aplicado = round(cotacao, 2)
        prod_disp = _live_weight(a["current_weight"], rendimento/100)
        lucro = round(receita - tc, 2)
        sim_rows.append({"ID":a["id"],"Categoria":band,
            "Peso (kg)":a["current_weight"], f"Venda ({ul})":prod_disp,
            "Preço (R$/kg)":preco_aplicado if base=="categoria" else None,
            "Receita (R$)":receita,"Custo Total (R$)":round(tc,2),"Lucro (R$)":lucro,
            "Margem (%)":round(lucro/receita*100,1) if receita else 0})
    df_sim=pd.DataFrame(sim_rows)
    if base != "categoria":
        df_sim = df_sim.drop(columns=["Preço (R$/kg)"])

    if sem_preco:
        st.warning(f"⚠️ Sem preço de categoria (receita R$ 0): **{', '.join(sem_preco)}**. "
                   f"Defina os valores em **Preços/Categoria**.")

    tot_rec=df_sim["Receita (R$)"].sum(); tot_luc=df_sim["Lucro (R$)"].sum()
    tot_cost=df_sim["Custo Total (R$)"].sum()
    margem_media=df_sim["Margem (%)"].mean()

    sk=st.columns(4)
    sk[0].metric("Receita Total", f"R$ {tot_rec:,.2f}")
    sk[1].metric("Custo Total",   f"R$ {tot_cost:,.2f}")
    sk[2].metric("Lucro / Prejuízo Total", f"R$ {tot_luc:,.2f}",
        delta=f"{tot_luc:+,.2f}", delta_color="normal")
    sk[3].metric("Margem Média", f"{margem_media:.1f}%",
        delta=f"{margem_media:+.1f}%", delta_color="normal")

    if tot_luc < 0:
        st.error(f"⚠️ Projeção de **PREJUÍZO** de R$ {abs(tot_luc):,.2f}. "
                 f"Reveja os preços, os custos ou o ponto de venda.")

    fmt_prod = "%.2f" if arroba_mode else "%.1f"
    cfg = {"Peso (kg)":st.column_config.NumberColumn(format="%.1f"),
           f"Venda ({ul})":st.column_config.NumberColumn(format=fmt_prod),
           "Receita (R$)":st.column_config.NumberColumn(format="R$ %.2f"),
           "Custo Total (R$)":st.column_config.NumberColumn(format="R$ %.2f"),
           "Lucro (R$)":st.column_config.NumberColumn(format="R$ %.2f")}
    if base == "categoria":
        cfg["Preço (R$/kg)"]=st.column_config.NumberColumn(format="R$ %.2f")
    st.dataframe(df_sim,use_container_width=True,hide_index=True,column_config=cfg)



def _fin_breakeven(animals):
    ul = _unit_label()
    st.subheader("⚖️ Ponto de Equilíbrio (Breakeven)")
    be_rows=[]
    costs = db._costs_by_animal()
    for a in animals:
        tc   = costs.get(a["id"], 0.0)
        prod = _live_weight(a["current_weight"], a.get("carcass_yield") or 0.52)
        be   = round(tc/prod, 2) if prod else 0
        be_rows.append({"ID":a["id"],"Raça":a["breed"],
            "Peso (kg)":a["current_weight"],
            f"Prod. ({ul})":prod,
            "Custo Total (R$)":tc,
            _breakeven_label():be})
    df_be    = pd.DataFrame(be_rows)
    be_col   = _breakeven_label()
    prod_col = f"Prod. ({ul})"
    fmt_prod = "%.2f" if _use_arroba() else "%.1f"
    fig_be=px.bar(df_be.sort_values(be_col),
        x="ID",y=be_col,color=be_col,
        color_continuous_scale=ESCALA_BOM_RUIM,
        labels={be_col:f"R$ mínimo por {ul}"})
    fig_be.update_layout(**PLOTLY,height=300,coloraxis_showscale=False,
        xaxis=dict(gridcolor=c["superficie"]),yaxis=dict(gridcolor=c["superficie"]))
    st.plotly_chart(fig_be,use_container_width=True)
    st.dataframe(df_be,use_container_width=True,hide_index=True,
        column_config={"Peso (kg)":st.column_config.NumberColumn(format="%.1f"),
            prod_col:st.column_config.NumberColumn(format=fmt_prod),
            "Custo Total (R$)":st.column_config.NumberColumn(format="R$ %.2f"),
            be_col:st.column_config.NumberColumn(format="R$ %.2f")})



def _fin_desempenho_origem():
    st.subheader("🏆 Ranking de Fornecedor / Origem")
    st.caption("Comparativo por origem sobre **todo o histórico** (ativos, vendidos e mortos): "
               "quem entrega o melhor **GMD**, a menor **mortalidade** e o menor "
               "**custo por @ produzida**.")
    rank=db.get_fornecedor_ranking()
    if not rank:
        st.info("Sem animais com fornecedor informado ainda.")
    else:
        rows=[{"Fornecedor":r["fornecedor"],"Animais":r["n"],
               "Ativos":r["ativos"],"Vendidos":r["vendidos"],"Mortos":r["mortos"],
               "GMD Médio (kg/dia)":r["gmd_medio"],
               "Mortalidade (%)":r["taxa_mortalidade"],
               "@ produzidas":r["arrobas_produzidas"],
               "Custo/@ produzida (R$)":r["custo_por_arroba"]} for r in rank]
        df_p=pd.DataFrame(rows)
        st.dataframe(df_p,use_container_width=True,hide_index=True,
            column_config={
                "GMD Médio (kg/dia)":st.column_config.NumberColumn(format="%.3f"),
                "Mortalidade (%)":st.column_config.NumberColumn(format="%.1f%%"),
                "@ produzidas":st.column_config.NumberColumn(format="%.2f"),
                "Custo/@ produzida (R$)":st.column_config.NumberColumn(format="R$ %.2f")})

        c1,c2=st.columns(2)
        with c1:
            fig_p=px.bar(df_p,x="Fornecedor",y="GMD Médio (kg/dia)",
                color="GMD Médio (kg/dia)",
                color_continuous_scale=ESCALA_RUIM_BOM,
                text="GMD Médio (kg/dia)")
            fig_p.update_traces(texttemplate="%{text:.3f}",textposition="outside")
            fig_p.update_layout(**PLOTLY,height=300,coloraxis_showscale=False,
                title="GMD médio por fornecedor",
                xaxis=dict(gridcolor=c["superficie"]),yaxis=dict(gridcolor=c["superficie"]))
            st.plotly_chart(fig_p,use_container_width=True)
        with c2:
            fig_m=px.bar(df_p,x="Fornecedor",y="Mortalidade (%)",
                color="Mortalidade (%)",
                color_continuous_scale=ESCALA_BOM_RUIM,
                text="Mortalidade (%)")
            fig_m.update_traces(texttemplate="%{text:.1f}%",textposition="outside")
            fig_m.update_layout(**PLOTLY,height=300,coloraxis_showscale=False,
                title="Taxa de mortalidade por fornecedor",
                xaxis=dict(gridcolor=c["superficie"]),yaxis=dict(gridcolor=c["superficie"]))
            st.plotly_chart(fig_m,use_container_width=True)

        melhor=next((r for r in rank if r["arrobas_produzidas"]>0),None)
        if melhor:
            st.success(f"🥇 Melhor GMD médio: **{rank[0]['fornecedor']}** "
                       f"({rank[0]['gmd_medio']:.3f} kg/dia). "
                       f"Compare com o **custo por @** e a **mortalidade** na tabela para "
                       f"decidir de quem vale a pena comprar de novo.")



def _fin_rateio_de_lote(animals):
    """Rateio de custo de lote entre os animais (spec 0019 + adaptador 0041).

    `services.rateio.ratear` existia pura e testada desde a spec 0019, mas
    sem consumidor — e a lacuna que a motivou continuava aberta: **não
    existe hoje nenhum jeito de lançar um custo para "o lote inteiro"**
    (trato coletivo, medicamento aplicado a todos, frete). Só dá pra
    lançar custo por um animal de cada vez (Ficha do Animal → Adicionar
    Custo). Sem isso, custo de lote simplesmente não entra no individual, e
    o custo por arroba de cada animal fica subestimado — exatamente o
    problema que a spec 0019 documentou.

    `dias_no_lote` (critério `peso_dia`) não vem de lugar nenhum pronto —
    resolvido aqui (R31, fora do escopo da spec 0041): a movimentação mais
    recente do animal para este piquete, ou `entry_date` se ele nunca se
    moveu (está no piquete desde que entrou no rebanho).
    """
    st.caption("Lança um custo único (trato do lote, medicamento aplicado a todos, "
               "frete...) dividido entre os animais do piquete — proporcionalmente, "
               "não em partes iguais que ignoram peso.")

    lotes = db.get_all_lotes()
    if not lotes:
        st.info("Cadastre piquetes primeiro (em Lotes / Pastagem).")
        return

    lote_sel = st.selectbox("Piquete *", lotes,
        format_func=lambda l: f"{l['id']} — {l['name']}", key="rat_lote")
    animais_lote = [a for a in animals if a.get("lote_id") == lote_sel["id"]]
    if not animais_lote:
        st.info("Nenhum animal ativo neste piquete.")
        return

    c1, c2, c3 = st.columns(3)
    with c1:
        valor_total = st.number_input("Valor total (R$) *", min_value=0.0, step=10.0,
                                      format="%.2f", key="rat_valor")
    with c2:
        criterio = st.selectbox("Critério", ["peso_dia", "peso", "igual"],
            format_func=lambda k: {"peso_dia": "Peso × dias no piquete",
                                   "peso": "Peso atual",
                                   "igual": "Igual para todos"}[k], key="rat_criterio")
    with c3:
        tipo = st.selectbox("Tipo de custo", COST_TYPES, key="rat_tipo")
    desc = st.text_input("Descrição *", placeholder="Ex: Vermífugo aplicado no lote",
                         key="rat_desc").strip()
    data_custo = st.date_input("Data", value=date.today(), key="rat_data")
    referencia = data_custo.isoformat()

    animais_para_ratear = [{"id": a["id"], "peso": a["current_weight"]}
                           for a in animais_lote]
    if criterio == "peso_dia":
        animal_ids = [a["id"] for a in animais_lote]
        last_movements = db.get_last_movements_bulk(animal_ids)
        for item, a in zip(animais_para_ratear, animais_lote):
            mov_date = last_movements.get(a["id"])
            item["entrada_no_lote"] = mov_date if mov_date else a["entry_date"]
        animais_para_ratear = com_dias_no_lote(animais_para_ratear, referencia)

    if not valor_total:
        st.info("Informe o valor total para ver a prévia do rateio.")
        return

    preview = ratear(valor_total, animais_para_ratear, criterio)
    st.markdown(f"**Prévia do rateio entre {len(preview)} animal(is)**")
    df_prev = pd.DataFrame(preview).rename(
        columns={"animal_id": "Animal", "valor": "Valor (R$)"})
    st.dataframe(df_prev, use_container_width=True, hide_index=True,
        column_config={"Valor (R$)": st.column_config.NumberColumn(format="R$ %.2f")})
    soma = sum(p["valor"] for p in preview)
    st.caption(f"Soma do rateio: R$ {soma:,.2f} (fecha exatamente com o valor total).")

    if st.button(f"💾 Lançar custo rateado para {len(preview)} animal(is)",
                 type="primary", disabled=not desc, key="rat_salvar"):
        for item in preview:
            db.add_animal_cost(item["animal_id"], tipo, desc, item["valor"], referencia,
                               notes=f"Rateio do piquete {lote_sel['id']} ({criterio}), "
                                     f"total R$ {valor_total:,.2f}")
        db.clear_cache()
        st.success(f"✅ R$ {valor_total:,.2f} rateado entre {len(preview)} animal(is).")
        st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# ESTOQUE DE INSUMOS
# ══════════════════════════════════════════════════════════════════════════════
def page_estoque():
    st.markdown('<div class="page-title">📦 Estoque de Insumos</div>', unsafe_allow_html=True)
    insumos=db.get_all_insumos()
    low=db.check_low_stock()
    if low:
        st.warning(f"⚠️ **{_plural(len(low),'insumo','insumos')} abaixo do estoque mínimo:** " +
            ", ".join(f"**{i['name']}** ({_num_br(i['current_stock'],0)} {i['unit']})" for i in low))

    et1,et2,et3,et4,et5=st.tabs(["📋 Inventário","📥 Entrada de Estoque","➕ Novo Insumo",
        "📈 Previsão de Ruptura","🛒 Compra com Nota Fiscal"])

    CAT_LABELS={"racao":"Ração","trato":"Trato (volumoso)","medicamento":"Medicamento",
                "vacina":"Vacina","mineral":"Mineral","outro":"Outro"}
    with et1:
        cats_present=sorted({i["category"] for i in insumos})
        fcat_ins=st.selectbox("Filtrar por categoria",
            ["Todas"]+cats_present,
            format_func=lambda c:c if c=="Todas" else CAT_LABELS.get(c,c))
        rows_i=[]
        for i in insumos:
            if fcat_ins!="Todas" and i["category"]!=fcat_ins: continue
            pct=i["current_stock"]/i["min_stock"]*100 if i["min_stock"] else 100
            rows_i.append({"Insumo":i["name"],"Categoria":CAT_LABELS.get(i["category"],i["category"]),
                "Estoque":i["current_stock"],"Unidade":i["unit"],
                "Mínimo":i["min_stock"],
                "Status":"🔴 Crítico" if pct<50 else "🟡 Baixo" if pct<100 else "🟢 OK",
                "Custo/Un (R$)":i["cost_per_unit"],
                "Valor Total (R$)":round(i["current_stock"]*i["cost_per_unit"],2)})
        df_i=pd.DataFrame(rows_i)
        st.dataframe(df_i,use_container_width=True,hide_index=True,
            column_config={"Estoque":st.column_config.NumberColumn(format="%.1f"),
                "Custo/Un (R$)":st.column_config.NumberColumn(format="R$ %.2f"),
                "Valor Total (R$)":st.column_config.NumberColumn(format="R$ %.2f")})
        tot_val=sum(r["Valor Total (R$)"] for r in rows_i)
        st.metric("Valor Total do Estoque",f"R$ {tot_val:,.2f}")

        # Gráfico % do mínimo
        df_bar=pd.DataFrame([{"Insumo":r["Insumo"],
            "% do Mínimo":min(r["Estoque"]/max(r["Mínimo"],0.01)*100,200)} for r in rows_i])
        fig_e=px.bar(df_bar.sort_values("% do Mínimo"),x="% do Mínimo",y="Insumo",
            orientation="h",color="% do Mínimo",
            color_continuous_scale=ESCALA_RUIM_BOM,range_color=[0,200])
        fig_e.add_vline(x=100,line_dash="dash",line_color=c["atencao"],
            annotation_text="Mínimo",annotation_position="top")
        fig_e.update_layout(**PLOTLY,height=280,coloraxis_showscale=False,
            xaxis=dict(gridcolor=c["superficie"],title="% do Estoque Mínimo"),
            yaxis=dict(gridcolor=c["superficie"],title=""))
        st.plotly_chart(fig_e,use_container_width=True)

    with et2:
        with st.form("f_entrada",clear_on_submit=True):
            ins=st.selectbox("Insumo",insumos,format_func=lambda x:f"{x['name']} ({x['current_stock']:.1f} {x['unit']})")
            ec1,ec2=st.columns(2)
            with ec1: qty=st.number_input("Quantidade",min_value=0.01,step=1.0,format="%.2f")
            with ec2: cpu=st.number_input("Custo por Unidade (R$)",min_value=0.0,step=0.01,format="%.2f",
                value=float(ins["cost_per_unit"]) if ins else 0.0)
            if st.form_submit_button("✅ Registrar Entrada",type="primary",use_container_width=True):
                db.add_insumo_entry(ins["id"],qty,cpu,st.session_state.user["name"])
                st.success(f"✅ +{qty:.1f} {ins['unit']} de {ins['name']}"); st.rerun()

    with et3:
        with st.form("f_new_ins",clear_on_submit=True):
            ni1,ni2=st.columns(2)
            with ni1:
                ni_name=st.text_input("Nome *",placeholder="Ex: Silagem de milho / Massa de soja")
                ni_cat=st.selectbox("Categoria",
                    ["medicamento","vacina","racao","trato","mineral","outro"],
                    format_func=lambda c:{"racao":"ração","trato":"trato (volumoso)"}.get(c,c),
                    help="'trato' = volumosos como silagem, massa de soja, bagaço de laranja")
            with ni2:
                ni_unit=st.selectbox("Unidade",
                    ["kg","ton","saco","ml","mg","g","dose","litro","comprimido"],
                    format_func=lambda u:{"ton":"tonelada (ton)"}.get(u,u))
                ni_stk=st.number_input("Estoque Inicial",min_value=0.0,step=1.0,format="%.1f")
            ni_min=st.number_input("Estoque Mínimo (alerta)",min_value=0.0,step=1.0,format="%.1f")
            ni_cpu=st.number_input("Custo por Unidade (R$)",min_value=0.0,step=0.01,format="%.2f")
            if st.form_submit_button("✅ Criar Insumo",type="primary",use_container_width=True):
                if ni_name:
                    db.add_new_insumo(ni_name,ni_cat,ni_unit,ni_stk,ni_min,ni_cpu)
                    st.success(f"✅ Insumo {ni_name} criado!"); st.rerun()
                else:
                    st.error("Nome é obrigatório.")

    with et4:
        st.caption("Dias até faltar cada insumo, no ritmo de consumo **planejado** "
            "(soma dos planos de trato ativos) — diferente do aviso de "
            "'abaixo do mínimo' acima, que só olha o saldo de hoje.")
        previsao = _previsao_estoque()
        URGENCIA_LABEL = {"critica": "🔴 Crítica", "atencao": "🟡 Atenção",
                          "ok": "🟢 OK", "sem_dados": "⚪ Sem dados"}
        criticos = [p for p in previsao if p["urgencia"] == "critica"]
        if criticos:
            st.warning(f"⚠️ **{_plural(len(criticos),'insumo','insumos')} em situação crítica:** " +
                ", ".join(f"**{p['nome']}**" for p in criticos))
        rows_p = [{
            "Insumo": p["nome"],
            "Dias Restantes": _num_br(p["dias_restantes"]) if p["dias_restantes"] is not None else "—",
            "Data de Ruptura": p["data_ruptura"] or "—",
            "Comprar Até": p["comprar_ate"] or "—",
            "Urgência": URGENCIA_LABEL.get(p["urgencia"], p["urgencia"]),
        } for p in previsao]
        st.dataframe(pd.DataFrame(rows_p),use_container_width=True,hide_index=True)
        st.caption("⚪ **Sem dados** = nenhum plano de trato ativo para o insumo, não é erro. "
            "**Comprar Até** hoje é igual à **Data de Ruptura** — o sistema ainda não guarda "
            "o prazo de reposição de cada insumo (fica para quando essa coluna existir).")

    with et5:
        _estoque_compra_com_nota(insumos)


def _estoque_compra_com_nota(insumos):
    """Trilha 3 (Estoque → Financeiro): compra com documento fiscal.

    Diferente da aba "Entrada de Estoque" (1 insumo por lançamento, sem
    fornecedor/documento/parcelamento), esta tela junta vários itens numa
    nota só e, ao registrar, grava estoque **e** contas a pagar na mesma
    operação (`repositories.compras.registrar` — ROADMAP §5, "Pronto quando").
    """
    st.caption("Lança vários insumos de uma nota só, aplica custo médio ponderado "
               "(mesma regra da Entrada de Estoque, ADR 0003) e já gera as "
               "parcelas em Contas a Pagar (aba 📋 em Financeiro).")

    if not insumos:
        st.info("Cadastre um insumo antes (aba ➕ Novo Insumo).")
        return

    itens_key = "compra_itens_atual"
    st.session_state.setdefault(itens_key, [])

    st.markdown("**1. Itens da nota**")
    ic1, ic2, ic3, ic4 = st.columns([3, 1, 1, 1])
    with ic1:
        ins_add = st.selectbox("Insumo", insumos,
            format_func=lambda x: f"{x['name']} ({x['unit']})",
            key="compra_add_insumo")
    with ic2:
        qtd_add = st.number_input("Quantidade", min_value=0.0, step=1.0,
            format="%.2f", key="compra_add_qtd")
    with ic3:
        custo_add = st.number_input("Custo Un. (R$)", min_value=0.0, step=0.01,
            format="%.2f", key="compra_add_custo",
            value=float(ins_add["cost_per_unit"]) if ins_add else 0.0)
    with ic4:
        st.write("")
        st.write("")
        if st.button("➕ Adicionar", key="compra_add_btn", use_container_width=True):
            if ins_add and qtd_add > 0:
                st.session_state[itens_key].append({
                    "insumo_id": ins_add["id"], "insumo_nome": ins_add["name"],
                    "unidade": ins_add["unit"], "quantidade": qtd_add,
                    "custo_unitario": custo_add})
                st.rerun()
            else:
                st.error("Escolha um insumo e uma quantidade maior que zero.")

    itens_atuais = st.session_state[itens_key]
    if itens_atuais:
        for idx, item in enumerate(itens_atuais):
            rc1, rc2 = st.columns([5, 1])
            with rc1:
                st.write(f"{item['insumo_nome']}: {item['quantidade']:.2f} "
                    f"{item['unidade']} × R$ {item['custo_unitario']:.2f} = "
                    f"R$ {item['quantidade']*item['custo_unitario']:,.2f}")
            with rc2:
                if st.button("🗑️", key=f"compra_remover_{idx}"):
                    itens_atuais.pop(idx)
                    st.rerun()
        total_nota = sum(i["quantidade"] * i["custo_unitario"] for i in itens_atuais)
        st.metric("Total da nota", f"R$ {total_nota:,.2f}")
    else:
        st.info("Nenhum item adicionado ainda — a nota precisa de ao menos 1 item.")

    st.markdown("**2. Documento e parcelamento**")
    fc1, fc2 = st.columns(2)
    with fc1:
        fornecedor_nome = st.text_input("Fornecedor", key="compra_fornecedor")
        doc_numero = st.text_input("Nº do documento fiscal", key="compra_doc_num")
        doc_serie = st.text_input("Série", key="compra_doc_serie")
    with fc2:
        data_emissao = st.date_input("Data de emissão", value=date.today(),
            key="compra_data_emissao")
        data_recebimento = st.date_input("Data de recebimento", value=date.today(),
            key="compra_data_recebimento")

    pc1, pc2 = st.columns(2)
    with pc1:
        num_parcelas = st.number_input("Número de parcelas", min_value=1, max_value=36,
            value=1, step=1, key="compra_num_parcelas")
    with pc2:
        primeira_parcela = st.date_input("Vencimento da 1ª parcela",
            value=date.today()+timedelta(days=30), key="compra_primeira_parcela")

    if st.button("✅ Registrar Compra", type="primary", use_container_width=True,
                 key="compra_registrar_btn", disabled=not itens_atuais):
        r = db.compras.registrar(
            data_emissao=data_emissao.isoformat(),
            data_recebimento=data_recebimento.isoformat(),
            itens=[{"insumo_id": i["insumo_id"], "quantidade": i["quantidade"],
                    "custo_unitario": i["custo_unitario"]} for i in itens_atuais],
            primeiro_vencimento=primeira_parcela.isoformat(),
            num_parcelas=int(num_parcelas),
            fornecedor_nome=fornecedor_nome, documento_numero=doc_numero,
            documento_serie=doc_serie, operator=st.session_state.user["name"])
        if r["ok"]:
            st.success(f"✅ Compra registrada — R$ {r['valor_total']:,.2f} em "
                f"{_plural(r['parcelas'],'parcela','parcelas')}. Estoque e contas a "
                f"pagar atualizados.")
            st.session_state[itens_key] = []
            st.rerun()
        else:
            st.error(f"❌ {r['erro']}")

# ══════════════════════════════════════════════════════════════════════════════
# ALERTAS
# ══════════════════════════════════════════════════════════════════════════════
def _consumo_diario_por_insumo() -> dict:
    """Consumo diário previsto de cada insumo, somando os planos de trato ativos.

    É consumo **planejado**, não realizado — que é justamente o que a regra de
    estoque precisa: ela pergunta se o saldo cobre os próximos dias de trato.

    Delega a `services.previsao_estoque_adaptador.consumo_diario_planejado`
    (spec 0039) em vez de calcular aqui. Antes a conta era feita inline, com
    `_FREQ_POR_DIA.get(p.get("frequency"), 1.0)` — frequência desconhecida
    (ex.: "quinzenal") caía no default `1.0` e virava consumo diário
    inventado (o defeito real da primeira tentativa da spec, PR 101).
    Ligar o adaptador aqui fecha o mesmo defeito neste consumidor: frequência
    fora de {"diario","semanal","mensal"} agora é ignorada, não vira 1×/dia.
    """
    insumos_por_id = {i["id"]: i for i in db.get_all_insumos()}
    planos_ativos = db.get_feeding_plans(active_only=True)
    return consumo_diario_planejado(insumos_por_id, planos_ativos, db.convert_quantity)


def _previsao_estoque() -> list[dict]:
    """Previsão de ruptura por insumo — dias restantes, data de ruptura, urgência.

    Liga `services/previsao_estoque.py::prever` (nunca chamado até aqui) através
    do adaptador da spec 0039. `prazo_reposicao_dias` não existe no schema ainda
    (fora do escopo daquela spec) — todo insumo entra com prazo 0, que `prever()`
    já trata como "desconhecido", não como erro.
    """
    insumos = db.get_all_insumos()
    consumo = _consumo_diario_por_insumo()
    montados = previsao_estoque_montar_insumos(insumos, consumo)
    return previsao_estoque_prever(montados, date.today().isoformat())


def _contexto_recomendacoes() -> dict:
    """Monta o retrato da fazenda que o motor de regras consome.

    O motor é função pura e não toca banco (R31) — quem apura é aqui.
    """
    consumo = _consumo_diario_por_insumo()
    hoje = date.today()

    animais = []
    animais_brutos = db.get_all_animals(status="ativo")
    a_ids = [a["id"] for a in animais_brutos]
    wd_batch = db.get_withdrawal_end_batch(a_ids)
    gmd_batch = db.calculate_gmd_bulk(a_ids)
    for a in animais_brutos:
        fim = wd_batch.get(a["id"])
        animais.append({
            "id": a["id"],
            "peso": a.get("current_weight"),
            "peso_alvo": a.get("target_weight"),
            "gmd": gmd_batch.get(a["id"]),
            "lote_id": a.get("lote_id"),
            "carencia_ate": fim.isoformat() if fim and fim >= hoje else None,
        })

    lotes = [{"id": l["id"],
              "capacidade_ua": l.get("capacity_ua"),
              "ua_atual": l.get("total_ua")}
             for l in db.get_all_lotes()]

    insumos = [{"id": i["id"], "nome": i["name"],
                "saldo": i.get("current_stock"),
                "consumo_diario": consumo.get(i["id"], 0.0)}
               for i in db.get_all_insumos()]

    return {
        "animais": animais,
        "lotes": lotes,
        "insumos": insumos,
        "preco_arroba": _to_float(db.get_setting("preco_arroba")),
        "custo_por_arroba": _custo_medio_por_arroba(),
        "hoje": hoje.isoformat(),
    }


def _to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _custo_medio_por_arroba():
    """Custo médio por arroba do rebanho ativo, ou None se não der para apurar."""
    ativos = db.get_all_animals(status="ativo")
    if not ativos:
        return None
    custo = arrobas = 0.0
    costs = db._costs_by_animal()
    for a in ativos:
        custo += costs.get(a["id"], 0.0) or 0
        arrobas += db.kg_to_arrobas(a["current_weight"],
                                    a.get("carcass_yield") or 0.52) or 0
    return round(custo / arrobas, 2) if arrobas else None


_GRAVIDADE_CARD = {"alta": "card-red", "media": "card-yellow", "baixa": "card-green"}


# §7.3 — cada pendência com o que é, e o prazo em que passa a valer. O prazo é
# metade da informação: sem ele, "12 animais sem identificação oficial" parece
# irregularidade quando ainda faltam anos para ser exigível.
_PENDENCIAS_7_3 = {
    "sem_mae_vinculada": (
        "Nascidos sem mãe vinculada",
        "§7.1 exige vínculo materno para animal nascido na propriedade. "
        "Vincule pela ficha do animal.", True),
    "nascimento_estimado": (
        "Nascimento com data estimada",
        "§7.1: parto não acompanhado entra como estimado. Não é erro — é um "
        "dado a confirmar quando houver como.", False),
    "sem_raca": (
        "Sem raça informada",
        "Dado de cadastro básico, e a raça entra em quase todo relatório "
        "zootécnico.", True),
    "sem_propriedade_de_nascimento": (
        "Sem propriedade de nascimento",
        "§3: o nascimento acontece em algum lugar, e é esse lugar que a "
        "rastreabilidade persegue.", True),
    "sem_identificacao_oficial": (
        "Sem identificação oficial",
        "§4.1 — exigível para trânsito a partir de **01/01/2033**, e o formato "
        "oficial ainda não foi publicado (§23). Está aqui para dimensionar o "
        "esforço, não para ser resolvido hoje.", False),
}


def _painel_pendencias_7_3():
    """Pendências de conformidade (PNIB §7.3).

    Fica numa aba própria, e **fora do contador da barra lateral**, de
    propósito. `sem_identificacao_oficial` lista o rebanho inteiro e só vira
    exigência em 2033: somá-lo ao badge deixaria o número permanentemente alto,
    e contador que nunca zera é contador que ninguém lê. O mesmo raciocínio já
    valeu para o alerta de sincronização do §8.
    """
    st.caption("O que falta para os dados estarem completos. Conformidade não é "
               "só registrar certo — é **saber o que falta** antes de a "
               "fiscalização perguntar (§7.3).")

    pend = db.nascimentos.pendencias()
    exigiveis = {k: v for k, v in pend.items()
                 if v and _PENDENCIAS_7_3.get(k, ("", "", True))[2]}

    if not any(pend.values()):
        st.success("✅ Nenhuma pendência de conformidade.")
        return

    if exigiveis:
        st.warning(f"⚠️ {_plural(len(exigiveis), 'tipo', 'tipos')} de pendência "
                   "com exigência já vigente.")
    else:
        st.success("✅ Nenhuma pendência de exigência vigente. O que resta abaixo "
                   "tem prazo futuro ou é informativo.")

    for chave, ids in pend.items():
        if not ids:
            continue
        titulo, explicacao, vigente = _PENDENCIAS_7_3.get(
            chave, (chave, "", True))
        icone = "🔴" if vigente else "⏳"
        with st.expander(f"{icone} {titulo} — {len(ids)}"):
            st.markdown(explicacao)
            st.dataframe(pd.DataFrame({"Brinco": ids}),
                         use_container_width=True, hide_index=True)

def page_alertas():
    st.markdown('<div class="page-title">🔔 Alertas Ativos</div>', unsafe_allow_html=True)
    # Duas perguntas diferentes: "o que faço hoje" e "o que falta nos dados".
    # Misturá-las faria a segunda, que é lenta e cumulativa, abafar a primeira.
    tab_op, tab_conf = st.tabs(["🔔 Operacionais", "📋 Conformidade (§7.3)"])
    with tab_conf:
        _painel_pendencias_7_3()
    with tab_op:
        _alertas_operacionais()


def _alertas_operacionais():
    # ── Recomendações do motor de regras (services/recomendacoes.py) ──────────
    st.subheader("🧭 Recomendações")
    st.caption("Regras explícitas sobre o estado atual da fazenda — cada uma diz o "
               "motivo e os números que a dispararam.")
    try:
        recs = avaliar_recomendacoes(_contexto_recomendacoes())
    except Exception as e:   # regra nova com dado faltando não pode derrubar a página
        recs = []
        st.warning(f"Não foi possível avaliar as recomendações: {e}")

    if recs:
        ordem = {"alta": 0, "media": 1, "baixa": 2}
        for r in sorted(recs, key=lambda x: ordem.get(x.get("severidade"), 9)):
            classe = _GRAVIDADE_CARD.get(r.get("severidade"), "card-yellow")
            acao = r.get("acao")
            st.markdown(
                f'<div class="{classe}"><b>{r.get("titulo","")}</b><br>'
                f'{r.get("motivo","")}'
                + (f'<br><i>👉 {acao}</i>' if acao else "")
                + '</div>', unsafe_allow_html=True)
    else:
        st.success("✅ Nenhuma recomendação no momento.")

    st.markdown("---")

    alerts=db.get_alert_animals()
    low   =db.check_low_stock()

    # Sumidos
    st.subheader(f"🔴 Animais Sumidos ({len(alerts['sumidos'])})")
    st.caption("Sem pesagem registrada há mais de 30 dias.")
    if alerts["sumidos"]:
        df_sum=pd.DataFrame([{"ID":a["id"],"Raça":a["breed"],"Lote":a.get("lote_id") or "—",
            "Último Peso (kg)":a["current_weight"],
            "Dias sem Pesagem":a["days_since_weighing"]} for a in alerts["sumidos"]])
        st.dataframe(df_sum,use_container_width=True,hide_index=True)
        for a in alerts["sumidos"]:
            c1,c2=st.columns([3,1])
            with c1:
                st.markdown(f'<div class="card-red">🔴 <b>{a["id"]}</b> — {a["breed"]} — '
                    f'Sem pesagem há <b>{a["days_since_weighing"]} dias</b></div>',
                    unsafe_allow_html=True)
            with c2:
                if st.button("📱 Ir para Campo",key=f"alr_sum_{a['id']}",use_container_width=True):
                    st.session_state.campo_id=a["id"]; _go("campo"); st.rerun()
    else:
        st.success("✅ Nenhum animal sumido.")

    st.markdown("---")

    # Carência
    st.subheader(f"🟡 Em Período de Carência ({len(alerts['carencia'])})")
    if alerts["carencia"]:
        for a in alerts["carencia"]:
            st.markdown(f'<div class="card-yellow">🟡 <b>{a["id"]}</b> — {a["breed"]} — '
                f'Carência até <b>{a["withdrawal_end"]}</b> '
                f'(<b>{a["days_remaining"]} dias restantes</b>)</div>',
                unsafe_allow_html=True)
    else:
        st.success("✅ Nenhum animal em carência.")

    st.markdown("---")

    # Prontos para abate
    st.subheader(f"🟢 Prontos para Abate ({len(alerts['prontos'])})")
    st.caption("Atingiram o peso-alvo e estão livres de carência.")
    if alerts["prontos"]:
        df_pro=pd.DataFrame([{"ID":a["id"],"Raça":a["breed"],
            "Peso Atual (kg)":a["current_weight"],"Peso-Alvo (kg)":a.get("target_weight") or 500,
            "@ Atuais":a["arrobas"]} for a in alerts["prontos"]])
        st.dataframe(df_pro,use_container_width=True,hide_index=True,
            column_config={"Peso Atual (kg)":st.column_config.NumberColumn(format="%.1f"),
                "@ Atuais":st.column_config.NumberColumn(format="%.2f")})
    else:
        st.info("Nenhum animal atingiu o peso-alvo ainda.")

    st.markdown("---")

    # Estoque crítico
    st.subheader(f"📦 Estoque Abaixo do Mínimo ({len(low)})")
    if low:
        for i in low:
            pct=i["current_stock"]/i["min_stock"]*100 if i["min_stock"] else 0
            st.markdown(f'<div class="card-yellow">⚠️ <b>{i["name"]}</b> — '
                f'Estoque: <b>{i["current_stock"]:.1f} {i["unit"]}</b> '
                f'(mínimo: {i["min_stock"]:.0f}) — <b>{pct:.0f}% do mínimo</b></div>',
                unsafe_allow_html=True)
        if st.button("📦 Ir para Estoque",type="primary"):
            _go("estoque"); st.rerun()
    else:
        st.success("✅ Todos os insumos com estoque adequado.")

    # Baixo desempenho (GMD abaixo da meta) — só admin gerencia a meta
    st.markdown("---")
    meta = db.get_gmd_target()
    low_perf = db.get_low_performance(meta)
    st.subheader(f"📉 Baixo Desempenho ({len(low_perf)})")
    st.caption(f"Animais com GMD abaixo da meta ({meta:.3f} kg/dia).")
    if low_perf:
        df_lp = pd.DataFrame([{"ID":a["id"],"Raça":a["breed"],
            "Lote":a.get("lote_id") or "—","Peso (kg)":a["current_weight"],
            "GMD (kg/dia)":round(a["gmd"],3)} for a in low_perf])
        st.dataframe(df_lp, use_container_width=True, hide_index=True,
            column_config={"Peso (kg)":st.column_config.NumberColumn(format="%.1f"),
                "GMD (kg/dia)":st.column_config.NumberColumn(format="%.3f")})
        if st.session_state.user["role"]=="admin" and st.button("📈 Ir para Desempenho",type="primary"):
            _go("desempenho"); st.rerun()
    else:
        st.success("✅ Nenhum animal abaixo da meta de GMD.")

# ══════════════════════════════════════════════════════════════════════════════
# RELATÓRIOS  (CSV + PDF)
# ══════════════════════════════════════════════════════════════════════════════
def page_relatorios():
    st.markdown('<div class="page-title">📄 Relatórios e Exportação</div>', unsafe_allow_html=True)
    animals=db.get_all_animals(status=None)

    rt1,rt2,rt3=st.tabs(["🐄 Inventário","⚖️ Pesagens","💰 Financeiro"])

    def _download_row(title, df, key):
        dc1,dc2,dc3=st.columns(3)
        with dc1:
            st.download_button(f"⬇️ CSV",_df_to_csv(df),
                f"agrotop_{key}.csv","text/csv",use_container_width=True,
                key=f"csv_{key}")
        with dc2:
            xlsx_bytes=_df_to_xlsx(title,df)
            if xlsx_bytes:
                st.download_button(f"⬇️ Excel",xlsx_bytes,
                    f"agrotop_{key}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,key=f"xlsx_{key}")
            else:
                st.info("Instale `openpyxl` p/ Excel.")
        with dc3:
            pdf_bytes=_df_to_pdf(f"AgroTop — {title}",df)
            if pdf_bytes:
                st.download_button(f"⬇️ PDF",pdf_bytes,
                    f"agrotop_{key}.pdf","application/pdf",use_container_width=True,
                    key=f"pdf_{key}")
            else:
                st.info("Instale `fpdf2` p/ PDF.")

    with rt1:
        st.subheader("🐄 Inventário Completo do Rebanho")
        rows_inv=[]
        a_ids = [a["id"] for a in animals]
        gmd_batch = db.calculate_gmd_bulk(a_ids)
        wd_batch = db.get_withdrawal_end_batch(a_ids)
        for a in animals:
            gmd=gmd_batch.get(a["id"])
            wd=wd_batch.get(a["id"])
            rows_inv.append({"ID":a["id"],"Raça":a["breed"],
                "Sexo":"M" if a["sex"]=="M" else "F",
                "Categoria":db.get_age_category(a.get("birth_date")),
                "Idade":db.get_age_display(a),
                "Data Nascimento":a.get("birth_date") or "",
                "Nasc. Estimado":"Sim" if a.get("birth_estimated") else "Não",
                "Origem Idade":db.AGE_SOURCES.get(a.get("age_source","propriedade"),""),
                "Data Entrada":a["entry_date"],
                "Peso Entrada (kg)":a["entry_weight"],
                "Peso Atual (kg)":a["current_weight"],
                "Ganho (kg)":round(a["current_weight"]-a["entry_weight"],1),
                "@ Atuais":db.kg_to_arrobas(a["current_weight"]),
                "GMD (kg/dia)":gmd or 0,"Status":a["status"],
                "Lote":a.get("lote_id") or "",
                "Fornecedor":a.get("fornecedor_name") or "",
                "NF":a.get("nf_number") or "",
                "GTA":a.get("gta_number") or "",
                "Carência até":wd.isoformat() if wd else ""})
        df_inv=pd.DataFrame(rows_inv)
        st.dataframe(df_inv,use_container_width=True,hide_index=True,height=350)
        _download_row("Inventário",df_inv,"inventario")

    with rt2:
        st.subheader("⚖️ Histórico de Pesagens")
        raw=db.get_all_weighings()
        if raw:
            df_p=pd.DataFrame(raw)[["animal_id","weigh_date","weight","method","lote_id","operator","notes"]].copy()
            df_p["method"]=df_p["method"].fillna("pesado").map(lambda m: db.WEIGH_METHODS.get(m,m))
            df_p.columns=["Animal","Data","Peso (kg)","Método","Lote","Operador","Obs"]
            st.dataframe(df_p,use_container_width=True,hide_index=True,height=350)
            _download_row("Pesagens",df_p,"pesagens")

    with rt3:
        ul = _unit_label()
        st.subheader("💰 Relatório Financeiro")
        price_lbl = f"Cotação (R$/{ul}) para o relatório"
        default_p = DEFAULT_PRICE_ARROBA if _use_arroba() else DEFAULT_PRICE_KG
        cotacao_r = st.number_input(price_lbl, min_value=0.01, max_value=5000.0,
            value=default_p, step=1.0)
        rend_r = 52
        if _use_arroba():
            rend_r = st.slider("Rendimento de Carcaça (%)", 40, 65, 52,
                key="rend_relatorio")
        rows_fin=[]
        costs = db._costs_by_animal()
        for a in animals:
            if a["status"] not in ("ativo","carencia"): continue
            tc   = costs.get(a["id"], 0.0)
            prod = _live_weight(a["current_weight"], rend_r/100)
            be   = round(tc/prod, 2) if prod else 0
            receita = round(prod * cotacao_r, 2)
            lucro   = round(receita - tc, 2)
            rows_fin.append({"ID":a["id"],"Raça":a["breed"],
                "Peso Atual (kg)":a["current_weight"],
                f"Prod. ({ul})":prod,
                "Custo Total (R$)":tc,
                _breakeven_label():be,
                f"Receita @ R${cotacao_r:.0f}/{ul}":receita,
                "Lucro Estimado (R$)":lucro})
        df_fin=pd.DataFrame(rows_fin)
        if not df_fin.empty:
            st.dataframe(df_fin,use_container_width=True,hide_index=True)
            _download_row("Financeiro",df_fin,"financeiro")

# ══════════════════════════════════════════════════════════════════════════════
# CADASTRAR
# ══════════════════════════════════════════════════════════════════════════════
def _age_inputs(entry_date, key_prefix=""):
    """Renderiza os campos de idade conforme o método escolhido.
    Retorna (birth_date_str|None, birth_estimated, age_source, erro|None).
    Deve ser chamado FORA de um st.form para permitir troca dinâmica."""
    MESES = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
    metodo = st.radio(
        "🎂 Como definir a idade?",
        list(db.AGE_SOURCES.keys()),
        format_func=lambda k: db.AGE_SOURCES[k],
        key=f"{key_prefix}age_method",
        horizontal=False,
    )

    bd_str, estimated, err = None, 0, None

    if metodo == "propriedade":
        bd = st.date_input("Data de Nascimento (exata)", value=None,
                           key=f"{key_prefix}bd_exact",
                           help="Para animais nascidos na propriedade")
        if bd:
            bd_str, estimated = bd.isoformat(), 0
        else:
            err = "Informe a data de nascimento exata."

    elif metodo == "estimado":
        st.caption("Estime o mês/ano aproximado do nascimento.")
        c1, c2 = st.columns(2)
        with c1:
            mes = st.selectbox("Mês aproximado", range(1,13),
                format_func=lambda m: MESES[m-1], key=f"{key_prefix}est_m")
        with c2:
            ano = st.number_input("Ano", min_value=2000, max_value=date.today().year,
                value=date.today().year-1, step=1, key=f"{key_prefix}est_y")
        try:
            bd_str = date(int(ano), int(mes), 15).isoformat()
            estimated = 1
        except ValueError:
            err = "Mês/ano inválido."

    elif metodo == "operador":
        st.caption("Informe a idade atual estimada do animal (hoje).")
        meses = st.number_input("Idade atual (meses)", min_value=0, max_value=360,
            value=24, step=1, key=f"{key_prefix}op_m")
        bd_str = db.birth_date_from_age(int(meses), date.today())
        estimated = 1
        st.info(f"📌 Nascimento estimado: **{bd_str}** · Categoria: "
                f"**{db.get_age_category(bd_str)}**")

    elif metodo == "nf_gta":
        st.caption("Informe a idade que consta na NF / GTA e a data do documento.")
        c1, c2 = st.columns(2)
        with c1:
            meses = st.number_input("Idade na NF/GTA (meses)", min_value=0, max_value=360,
                value=18, step=1, key=f"{key_prefix}nf_m")
        with c2:
            doc_date = st.date_input("Data do documento", value=entry_date,
                key=f"{key_prefix}nf_d")
        bd_str = db.birth_date_from_age(int(meses), doc_date)
        estimated = 1
        st.info(f"📌 Nascimento estimado: **{bd_str}** · Idade hoje: "
                f"**{db.get_age_months(bd_str)} meses** · Categoria: "
                f"**{db.get_age_category(bd_str)}**")

    return bd_str, estimated, metodo, err


_GRAVIDADE_ICONE = {"bloqueio": "🔴", "alerta": "🟡", "informativo": "🔵"}


def _cadastro_nascimento():
    """Registro de nascimento (PNIB §7).

    A regra é `services/genealogia.py`, via `repositories/nascimentos.py`. Aqui
    só a tela — e a decisão de interface que o §7.2 impõe: **bloqueio impede,
    alerta pede confirmação.** O texto do §7.2 é explícito: o sistema deve
    "emitir alerta, sem substituir a avaliação técnica". Quem avalia é o
    técnico; o software mostra o que sabe.
    """
    st.caption("Nascimento na propriedade. A mãe precisa estar cadastrada e ativa — "
               "o vínculo materno é exigência do §7 do PNIB e não pode ser preenchido depois "
               "sem deixar rastro.")

    femeas = [a for a in db.get_all_animals(status="ativo") if a.get("sex") == "F"]
    if not femeas:
        st.warning("Nenhuma fêmea ativa no rebanho. Cadastre a mãe antes de registrar a cria.")
        return

    c1, c2, c3 = st.columns([2, 1, 1])
    with c1:
        rotulos = {f"{a['id']} — {a['breed']}": a for a in femeas}
        mae = rotulos[st.selectbox("🐄 Mãe *", list(rotulos), key="nasc_mae")]
    with c2:
        data_parto = st.date_input("📅 Data do parto *", value=date.today(),
                                   max_value=date.today(), key="nasc_data")
    with c3:
        hora = st.text_input("Hora", placeholder="14:30", key="nasc_hora").strip()

    c4, c5, c6 = st.columns(3)
    with c4:
        tipo_parto = st.selectbox("Tipo de parto", ["normal", "assistido", "cesarea"],
                                  key="nasc_tipo")
    with c5:
        condicao = st.selectbox("Condição", ["nascido_vivo", "natimorto"],
                                format_func=lambda v: "Nascido vivo" if v == "nascido_vivo"
                                else "Natimorto", key="nasc_cond")
    with c6:
        n_crias = st.number_input("Nº de crias", min_value=1, max_value=4, value=1,
                                  key="nasc_n",
                                  help="Duas ou mais geram animais distintos ligados ao "
                                       "MESMO parto — é o que o §7.2 exige para gêmeos.")

    data_estimada = st.checkbox(
        "Data estimada (parto não acompanhado)", key="nasc_est",
        help="§7.1: marcar quando ninguém presenciou. Fica registrado como estimado "
             "e aparece nas pendências.")

    # ── Prévia da validação, ANTES de o usuário preencher as crias ───────────
    problemas = db.nascimentos.avaliar(mae["uuid"], data_parto.isoformat(),
                                       mae.get("property_id"))
    bloqueios = [p for p in problemas if p["gravidade"] == "bloqueio"]
    alertas = [p for p in problemas if p["gravidade"] == "alerta"]

    if problemas:
        st.markdown("**Verificação do vínculo materno (§7.2)**")
        for p in problemas:
            icone = _GRAVIDADE_ICONE.get(p["gravidade"], "•")
            texto = f"{icone} {p['mensagem']}"
            if p["gravidade"] == "bloqueio":
                st.error(texto)
            elif p["gravidade"] == "alerta":
                st.warning(texto)
            else:
                st.info(texto)

    if bloqueios:
        st.error("🚫 Não é possível registrar enquanto houver bloqueio. "
                 "Corrija o cadastro da mãe ou a data do parto.")
        return

    confirmado = False
    if alertas:
        confirmado = st.checkbox(
            "Avaliei os alertas acima e confirmo o registro", key="nasc_conf",
            help="§7.2: o sistema alerta, sem substituir a avaliação técnica. "
                 "A confirmação fica registrada.")

    st.markdown("---")
    st.markdown(f"**{'Crias' if n_crias > 1 else 'Cria'}**")

    crias = []
    for i in range(int(n_crias)):
        k1, k2, k3, k4 = st.columns([2, 1, 1, 1])
        with k1:
            brinco = st.text_input(f"🏷️ Brinco {i+1} *", key=f"nasc_id_{i}").strip().upper()
        with k2:
            sexo = st.selectbox("Sexo", ["M", "F"], key=f"nasc_sexo_{i}",
                                format_func=lambda v: "♂" if v == "M" else "♀")
        with k3:
            raca = st.selectbox("Raça", BREEDS, key=f"nasc_raca_{i}",
                                index=BREEDS.index(mae["breed"]) if mae["breed"] in BREEDS else 0)
        with k4:
            peso = st.number_input("Peso (kg)", min_value=0.0, max_value=100.0,
                                   step=0.5, value=0.0, key=f"nasc_peso_{i}")
        crias.append({"id": brinco, "sexo": sexo, "raca": raca,
                      "peso": peso or None})

    obs = st.text_area("Observações", key="nasc_obs").strip()

    brincos = [cr["id"] for cr in crias if cr["id"]]
    faltando = len(brincos) < int(n_crias)
    repetidos = len(brincos) != len(set(brincos))
    if repetidos:
        st.error("🚫 Dois brincos iguais na mesma ninhada.")

    pode = (not faltando and not repetidos
            and (not alertas or confirmado))
    if faltando:
        st.caption("Informe o brinco de cada cria para habilitar o registro.")

    if st.button("✅ Registrar nascimento", type="primary", disabled=not pode,
                 key="nasc_salvar"):
        r = db.nascimentos.registrar(
            mae["uuid"], data_parto.isoformat(), crias,
            hora=hora, tipo_parto=tipo_parto, condicao=condicao,
            propriedade_id=mae.get("property_id"),
            responsavel=st.session_state.user["name"],
            data_estimada=data_estimada, observacoes=obs,
            ignorar_alertas=confirmado)

        if r.get("ok"):
            nomes = ", ".join(brincos)
            st.success(f"✅ Nascimento registrado: {nomes}"
                       + (" — gêmeos no mesmo parto" if len(brincos) > 1 else ""))
            st.rerun()
        elif r.get("exige_confirmacao"):
            st.warning("Há alertas: marque a confirmação acima para prosseguir.")
        else:
            st.error(f"🚫 {r.get('erro', 'Não foi possível registrar.')}")


def page_cadastrar():
    st.markdown('<div class="page-title">➕ Cadastrar Novo Animal</div>', unsafe_allow_html=True)
    # Duas formas de um animal entrar no rebanho, e elas são diferentes: comprado
    # tem fornecedor e preço; nascido tem mãe, parto e validação do §7 do PNIB.
    tab_compra, tab_nasc = st.tabs(["🛒 Comprado / Recebido", "🐮 Nascimento na fazenda"])
    with tab_nasc:
        _cadastro_nascimento()
    with tab_compra:
        _cadastro_compra()


def _cadastro_compra():
    fornecedores=db.get_all_fornecedores()
    lotes=[l for l in db.get_all_lotes() if l["status"]=="ativo"]

    # Campos fora do form (para reagir à troca de método de idade/peso)
    c_top1, c_top2 = st.columns(2)
    with c_top1:
        aid=st.text_input("🏷️ ID / Brinco *",placeholder="Ex: BR0015").strip().upper()
        breed=st.selectbox("🐄 Raça *",BREEDS)
        sex=st.radio("Sexo *",["♂ Macho","♀ Fêmea"],horizontal=True)
    with c_top2:
        entry_date=st.date_input("📅 Data de Entrada *",value=date.today())
        target_weight=st.number_input("🎯 Peso-Alvo de Abate (kg)",
            min_value=0.0,max_value=2000.0,value=500.0,step=5.0,format="%.1f")

    st.markdown("**📆 Definição de Idade / Categoria**")
    birth_date_str, birth_est, age_src, age_err = _age_inputs(entry_date, "cad_")
    is_propriedade = (age_src == "propriedade")

    # ── Peso de entrada (com método) ──────────────────────────────────────────
    st.markdown("**⚖️ Peso de Entrada**")
    st.caption("Não é obrigatório pesar na balança — pode estimar ou usar medição.")
    peso_metodo = st.radio("Como obter o peso?",
        list(db.WEIGH_METHODS.keys()),
        format_func=lambda m: db.WEIGH_METHODS[m],
        horizontal=True, key="cad_peso_metodo")

    if peso_metodo == "medicao":
        pm1, pm2, pm3 = st.columns(3)
        with pm1:
            pt_c=st.number_input("Perímetro torácico (cm)",min_value=0.0,max_value=350.0,
                value=180.0,step=1.0,key="cad_pt")
        with pm2:
            comp_c=st.number_input("Comprimento corporal (cm)",min_value=0.0,max_value=350.0,
                value=150.0,step=1.0,key="cad_comp")
        entry_weight = db.estimate_weight_by_measurement(pt_c, comp_c)
        with pm3:
            st.metric("Peso estimado", f"{entry_weight:.1f} kg")
        medida_nota = f"PT={pt_c:.0f}cm Comp={comp_c:.0f}cm"
    else:
        lbl = "Peso na balança (kg) *" if peso_metodo=="pesado" else "Peso estimado (kg) *"
        entry_weight=st.number_input(lbl,min_value=0.1,max_value=2000.0,
            step=0.5,format="%.1f",key="cad_peso_valor")
        medida_nota = ""

    with st.form("f_cad",clear_on_submit=False):
        cf1, cf2 = st.columns(2)
        with cf1:
            # Valor de compra só para animais adquiridos (não nascidos na propriedade)
            if not is_propriedade:
                compra_modo = st.selectbox("💰 Como foi a compra?",
                    ["cabeca","kg"],
                    format_func=lambda m: "Por cabeça (valor fechado)" if m=="cabeca" else "Por kg (peso × preço/kg)")
                if compra_modo == "kg":
                    preco_kg_compra = st.number_input("Preço por kg (R$)",
                        min_value=0.0, step=0.10, format="%.2f")
                    purchase_price = round(entry_weight * preco_kg_compra, 2)
                    st.caption(f"→ Total: **R\\$ {purchase_price:,.2f}** ({entry_weight:.0f} kg × R\\$ {preco_kg_compra:.2f})")
                else:
                    purchase_price=st.number_input("Valor de Compra (R$)",
                        min_value=0.0,step=10.0,format="%.2f")
            else:
                purchase_price=0.0
                compra_modo="propriedade"
                st.caption("💰 Valor de compra não se aplica a animais nascidos na propriedade.")
            lote_sel=st.selectbox("🌿 Lote de Destino",
                [None]+lotes,format_func=lambda x:"— Sem lote —" if x is None else f"{x['id']} — {x['name']}")
        with cf2:
            forn_sel=st.selectbox("🚚 Fornecedor / Origem",
                [None]+fornecedores,format_func=lambda x:"— Não informado —" if x is None else f"{x['name']} ({x['city']}/{x['state']})")
            notes=st.text_area("📝 Observações",height=70,
                placeholder="Opcional",value=medida_nota)

        if not is_propriedade:
            st.caption("📄 Documentos de compra (opcional)")
            cd1, cd2 = st.columns(2)
            with cd1:
                nf_number=st.text_input("Número da NF",placeholder="Ex: 012345",
                    help="Nota Fiscal — opcional").strip()
            with cd2:
                gta_number=st.text_input("Número da GTA",placeholder="Ex: MT-0009876",
                    help="Guia de Trânsito Animal — opcional").strip()
        else:
            nf_number=gta_number=""

        if st.form_submit_button("✅ Cadastrar Animal",type="primary",use_container_width=True):
            errs=[]
            if not aid:             errs.append("ID do animal é obrigatório.")
            elif db.get_animal(aid):errs.append(f"Animal **{aid}** já existe.")
            if entry_weight<=0:     errs.append("Peso de entrada deve ser > 0.")
            if age_err:             errs.append(age_err)
            if errs:
                for e in errs: st.error(f"❌ {e}")
            else:
                db.add_animal(
                    aid, breed,
                    "M" if "Macho" in sex else "F",
                    birth_date_str,
                    entry_date.strftime("%Y-%m-%d"),
                    entry_weight, target_weight, purchase_price,
                    lote_sel["id"] if lote_sel else None,
                    forn_sel["id"] if forn_sel else None,
                    notes,
                    birth_estimated=birth_est,
                    age_source=age_src,
                    nf_number=nf_number,
                    gta_number=gta_number,
                    weight_method=peso_metodo,
                    purchase_mode=compra_modo,
                )
                cat = db.get_age_category(birth_date_str)
                st.success(f"✅ Animal **{aid}** cadastrado! Categoria: **{cat}** · "
                           f"Peso: {entry_weight:.1f} kg ({db.WEIGH_METHODS[peso_metodo]})")
                st.balloons()

    # Cadastrar Fornecedor rápido
    with st.expander("➕ Cadastrar novo Fornecedor / Origem"):
        with st.form("f_forn",clear_on_submit=True):
            ff1,ff2,ff3=st.columns(3)
            with ff1: fn=st.text_input("Nome *")
            with ff2: fc=st.text_input("Cidade")
            with ff3: fs=st.selectbox("Estado",["MT","MS","GO","MG","SP","PR","RS","BA","TO","PA","RO","Outro"])
            fcontact=st.text_input("Contato")
            if st.form_submit_button("✅ Salvar Fornecedor",use_container_width=True):
                if fn:
                    db.add_fornecedor(fn,fc,fs,fcontact)
                    st.success(f"✅ {fn} cadastrado!"); st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# ADMIN
# ══════════════════════════════════════════════════════════════════════════════
def _admin_users():
    st.subheader("👥 Gestão de Usuários")
    users = db.get_all_users()

    # ── Lista de usuários ─────────────────────────────────────────────────────
    df_u = pd.DataFrame([{
        "ID": u["id"], "Usuário": u["username"], "Nome": u["name"],
        "Papel": "Administrador" if u["role"]=="admin" else "Operador",
    } for u in users])
    st.dataframe(df_u, use_container_width=True, hide_index=True)

    st.markdown("---")
    col_edit, col_new = st.columns(2)

    # ── Editar usuário existente ──────────────────────────────────────────────
    with col_edit:
        st.markdown("**✏️ Editar Usuário**")
        sel = st.selectbox("Usuário", users,
            format_func=lambda u: f"{u['username']} — {u['name']}", key="edit_user_sel")
        with st.form("f_edit_user", clear_on_submit=False):
            e_username = st.text_input("Usuário (login)", value=sel["username"]).strip()
            e_name = st.text_input("Nome", value=sel["name"]).strip()
            e_role = st.selectbox("Papel", ["operator","admin"],
                index=0 if sel["role"]=="operator" else 1,
                format_func=lambda r: "Operador" if r=="operator" else "Administrador")
            st.caption("Deixe a senha em branco para mantê-la. Preencha para redefinir.")
            e_pwd = st.text_input("Nova senha", type="password", placeholder="••••••••")
            e_pwd2 = st.text_input("Confirmar nova senha", type="password", placeholder="••••••••")
            if st.form_submit_button("💾 Salvar Alterações", type="primary", use_container_width=True):
                errs = []
                if not e_username: errs.append("Usuário não pode ficar vazio.")
                elif db.username_exists(e_username, exclude_id=sel["id"]):
                    errs.append(f"O login '{e_username}' já está em uso.")
                if not e_name: errs.append("Nome não pode ficar vazio.")
                if e_pwd and e_pwd != e_pwd2: errs.append("As senhas não coincidem.")
                # Impede remover o último admin
                if sel["role"]=="admin" and e_role=="operator" and db.count_admins()<=1:
                    errs.append("Não é possível rebaixar o único administrador.")
                if errs:
                    for x in errs: st.error(f"❌ {x}")
                else:
                    if e_username != sel["username"]:
                        db.update_username(sel["id"], e_username)
                    db.update_user(sel["id"], e_name, e_role, e_pwd or None)
                    st.success(f"✅ Usuário '{e_username}' atualizado!")
                    # Se editou a si mesmo, atualiza a sessão
                    if sel["id"] == st.session_state.user["id"]:
                        st.session_state.user = db.get_user(sel["id"])
                    st.rerun()

        # Excluir usuário
        with st.expander("🗑️ Excluir usuário"):
            del_sel = st.selectbox("Usuário a excluir", users,
                format_func=lambda u: f"{u['username']} — {u['name']}", key="del_user_sel")
            st.warning("Esta ação é permanente.")
            if st.button("Excluir definitivamente", type="secondary"):
                if del_sel["id"] == st.session_state.user["id"]:
                    st.error("Você não pode excluir a sua própria conta logada.")
                elif del_sel["role"]=="admin" and db.count_admins()<=1:
                    st.error("Não é possível excluir o único administrador.")
                else:
                    db.delete_user(del_sel["id"])
                    st.success(f"Usuário '{del_sel['username']}' excluído."); st.rerun()

    # ── Novo usuário ──────────────────────────────────────────────────────────
    with col_new:
        st.markdown("**➕ Novo Usuário**")
        with st.form("f_new_user", clear_on_submit=True):
            n_username = st.text_input("Usuário (login) *", placeholder="ex: op2").strip()
            n_name = st.text_input("Nome *", placeholder="ex: Maria Operadora").strip()
            n_role = st.selectbox("Papel *", ["operator","admin"],
                format_func=lambda r: "Operador" if r=="operator" else "Administrador")
            n_pwd = st.text_input("Senha *", type="password", placeholder="••••••••")
            n_pwd2 = st.text_input("Confirmar senha *", type="password", placeholder="••••••••")
            if st.form_submit_button("✅ Criar Usuário", type="primary", use_container_width=True):
                errs = []
                if not n_username: errs.append("Usuário é obrigatório.")
                elif db.username_exists(n_username): errs.append(f"O login '{n_username}' já existe.")
                if not n_name: errs.append("Nome é obrigatório.")
                if not n_pwd: errs.append("Senha é obrigatória.")
                elif n_pwd != n_pwd2: errs.append("As senhas não coincidem.")
                if errs:
                    for x in errs: st.error(f"❌ {x}")
                else:
                    db.add_user(n_username, n_pwd, n_name, n_role)
                    st.success(f"✅ Usuário '{n_username}' criado!")
                    st.rerun()


def page_clima():
    if st.session_state.user["role"] != "admin":
        st.error("🔒 Acesso restrito ao Administrador."); return
    st.markdown('<div class="page-title">🌧️ Clima & Chuva</div>', unsafe_allow_html=True)
    lotes = db.get_all_lotes()

    ct1, ct2, ct3 = st.tabs(["🌦️ Previsão do Tempo", "💧 Registrar Chuva", "📊 Histórico de Chuva"])

    # ── Previsão do tempo ─────────────────────────────────────────────────────
    with ct1:
        lat = db.get_setting("farm_lat")
        lon = db.get_setting("farm_lon")
        with st.expander("📍 Localização da fazenda (para a previsão)",
                         expanded=not (lat and lon)):
            st.caption("Pegue as coordenadas no Google Maps: clique no local do mapa e "
                       "copie os números que aparecem (latitude, longitude).")
            lc1, lc2 = st.columns(2)
            with lc1:
                nlat = st.number_input("Latitude", value=float(lat) if lat else -15.60000,
                    format="%.5f", step=0.001)
            with lc2:
                nlon = st.number_input("Longitude", value=float(lon) if lon else -56.10000,
                    format="%.5f", step=0.001)
            if st.button("💾 Salvar localização", type="primary"):
                db.set_setting("farm_lat", nlat)
                db.set_setting("farm_lon", nlon)
                st.success("📍 Localização salva!"); st.rerun()

        if lat and lon:
            fc = _fetch_forecast(float(lat), float(lon))
            if fc and "daily" in fc:
                d = fc["daily"]
                df = pd.DataFrame({
                    "Data": pd.to_datetime(d["time"]),
                    "Chuva (mm)": d["precipitation_sum"],
                    "Prob. chuva (%)": d["precipitation_probability_max"],
                    "Mín (°C)": d["temperature_2m_min"],
                    "Máx (°C)": d["temperature_2m_max"],
                })
                hoje = df.iloc[0]; amanha = df.iloc[1] if len(df) > 1 else df.iloc[0]
                mk = st.columns(4)
                mk[0].metric("Hoje", f"{hoje['Máx (°C)']:.0f}° / {hoje['Mín (°C)']:.0f}°",
                             help="Máxima / mínima")
                mk[1].metric("Chuva hoje", f"{hoje['Chuva (mm)']:.0f} mm",
                             delta=f"{hoje['Prob. chuva (%)']:.0f}% prob.")
                mk[2].metric("Chuva amanhã", f"{amanha['Chuva (mm)']:.0f} mm",
                             delta=f"{amanha['Prob. chuva (%)']:.0f}% prob.")
                mk[3].metric("Chuva prevista (7 dias)", f"{sum(d['precipitation_sum']):.0f} mm")

                fig = px.bar(df, x="Data", y="Chuva (mm)", color="Prob. chuva (%)",
                    color_continuous_scale=[c["texto_secundario"], c["info"], c["info_secundario"]],
                    labels={"Chuva (mm)":"Chuva prevista (mm)"})
                fig.update_layout(**_layout(height=280, xaxis=dict(gridcolor=c["superficie"]),
                    yaxis=dict(gridcolor=c["superficie"])))
                st.plotly_chart(fig, use_container_width=True)
                dfx = df.copy(); dfx["Data"] = dfx["Data"].dt.strftime("%d/%m (%a)")
                st.dataframe(dfx, use_container_width=True, hide_index=True,
                    column_config={c: st.column_config.NumberColumn(format="%.0f")
                                   for col in ["Chuva (mm)","Prob. chuva (%)","Mín (°C)","Máx (°C)"]})
                st.caption("Fonte: Open-Meteo · a mesma previsão vale para todos os piquetes da fazenda.")
            else:
                st.warning("Não foi possível obter a previsão agora (sem internet ou serviço "
                           "indisponível). Tente novamente em alguns minutos.")
        else:
            st.info("Defina a **localização da fazenda** acima para ver a previsão do tempo.")

    # ── Registrar chuva (por piquete) ─────────────────────────────────────────
    with ct2:
        st.caption("Registre a leitura do pluviômetro. Se houver um pluviômetro por piquete, "
                   "escolha o piquete; se for um só, deixe em **Geral / Sede**.")
        with st.form("f_rain", clear_on_submit=True):
            r1, r2 = st.columns(2)
            with r1: rdate = st.date_input("Data da leitura", value=date.today())
            with r2: rmm = st.number_input("Chuva medida (mm)", min_value=0.0, step=1.0, format="%.1f")
            lote_sel = st.selectbox("Piquete", [None]+lotes,
                format_func=lambda x: "Geral / Sede" if x is None else f"{x['id']} — {x['name']}")
            rnotes = st.text_input("Observações", placeholder="Opcional")
            if st.form_submit_button("💧 Registrar chuva", type="primary", use_container_width=True):
                db.add_rain(rdate.strftime("%Y-%m-%d"), rmm,
                    lote_sel["id"] if lote_sel else None,
                    st.session_state.user["name"], rnotes)
                st.success(f"✅ {rmm:.1f} mm registrados."); st.rerun()

    # ── Histórico de chuva ────────────────────────────────────────────────────
    with ct3:
        h1, h2 = st.columns(2)
        with h1: start = st.date_input("De", value=date(date.today().year,1,1), key="rain_start")
        with h2: end = st.date_input("Até", value=date.today(), key="rain_end")
        chuvas = db.get_rain(start.isoformat(), end.isoformat())
        total = db.get_rain_total(start.isoformat(), end.isoformat())
        st.metric("🌧️ Chuva acumulada no período", f"{total:.0f} mm")
        if chuvas:
            df = pd.DataFrame(chuvas)
            df["read_date"] = pd.to_datetime(df["read_date"])
            df["piquete"] = df["lote_name"].fillna("Geral / Sede")
            gc1, gc2 = st.columns(2)
            with gc1:
                st.markdown("**Por mês**")
                dfm = (df.groupby(df["read_date"].dt.to_period("M").astype(str))["rain_mm"]
                       .sum().reset_index())
                dfm.columns = ["Mês","Chuva (mm)"]
                figm = px.bar(dfm, x="Mês", y="Chuva (mm)",
                    color_discrete_sequence=[c["info_secundario"]])
                figm.update_layout(**_layout(height=250, xaxis=dict(gridcolor=c["superficie"]),
                    yaxis=dict(gridcolor=c["superficie"])))
                st.plotly_chart(figm, use_container_width=True)
            with gc2:
                st.markdown("**Por piquete**")
                dfl = df.groupby("piquete")["rain_mm"].sum().reset_index().sort_values("rain_mm")
                figl = px.bar(dfl, x="rain_mm", y="piquete", orientation="h",
                    color="rain_mm", color_continuous_scale=[c["texto_secundario"], c["info_secundario"]],
                    labels={"rain_mm":"Chuva (mm)","piquete":""})
                figl.update_layout(**_layout(height=250, coloraxis_showscale=False,
                    xaxis=dict(gridcolor=c["superficie"]), yaxis=dict(gridcolor=c["superficie"])))
                st.plotly_chart(figl, use_container_width=True)
            dft = df[["read_date","rain_mm","piquete","operator"]].copy()
            dft["read_date"] = dft["read_date"].dt.strftime("%d/%m/%Y")
            dft.columns = ["Data","Chuva (mm)","Piquete","Registrado por"]
            st.dataframe(dft, use_container_width=True, hide_index=True,
                column_config={"Chuva (mm)":st.column_config.NumberColumn(format="%.1f")})
        else:
            st.info("Nenhum registro de chuva no período. Registre na aba **Registrar Chuva**.")


def page_sanitario():
    if st.session_state.user["role"] != "admin":
        st.error("🔒 Acesso restrito ao Administrador."); return
    st.markdown('<div class="page-title">💉 Calendário Sanitário</div>', unsafe_allow_html=True)
    insumos = [i for i in db.get_all_insumos() if i["category"] in ("vacina","medicamento")]

    tp1, tp2 = st.tabs(["📋 Plano de Vacinação", "⚙️ Protocolos"])

    # ── Plano de vacinação ────────────────────────────────────────────────────
    with tp1:
        prots = db.get_protocols(active_only=True)
        if not prots:
            st.info("Nenhum protocolo ativo. Cadastre em **Protocolos** (aba ao lado) "
                    "as vacinações obrigatórias da sua região.")
        for p in prots:
            plan = db.get_protocol_plan(p)
            freq = db.PROTOCOL_FREQUENCIES.get(p["frequency"], p["frequency"])
            sexo = db.SEX_TARGETS.get(p["sex_target"], p["sex_target"])
            dose_desc = (f"{p['dose_value']:g} {p['dose_unit']} a cada {p['dose_ref_kg']:.0f} kg"
                         if (p.get("dose_ref_kg") or 0) > 0
                         else f"{p['dose_value']:g} {p['dose_unit']} por animal")
            st.markdown(f"#### 💉 {p['name']}")
            st.caption(f"{sexo} · {p['age_min']}–{p['age_max']} meses · {freq} · dose: {dose_desc}"
                       + (f" · insumo: {p['insumo_name']}" if p.get('insumo_name') else ""))
            k = st.columns(5)
            k[0].metric("Elegíveis", plan["n_eligible"])
            k[1].metric("Pendentes", plan["n_pending"])
            k[2].metric("Doses necessárias", f"{plan['doses_needed']:g} {p['dose_unit']}")
            if p.get("insumo_id"):
                k[3].metric("Estoque", f"{plan['stock']:g} {p.get('insumo_unit') or ''}")
                if plan["shortfall"] > 0:
                    k[4].metric("Faltam comprar", f"{plan['shortfall']:g}", delta="repor",
                                delta_color="inverse")
                else:
                    k[4].metric("Estoque", "✅ suficiente")
            else:
                k[3].metric("Estoque", "— (sem insumo)")
            if plan["idade_desconhecida"]:
                st.caption(f"⚠️ {_plural(plan['idade_desconhecida'],'animal','animais')} do sexo-alvo "
                           f"sem idade definida — verifique se precisam.")
            if plan["n_pending"] > 0:
                with st.form(f"camp_{p['id']}", clear_on_submit=True):
                    cc1, cc2 = st.columns([1,2])
                    with cc1:
                        cd = st.date_input("Data", value=date.today(), key=f"cd_{p['id']}")
                    with cc2:
                        if plan["shortfall"] > 0:
                            st.warning(f"Estoque insuficiente (faltam {plan['shortfall']:g}). "
                                       "A aplicação prossegue e o estoque vai a zero.")
                    if st.form_submit_button(
                            f"💉 Aplicar campanha em {plan['n_pending']} animais",
                            type="primary", use_container_width=True):
                        r = db.apply_protocol_campaign(p["id"], cd.strftime("%Y-%m-%d"),
                                                       st.session_state.user["name"])
                        st.success(f"✅ Aplicado em {r['n']} animais ({r['doses']:g} {p['dose_unit']}).")
                        st.rerun()
            else:
                st.success("✅ Todos os elegíveis já estão em dia com este protocolo.")
            st.markdown("---")

    # ── Gestão de protocolos ──────────────────────────────────────────────────
    with tp2:
        st.markdown("**Protocolos cadastrados**")
        todos = db.get_protocols(active_only=False)
        for p in todos:
            c1, c2, c3 = st.columns([5,1,1])
            with c1:
                ativo = "🟢" if p["active"] else "⚪"
                st.markdown(f"{ativo} **{p['name']}** — {db.SEX_TARGETS.get(p['sex_target'],'')} · "
                            f"{p['age_min']}–{p['age_max']}m · {db.PROTOCOL_FREQUENCIES.get(p['frequency'],'')}")
            with c2:
                if st.button("Pausar" if p["active"] else "Ativar", key=f"tgp_{p['id']}",
                             use_container_width=True):
                    db.set_protocol_active(p["id"], 0 if p["active"] else 1); st.rerun()
            with c3:
                if st.button("🗑️", key=f"delp_{p['id']}", use_container_width=True):
                    db.delete_protocol(p["id"]); st.rerun()

        st.markdown("---")
        st.markdown("**➕ Novo Protocolo**")
        with st.form("f_prot", clear_on_submit=True):
            pc1, pc2, pc3 = st.columns(3)
            with pc1:
                nome = st.text_input("Nome *", placeholder="Ex: Brucelose B19, Aftosa")
                sexo_t = st.selectbox("Sexo-alvo", list(db.SEX_TARGETS.keys()),
                    format_func=lambda s: db.SEX_TARGETS[s])
                freq = st.selectbox("Frequência", list(db.PROTOCOL_FREQUENCIES.keys()),
                    format_func=lambda f: db.PROTOCOL_FREQUENCIES[f])
            with pc2:
                a_min = st.number_input("Idade mínima (meses)", min_value=0, max_value=999, value=0)
                a_max = st.number_input("Idade máxima (meses)", min_value=0, max_value=999, value=999)
                carencia_p = st.number_input("Carência (dias)", min_value=0, max_value=180, value=0)
            with pc3:
                dose_v = st.number_input("Dose", min_value=0.0, value=2.0, step=0.5, format="%.2f")
                dose_kg = st.number_input("A cada X kg (0 = dose fixa)", min_value=0.0, value=0.0,
                    step=10.0, help="Ex: '1 ml a cada 50 kg' → dose=1, aqui=50. 0 = mesma dose p/ todos")
                dose_u = st.selectbox("Unidade", ["ml","dose","mg","comprimido"])
            ins_link = st.selectbox("Insumo do estoque (para projeção de doses)",
                [None]+insumos,
                format_func=lambda x: "— Sem vínculo —" if x is None else f"{x['name']} ({x['current_stock']:g} {x['unit']})")
            via = st.selectbox("Via", ROUTES)
            notas = st.text_input("Observações", placeholder="Opcional")
            if st.form_submit_button("✅ Criar Protocolo", type="primary", use_container_width=True):
                if not nome:
                    st.error("Informe o nome do protocolo.")
                elif a_max < a_min:
                    st.error("Idade máxima deve ser ≥ mínima.")
                else:
                    db.add_protocol(nome.strip(), sexo_t, a_min, a_max, dose_v, dose_kg, dose_u,
                        ins_link["id"] if ins_link else None, freq, carencia_p, via, notas)
                    st.success(f"✅ Protocolo '{nome}' criado!")
                    st.rerun()


def _render_tab_projecao_abate(animals):
    st.caption("Estimativa de quando cada animal atinge o **peso-alvo**, mantido o GMD recente. "
               "Também mostramos o **GMD total** (de vida) como referência da trajetória.")
    rows = []
    bulk_data = db.projecao_abate_bulk(animals)
    _SITUACAO_ROTULO = {
        "perdendo_peso": "⚠️ Perdendo peso",
        "sem_ganho": "— (sem GMD)",
    }
    for a in animals:
        data = bulk_data[a["id"]]
        p = data["projecao"]
        g_total = data["gmd_total"]
        data_estimada = p["data"] or _SITUACAO_ROTULO.get(p["situacao"], "— (sem GMD)")
        rows.append({"ID":a["id"],"Raça":a["breed"],
            "Peso Atual (kg)":a["current_weight"],
            "Peso-Alvo (kg)":a.get("target_weight") or 500,
            "Falta (kg)":p["falta"],
            "GMD recente":round(p["gmd"],3) if p["gmd"] else None,
            "GMD total":round(g_total,3) if g_total else None,
            "Dias p/ abate":p["dias"] if p["dias"] is not None else None,
            "Data estimada":data_estimada})
    df = pd.DataFrame(rows)
    st.dataframe(df, use_container_width=True, hide_index=True, height=420,
        column_config={
            "Peso Atual (kg)":st.column_config.NumberColumn(format="%.1f"),
            "Peso-Alvo (kg)":st.column_config.NumberColumn(format="%.0f"),
            "Falta (kg)":st.column_config.NumberColumn(format="%.1f"),
            "GMD recente":st.column_config.NumberColumn(format="%.3f"),
            "GMD total":st.column_config.NumberColumn(format="%.3f")})
    prontos = [r for r in rows if r["Dias p/ abate"] == 0]
    if prontos:
        st.success(f"🟢 {_plural(len(prontos),'animal já pronto','animais já prontos')} para abate "
                   f"(peso-alvo atingido).")
    perdendo = [r for r in rows if r["Data estimada"] == "⚠️ Perdendo peso"]
    if perdendo:
        st.warning(f"⚠️ {_plural(len(perdendo),'animal está','animais estão')} perdendo peso — "
                   f"diferente de faltar dado, é sinal de saúde/pasto/verminose a investigar.")


def _render_tab_comparativo_piquete():
    st.caption("GMD médio × investimento em nutrição de cada piquete. Um pasto com muito "
               "trato tende a ter **GMD maior**, mas também **custo por GMD maior** — "
               "aqui você compara a eficiência.")
    perf = db.get_performance_by_lote()
    if not perf:
        st.info("Sem dados por piquete ainda.")
    else:
        rows = [{"Piquete":f"{p['lote_id']} — {p['lote_name']}","Animais":p["n"],
                 "GMD médio (kg/dia)":p["gmd_medio"],
                 "Nutrição/animal (R$)":p["custo_nut_por_animal"],
                 "Custo por GMD (R$)":p["custo_por_gmd"]} for p in perf]
        df = pd.DataFrame(rows)
        st.dataframe(df, use_container_width=True, hide_index=True,
            column_config={
                "GMD médio (kg/dia)":st.column_config.NumberColumn(format="%.3f"),
                "Nutrição/animal (R$)":st.column_config.NumberColumn(format="R$ %.2f"),
                "Custo por GMD (R$)":st.column_config.NumberColumn(format="R$ %.2f")})
        fig = go.Figure()
        fig.add_bar(x=[p["lote_name"] for p in perf], y=[p["gmd_medio"] for p in perf],
            name="GMD médio", marker_color=c["primaria"], yaxis="y")
        fig.add_trace(go.Scatter(x=[p["lote_name"] for p in perf],
            y=[p["custo_nut_por_animal"] for p in perf], name="Nutrição/animal (R$)",
            mode="lines+markers", line=dict(color=c["atencao"],width=3), yaxis="y2"))
        fig.update_layout(**_layout(height=320,
            legend=dict(orientation="h",y=1.1),
            xaxis=dict(gridcolor=c["superficie"]),
            yaxis=dict(title="GMD (kg/dia)",gridcolor=c["superficie"]),
            yaxis2=dict(title="R$/animal",overlaying="y",side="right",showgrid=False)))
        st.plotly_chart(fig, use_container_width=True)
        if all(p["custo_nutricao"]==0 for p in perf):
            st.info("💡 O custo de nutrição por piquete começa a ser contabilizado a partir "
                    "das próximas confirmações de trato (na aba Trato do Modo Campo).")


def _render_tab_simulador_terminacao(animals):
    st.caption("Compare a viabilidade econômica de **terminar o boi** em pasto, "
               "semiconfinamento ou confinamento. Ajuste GMD, custo/dia e rendimento "
               "de cada estratégia — os valores são **editáveis** e salvos para as "
               "próximas simulações.")

    pesos = sorted(a["current_weight"] for a in animals)
    peso_medio = round(pesos[len(pesos)//2], 0) if pesos else 380.0
    metas = [a.get("target_weight") for a in animals if a.get("target_weight")]
    meta_pad = round(sum(metas)/len(metas), 0) if metas else 500.0
    try:
        arroba_pad = float(db.get_setting("preco_arroba", "300"))
    except (TypeError, ValueError):
        arroba_pad = 300.0

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        peso_atual = st.number_input("Peso atual (kg)", min_value=50.0, max_value=900.0,
            value=float(peso_medio), step=10.0,
            help="Padrão: peso mediano do rebanho ativo")
    with c2:
        peso_meta = st.number_input("Peso de abate (kg)", min_value=100.0, max_value=1000.0,
            value=float(meta_pad), step=10.0)
    with c3:
        preco_arroba = st.number_input("Preço da @ (R$)", min_value=0.0, max_value=2000.0,
            value=float(arroba_pad), step=5.0, format="%.2f",
            help="Preço do boi gordo por arroba na venda")
    with c4:
        custo_boi = st.number_input("Custo do boi magro (R$)", min_value=0.0,
            max_value=100000.0, value=0.0, step=50.0, format="%.2f",
            help="Opcional — aquisição/valor do animal hoje. Igual para todos os "
                 "cenários; deixe 0 para analisar só a etapa de terminação.")

    st.markdown("**Cenários** — edite GMD (kg/dia), custo/dia (R$) e rendimento de carcaça (%)")
    cen = db.get_terminacao_cenarios()
    df_cen = pd.DataFrame(cen)[["nome", "gmd", "custo_dia", "rendimento"]]
    edited = st.data_editor(df_cen, use_container_width=True, hide_index=True,
        num_rows="dynamic", key="term_editor",
        column_config={
            "nome": st.column_config.TextColumn("Estratégia"),
            "gmd": st.column_config.NumberColumn("GMD (kg/dia)", min_value=0.0,
                max_value=3.0, step=0.05, format="%.3f"),
            "custo_dia": st.column_config.NumberColumn("Custo/dia (R$)", min_value=0.0,
                step=0.5, format="R$ %.2f"),
            "rendimento": st.column_config.NumberColumn("Rendimento (%)", min_value=0.30,
                max_value=0.70, step=0.01, format="%.2f")})

    cA, cB = st.columns([1, 3])
    with cA:
        if st.button("💾 Salvar cenários", use_container_width=True):
            db.set_terminacao_cenarios(edited.to_dict("records"))
            db.set_setting("preco_arroba", round(preco_arroba, 2))
            st.success("Cenários e preço da @ salvos!"); st.rerun()

    cenarios = [r for r in edited.to_dict("records") if r.get("nome")]
    sim = db.simular_terminacao(peso_atual, peso_meta, preco_arroba, cenarios, custo_boi)

    if peso_meta - peso_atual <= 0:
        st.warning("O peso de abate precisa ser maior que o peso atual.")
    elif not any(s["dias"] for s in sim):
        st.info("Informe um GMD maior que zero em pelo menos um cenário.")
    else:
        ganho = round(peso_meta - peso_atual, 1)
        st.markdown(f"Ganho necessário: **{ganho:.0f} kg** por cabeça.")
        rows = [{"Estratégia":s["nome"],"Dias no trato":s["dias"],
                 "@ produzidas":s["arrobas_produzidas"],
                 "Custo alimentar (R$)":s["custo_alimentar"],
                 "Custo/@ produzida (R$)":s["custo_por_arroba"],
                 "Receita (R$)":s["receita"],"Lucro (R$)":s["lucro"],
                 "Lucro/dia (R$)":s["lucro_por_dia"],
                 "Margem (%)":s["margem"]} for s in sim]
        df = pd.DataFrame(rows)
        st.dataframe(df, use_container_width=True, hide_index=True,
            column_config={
                "@ produzidas":st.column_config.NumberColumn(format="%.2f"),
                "Custo alimentar (R$)":st.column_config.NumberColumn(format="R$ %.2f"),
                "Custo/@ produzida (R$)":st.column_config.NumberColumn(format="R$ %.2f"),
                "Receita (R$)":st.column_config.NumberColumn(format="R$ %.2f"),
                "Lucro (R$)":st.column_config.NumberColumn(format="R$ %.2f"),
                "Lucro/dia (R$)":st.column_config.NumberColumn(format="R$ %.2f"),
                "Margem (%)":st.column_config.NumberColumn(format="%.1f%%")})

        validos = [s for s in sim if s["lucro"] is not None]
        best = validos[0] if validos else None
        if best and best["viavel"]:
            st.success(f"🏆 Estratégia mais rentável: **{best['nome']}** — "
                       f"lucro de **R\\$ {best['lucro']:,.2f}** em **{best['dias']} dias** "
                       f"(R\\$ {best['lucro_por_dia']:,.2f}/dia).")
        elif best:
            st.warning(f"⚠️ Nenhuma estratégia dá lucro positivo com estes parâmetros. "
                       f"A menos ruim é **{best['nome']}** (R\\$ {best['lucro']:,.2f}).")

        fig = go.Figure()
        nomes = [s["nome"] for s in validos]
        fig.add_bar(x=nomes, y=[s["lucro"] for s in validos], name="Lucro (R$)",
            marker_color=[c["primaria"] if s["viavel"] else c["perigo"] for s in validos],
            yaxis="y")
        fig.add_trace(go.Scatter(x=nomes, y=[s["dias"] for s in validos],
            name="Dias no trato", mode="lines+markers",
            line=dict(color=c["atencao"], width=3), yaxis="y2"))
        fig.update_layout(**_layout(height=320, legend=dict(orientation="h", y=1.1),
            xaxis=dict(gridcolor=c["superficie"]),
            yaxis=dict(title="Lucro (R$)", gridcolor=c["superficie"], zeroline=True,
                zerolinecolor=c["borda_suave"]),
            yaxis2=dict(title="Dias", overlaying="y", side="right", showgrid=False)))
        st.plotly_chart(fig, use_container_width=True)
        st.caption("Receita = peso de abate × rendimento ÷ 15 × preço da @. "
                   "Lucro = receita − custo alimentar (dias × custo/dia) − custo do boi magro. "
                   "O confinamento costuma dar **mais lucro/dia** (gira o capital mais rápido), "
                   "mesmo com custo/dia maior; o pasto costuma ter **menor custo por @ produzida**.")


def _render_tab_correlacao_chuva_gmd():
    st.caption("Associação entre a chuva do mês e o GMD médio do rebanho no mesmo "
               "mês — usa todo o histórico de leituras de chuva e pesagens. "
               "Correlação não demonstra causalidade.")

    leituras = db.get_rain()
    pesagens = db.get_all_weighings()
    series = series_mensais(leituras, pesagens)

    if not series:
        st.info("Ainda não há mês com leitura de chuva **e** GMD calculável ao "
               "mesmo tempo — registre chuva (Clima) e pesagens no mesmo período.")
    else:
        resultado = correlacao_chuva_gmd(series)
        k1, k2 = st.columns(2)
        k1.metric("Coeficiente de correlação",
                 f"{resultado['coeficiente']:.2f}"
                 if resultado["coeficiente"] is not None else "—")
        k2.metric("Períodos avaliados", resultado["n"])
        st.info(f"ℹ️ {resultado['interpretacao']}")

        df_s = pd.DataFrame(series).sort_values("periodo")
        fig = go.Figure()
        fig.add_bar(x=df_s["periodo"], y=df_s["chuva_mm"], name="Chuva (mm)",
                   marker_color=c["primaria"], yaxis="y")
        fig.add_trace(go.Scatter(x=df_s["periodo"], y=df_s["gmd_medio"],
            name="GMD médio (kg/dia)", mode="lines+markers",
            line=dict(color=c["atencao"], width=3), yaxis="y2"))
        fig.update_layout(**_layout(height=320, legend=dict(orientation="h", y=1.1),
            xaxis=dict(gridcolor=c["superficie"], title="Mês"),
            yaxis=dict(title="Chuva (mm)", gridcolor=c["superficie"]),
            yaxis2=dict(title="GMD médio (kg/dia)", overlaying="y", side="right",
                       showgrid=False)))
        st.plotly_chart(fig, use_container_width=True)

        st.dataframe(df_s.rename(columns={
            "periodo": "Mês", "chuva_mm": "Chuva (mm)",
            "gmd_medio": "GMD médio (kg/dia)"}),
            use_container_width=True, hide_index=True)


def _render_tab_meta_gmd():
    meta_atual = db.get_gmd_target()
    c1, c2 = st.columns([1,2])
    with c1:
        nova = st.number_input("Meta de GMD (kg/dia)", min_value=0.0, max_value=3.0,
            value=float(meta_atual), step=0.05, format="%.3f",
            help="Ganho médio diário mínimo esperado")
        if st.button("💾 Salvar meta", use_container_width=True):
            db.set_setting("gmd_meta", round(nova, 3))
            st.success("Meta salva!"); st.rerun()
    with c2:
        st.caption("Animais com GMD **abaixo da meta** são candidatos a investigação "
                   "(saúde, verminose, pasto ruim) ou descarte. A meta também aparece "
                   "como alerta na página **Alertas**.")

    low = db.get_low_performance(meta_atual)
    st.markdown(f"**{_plural(len(low),'animal','animais')} abaixo da meta "
                f"({meta_atual:.3f} kg/dia)**")
    if low:
        rows = [{"ID":a["id"],"Raça":a["breed"],
                 "Categoria":db.get_age_category(a.get("birth_date")),
                 "Lote":a.get("lote_id") or "—","Peso (kg)":a["current_weight"],
                 "GMD (kg/dia)":round(a["gmd"],3)} for a in low]
        df = pd.DataFrame(rows)
        st.dataframe(df, use_container_width=True, hide_index=True,
            column_config={"Peso (kg)":st.column_config.NumberColumn(format="%.1f"),
                "GMD (kg/dia)":st.column_config.NumberColumn(format="%.3f")})
        fig = px.bar(df.sort_values("GMD (kg/dia)"), x="GMD (kg/dia)", y="ID",
            orientation="h", color="GMD (kg/dia)",
            color_continuous_scale=ESCALA_RUIM_BOM)
        fig.add_vline(x=meta_atual, line_dash="dash", line_color=c["primaria"],
            annotation_text="Meta", annotation_position="top")
        fig.update_layout(**PLOTLY, height=max(180,len(df)*30), coloraxis_showscale=False,
            xaxis=dict(gridcolor=c["superficie"]), yaxis=dict(gridcolor=c["superficie"],title=""))
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.success("✅ Todos os animais estão na meta ou acima!")


def page_desempenho():
    if st.session_state.user["role"] != "admin":
        st.error("🔒 Acesso restrito ao Administrador."); return
    st.markdown('<div class="page-title">📈 Desempenho do Rebanho</div>', unsafe_allow_html=True)
    animals = db.get_all_animals()
    if not animals:
        st.info("Sem animais ativos."); return

    dt1, dt2, dt3, dt4, dt5 = st.tabs(
        ["🎯 Meta & Baixo Desempenho", "📅 Projeção de Abate",
         "🌿 Comparativo por Piquete", "🐂 Simulador de Terminação", "🌧️ Chuva × GMD"])

    # ── Meta de GMD e baixo desempenho ────────────────────────────────────────
    with dt1:
        _render_tab_meta_gmd()

    # ── Projeção de abate ─────────────────────────────────────────────────────
    with dt2:
        _render_tab_projecao_abate(animals)

    # ── Comparativo por piquete ───────────────────────────────────────────────
    with dt3:
        _render_tab_comparativo_piquete()

    # ── Simulador de terminação ───────────────────────────────────────────────
    with dt4:
        _render_tab_simulador_terminacao(animals)

    # ── Correlação chuva × GMD ────────────────────────────────────────────────
    with dt5:
        _render_tab_correlacao_chuva_gmd()


def page_nutricao():
    if st.session_state.user["role"]!="admin":
        st.error("🔒 Acesso restrito ao Administrador."); return
    st.markdown('<div class="page-title">🌾 Nutrição — Plano de Trato por Piquete</div>',
                unsafe_allow_html=True)
    st.caption("Defina o que cada piquete recebe (silagem, ração, massa de soja, sal mineral...) "
               "e a frequência. Os operadores confirmam a aplicação no Modo Campo.")

    lotes = db.get_all_lotes()
    insumos = db.get_all_insumos()

    nt1, nt2, nt3, nt4, nt5 = st.tabs(["📋 Planos Ativos", "➕ Novo Item de Trato",
                                       "✅ Histórico de Checagens", "💰 Custo por Piquete",
                                       "🕘 Histórico da Dieta"])

    with nt1:
        plans_e_encerrados = db.get_feeding_plans(active_only=False)
        # "Planos Ativos" mostra a versão CORRENTE de cada item (vigente,
        # ativa ou pausada) — versão encerrada (vigente_ate preenchido) só
        # aparece na aba "🕘 Histórico da Dieta", senão as duas telas ficam
        # mostrando a mesma coisa com nomes diferentes.
        plans = [p for p in plans_e_encerrados if p.get("vigente_ate") is None]
        if not plans:
            st.info("Nenhum plano de nutrição cadastrado. Use a aba **Novo Item de Trato**.")
        else:
            # Agrupa por piquete
            lotes_com_plano = sorted({p["lote_id"] for p in plans})
            for lid in lotes_com_plano:
                lote_nome = next((l["name"] for l in lotes if l["id"]==lid), lid)
                itens = [p for p in plans if p["lote_id"]==lid]
                st.markdown(f"#### 🌿 {lid} — {lote_nome}")
                for p in itens:
                    freq = db.FEEDING_FREQUENCIES.get(p["frequency"], p["frequency"])
                    ativo = "🟢 ativo" if p["active"] else "⚪ pausado"
                    c1, c2, c3 = st.columns([5,1,1])
                    with c1:
                        st.markdown(
                            f'<div class="hist-item">'
                            f'<b>{p["product_name"]}</b> — {p["quantity"]:.0f} {p["unit"]} '
                            f'· <span style="color:{c["primaria"]}">{freq}</span> · {ativo} '
                            f'· desde {p["vigente_de"]}'
                            f'{"  · vinc. estoque: "+p["insumo_name"] if p.get("insumo_name") else ""}'
                            f'</div>', unsafe_allow_html=True)
                    with c2:
                        novo = 0 if p["active"] else 1
                        if st.button("Ativar" if not p["active"] else "Pausar",
                                     key=f"tgl_{p['id']}", use_container_width=True,
                                     help="Pausa/retoma esta mesma versão — não conta como mudança "
                                          "de dieta."):
                            db.set_feeding_plan_active(p["id"], novo); st.rerun()
                    with c3:
                        if st.button("🔚 Encerrar", key=f"enc_{p['id']}", use_container_width=True,
                                     help="Fecha a vigência deste item. Diferente de excluir: o "
                                          "histórico de custo continua reconstruível."):
                            db.encerrar_feeding_plan(p["id"]); st.rerun()

                    with st.expander(f"✏️ Nova versão — {p['product_name']}"):
                        st.caption("Muda a quantidade/frequência a partir de **hoje**, sem apagar "
                                  "o que valeu até ontem — o custo já calculado com a versão "
                                  "anterior não muda retroativamente.")
                        with st.form(f"f_versao_{p['id']}"):
                            nv1, nv2 = st.columns(2)
                            with nv1:
                                nv_qtd = st.number_input("Nova quantidade", min_value=0.0,
                                    value=float(p["quantity"]), step=1.0, format="%.1f",
                                    key=f"nv_qtd_{p['id']}")
                            with nv2:
                                freqs = list(db.FEEDING_FREQUENCIES.keys())
                                nv_freq = st.selectbox("Nova frequência", freqs,
                                    index=freqs.index(p["frequency"]) if p["frequency"] in freqs else 0,
                                    format_func=lambda f: db.FEEDING_FREQUENCIES[f],
                                    key=f"nv_freq_{p['id']}")
                            if st.form_submit_button("💾 Salvar nova versão", type="primary"):
                                if nv_qtd <= 0:
                                    st.error("A quantidade deve ser maior que zero.")
                                else:
                                    r = db.nova_versao_feeding_plan(
                                        p["id"], quantity=nv_qtd, frequency=nv_freq)
                                    if r["ok"]:
                                        st.success("✅ Nova versão salva.")
                                        st.rerun()
                                    else:
                                        st.error(f"❌ {r['erro']}")

    with nt2:
        if not lotes:
            st.warning("Cadastre piquetes primeiro (em Lotes / Pastagem).")
        else:
            with st.form("f_plan", clear_on_submit=True):
                fp1, fp2 = st.columns(2)
                with fp1:
                    lote_sel = st.selectbox("Piquete *", lotes,
                        format_func=lambda l: f"{l['id']} — {l['name']}")
                    prod = st.text_input("Produto *", placeholder="Ex: Silagem de milho")
                    freq = st.selectbox("Frequência *", list(db.FEEDING_FREQUENCIES.keys()),
                        format_func=lambda f: db.FEEDING_FREQUENCIES[f])
                with fp2:
                    qtd = st.number_input("Quantidade *", min_value=0.0, step=5.0, format="%.1f")
                    unid = st.selectbox("Unidade", ["kg","ton","saco","litro","g"])
                    ins_link = st.selectbox("Vincular a insumo (opcional)",
                        [None]+insumos,
                        format_func=lambda x: "— Sem vínculo —" if x is None else f"{x['name']} ({x['current_stock']:.0f} {x['unit']})",
                        help="Se vinculado, a confirmação do operador pode baixar do estoque")
                notes = st.text_input("Observações", placeholder="Opcional")
                if st.form_submit_button("✅ Adicionar ao Plano", type="primary", use_container_width=True):
                    if not prod or qtd<=0:
                        st.error("Informe o produto e a quantidade.")
                    else:
                        db.add_feeding_plan(lote_sel["id"], prod.strip(), qtd, unid, freq,
                            insumo_id=ins_link["id"] if ins_link else None, notes=notes)
                        st.success(f"✅ {prod} adicionado ao {lote_sel['name']} ({db.FEEDING_FREQUENCIES[freq]})")
                        st.rerun()

    with nt3:
        st.markdown("**Checagens registradas pelos operadores**")
        cc1, cc2 = st.columns(2)
        with cc1:
            start_c = st.date_input("De", value=date.today()-timedelta(days=30), key="chk_start")
        with cc2:
            end_c = st.date_input("Até", value=date.today(), key="chk_end")
        checks = db.get_feeding_checks(start_date=start_c.isoformat(), end_date=end_c.isoformat())
        if checks:
            df_c = pd.DataFrame(checks)[["check_date","lote_id","product_name","status","actual_quantity","operator"]].copy()
            df_c["status"] = df_c["status"].map(lambda s: db.FEEDING_CHECK_STATUS.get(s,s))
            df_c.columns = ["Data","Piquete","Produto","Status","Qtd Aplicada","Operador"]
            st.dataframe(df_c, use_container_width=True, hide_index=True)
        else:
            st.info("Nenhuma checagem registrada no período.")

    with nt4:
        _nutricao_custo_por_piquete(lotes, insumos)

    with nt5:
        _nutricao_historico_da_dieta(lotes)


def _nutricao_historico_da_dieta(lotes):
    """Linha do tempo da dieta de cada piquete (§5, Trilha 3) —
    `db.get_feeding_plan_historico`: toda versão de todo item, vigente ou
    encerrada, mais recente primeiro. É o que 'nova versão'/'encerrar' (aba
    📋 Planos Ativos) constroem, versão a versão."""
    st.caption("Toda mudança de quantidade ou frequência vira uma versão nova, nunca uma "
               "edição por cima da anterior — o custo já calculado com a versão antiga não "
               "muda quando a dieta muda.")

    if not lotes:
        st.info("Nenhum piquete cadastrado.")
        return

    lote_sel = st.selectbox("Piquete", lotes, format_func=lambda l: f"{l['id']} — {l['name']}",
                            key="hist_dieta_lote")
    historico = db.get_feeding_plan_historico(lote_sel["id"])
    if not historico:
        st.info("Nenhum item de trato foi cadastrado para este piquete ainda.")
        return

    for produto in sorted({h["product_name"] for h in historico}):
        versoes = [h for h in historico if h["product_name"] == produto]
        with st.expander(f"🌿 {produto} — {len(versoes)} "
                         f"{'versão' if len(versoes)==1 else 'versões'}",
                         expanded=any(v["vigente_ate"] is None for v in versoes)):
            rows = [{
                "Vigência": f"{v['vigente_de']} → {v['vigente_ate'] or 'hoje'}",
                "Quantidade": f"{v['quantity']:.1f} {v['unit']}",
                "Frequência": db.FEEDING_FREQUENCIES.get(v["frequency"], v["frequency"]),
                "Situação": "🟢 Vigente" if v["vigente_ate"] is None else "⚪ Encerrada",
            } for v in versoes]
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)


def _nutricao_custo_por_piquete(lotes, insumos):
    """Custo de dieta por piquete (spec 0037 + `services/dieta.py`, nunca chamado).

    `services.dieta.custo_por_cabeca_dia` espera ingredientes já por cabeça;
    `services.dieta_adaptador.ingredientes_por_cabeca` (spec 0037) é a ponte
    a partir de `feeding_plans`, que são por piquete.
    """
    st.caption("Custo diário de trato por cabeça, a partir dos planos **ativos** de "
               "cada piquete. Matéria seca aparece zerada — a coluna não existe no "
               "schema hoje (fora do escopo desta integração, decisão registrada no "
               "ROADMAP); custo em R$ não é afetado.")

    insumos_por_id = {i["id"]: i for i in insumos}
    lotes_com_plano_ativo = sorted({
        p["lote_id"] for p in db.get_feeding_plans(active_only=True)
    })
    if not lotes_com_plano_ativo:
        st.info("Nenhum piquete com plano de trato ativo.")
        return

    for lid in lotes_com_plano_ativo:
        lote_nome = next((l["name"] for l in lotes if l["id"] == lid), lid)
        animais_lote = db.get_all_animals(status="ativo", lote_id=lid)
        cabecas = len(animais_lote)
        planos = db.get_feeding_plans(lote_id=lid, active_only=True)

        ingredientes = ingredientes_por_cabeca(
            planos, insumos_por_id, cabecas, converter_quantidade=db.convert_quantity)
        resultado = custo_por_cabeca_dia(ingredientes)

        with st.expander(f"🌿 {lid} — {lote_nome} · {cabecas} cabeça(s) · "
                         f"R$ {resultado['custo_dia']:.2f}/cabeça/dia"):
            if cabecas == 0:
                st.warning("Piquete sem animais ativos — custo por cabeça não se "
                          "aplica (os planos continuam ativos, mas ninguém consome).")
                continue
            if not ingredientes:
                st.info("Nenhum plano deste piquete pôde ser convertido para a "
                       "unidade do insumo vinculado — confira o vínculo insumo↔plano "
                       "e a frequência de cada item.")
                continue

            gmds = db.calculate_gmd_bulk([a["id"] for a in animais_lote])
            validos = [g for g in gmds.values() if g is not None]
            gmd_medio = sum(validos) / len(validos) if validos else 0.0
            rendimento_medio = sum(
                a.get("carcass_yield") or 0.52 for a in animais_lote) / cabecas
            custo_arroba = custo_por_arroba_produzida(
                resultado["custo_dia"], gmd_medio, rendimento_medio)

            k1, k2, k3, k4 = st.columns(4)
            k1.metric("Custo/cabeça/dia", f"R$ {resultado['custo_dia']:.2f}")
            k2.metric("Matéria natural/dia", f"{resultado['kg_materia_natural']:.2f} kg")
            k3.metric("Matéria seca/dia", f"{resultado['kg_materia_seca']:.2f} kg")
            k4.metric("Custo/@ produzida",
                     f"R$ {custo_arroba:.2f}" if custo_arroba is not None else "—",
                     help="Precisa de GMD positivo no piquete para calcular.")

            if resultado["participacao"]:
                df_p = pd.DataFrame(resultado["participacao"])
                df_p.columns = ["Ingrediente", "Participação (%)"]
                fig = px.pie(df_p, names="Ingrediente", values="Participação (%)", hole=0.45)
                fig.update_layout(**PLOTLY, height=260)
                st.plotly_chart(fig, use_container_width=True)


# Rótulos dos doze estados do §5.2. O código é o que vai ao banco; o rótulo é o
# que o operador entende no curral.
_ESTADO_BRINCO = {
    "solicitado": "Solicitado ao fornecedor", "recebido": "Recebido, a conferir",
    "disponivel": "Disponível", "reservado": "Reservado",
    "aplicado": "Aplicado em animal", "perdido": "Perdido",
    "danificado": "Danificado", "substituido": "Substituído",
    "inutilizado": "Inutilizado (definitivo)", "devolvido": "Devolvido (definitivo)",
    "cancelado": "Cancelado (definitivo)", "bloqueado_orgao": "Bloqueado pelo órgão",
}
_TIPO_BRINCO = {"brinco_visual": "Brinco visual", "boton": "Botton",
                "conjunto": "Conjunto (visual + eletrônico)", "outro": "Outro"}


def page_brincos():
    """Estoque de dispositivos de identificação (PNIB §5).

    Página própria, e não uma aba do Estoque de Insumos, porque as duas coisas
    só se parecem no nome: insumo acaba e se repõe; brinco é um número
    controlado, com doze estados e ato definitivo no fim. Um alerta de estoque
    mínimo não diz nada sobre um brinco inutilizado.
    """
    st.markdown('<div class="page-title">🏷️ Brincos e Dispositivos</div>',
                unsafe_allow_html=True)

    inv = db.dispositivos.inventario()
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Em estoque", inv["em_estoque"])
    c2.metric("Aplicados", inv["aplicados"])
    c3.metric("Perdidos / danificados", inv["perdidos_ou_danificados"])
    c4.metric("Divergências", len(inv["divergencias"]))

    if inv["divergencias"]:
        st.warning(
            f"⚠️ {_plural(len(inv['divergencias']), 'dispositivo aplicado', 'dispositivos aplicados')} "
            "com divergência entre o código visual e o eletrônico. "
            "Foi registrado na aplicação e **não bloqueou** o trabalho (§5.3) — "
            "mas precisa de conferência no campo.")

    t_inv, t_aplicar, t_import, t_arquivo = st.tabs(
        ["📋 Inventário", "🏷️ Aplicar em animal", "📥 Importar lote",
         "📄 Importar arquivo"])

    with t_inv:
        _brincos_inventario(inv)
    with t_aplicar:
        _brincos_aplicar()
    with t_import:
        _brincos_importar()
    with t_arquivo:
        _brincos_importar_arquivo()


def _brincos_inventario(inv):
    if not inv["por_status"]:
        st.info("Nenhum dispositivo cadastrado. Comece pela aba **Importar lote**.")
        return

    st.markdown("**Por situação**")
    st.dataframe(
        pd.DataFrame([{"Situação": _ESTADO_BRINCO.get(k, k), "Quantidade": v}
                      for k, v in sorted(inv["por_status"].items(),
                                         key=lambda kv: -kv[1])]),
        use_container_width=True, hide_index=True)

    if inv["por_lote"]:
        st.markdown("**Por lote de compra**")
        st.dataframe(
            pd.DataFrame([{"Lote": l["lote"], "Total": l["total"],
                           "Ainda em estoque": l["em_estoque"]}
                          for l in inv["por_lote"]]),
            use_container_width=True, hide_index=True)

    st.markdown("---")
    st.markdown("**Mudar a situação de um dispositivo**")
    st.caption("Busque pelo número gravado no brinco — é como ele aparece na mão "
               "do operador.")

    codigo = st.text_input("🔎 Código do brinco", key="brc_busca").strip().upper()
    if not codigo:
        return

    d = db.dispositivos.por_codigo(codigo)
    if d is None:
        st.info(f"Nenhum dispositivo ativo com o código **{codigo}**. "
                "Inutilizado, devolvido ou cancelado não aparece nesta busca — "
                "o número deles não volta ao estoque.")
        return

    atual = d["status"]
    st.markdown(f"**{d['codigo_visual']}** · {_TIPO_BRINCO.get(d['tipo'], d['tipo'])} · "
                f"situação atual: **{_ESTADO_BRINCO.get(atual, atual)}**"
                + (f" · lote {d['lote']}" if d.get("lote") else ""))

    destinos = [e for e in ESTADOS_DISPOSITIVO
                if e != atual and _transicao_dispositivo(atual, e)["permitida"]]
    if not destinos:
        # Nada permitido tem duas causas muito diferentes, e confundi-las seria
        # dizer ao operador que o sistema quebrou quando na verdade a norma
        # está sendo cumprida.
        if _transicao_dispositivo(atual, "disponivel").get("exige_autorizacao"):
            st.error("🔒 Dispositivo bloqueado pelo órgão oficial. "
                     "Só o órgão libera — o sistema não desfaz isso sozinho (§5.2).")
        else:
            st.info(f"**{_ESTADO_BRINCO.get(atual, atual)}** é situação definitiva: "
                    "não admite mudança. É o que garante que este número não será "
                    "reaplicado.")
        return

    novo = st.selectbox("Nova situação", destinos,
                        format_func=lambda e: _ESTADO_BRINCO.get(e, e),
                        key="brc_novo")
    regra = _transicao_dispositivo(atual, novo)
    motivo = ""
    if regra["exige_motivo"]:
        motivo = st.text_input(
            "Motivo *", key="brc_motivo",
            help="Sem o motivo ninguém reconstrói depois por que um brinco pago "
                 "virou refugo.").strip()

    pode = not regra["exige_motivo"] or bool(motivo)
    if st.button("💾 Registrar mudança", type="primary", disabled=not pode,
                 key="brc_salvar"):
        r = db.dispositivos.mudar_status(
            d["id"], novo, motivo=motivo, usuario=st.session_state.user["name"])
        if r.get("ok"):
            db.clear_cache()
            st.success(f"✅ {d['codigo_visual']}: {_ESTADO_BRINCO.get(atual, atual)}"
                       f" → {_ESTADO_BRINCO.get(novo, novo)}")
            st.rerun()
        else:
            st.error(f"🚫 {r.get('erro') or r.get('motivo', 'Não foi possível mudar.')}")


def _brincos_aplicar():
    disp = db.dispositivos.disponiveis(limite=200)
    if not disp:
        st.info("Nenhum dispositivo disponível para aplicar. "
                "Importe um lote ou libere os que estão reservados.")
        return

    animais = db.get_all_animals(status="ativo")
    if not animais:
        st.info("Nenhum animal ativo para receber o dispositivo.")
        return

    c1, c2 = st.columns(2)
    with c1:
        rot_d = {f"{d['codigo_visual']} — {_TIPO_BRINCO.get(d['tipo'], d['tipo'])}": d
                 for d in disp}
        dispositivo = rot_d[st.selectbox("🏷️ Dispositivo *", list(rot_d),
                                         key="brap_disp")]
    with c2:
        rot_a = {f"{a['id']} — {a['breed']}": a for a in animais}
        animal = rot_a[st.selectbox("🐄 Animal *", list(rot_a), key="brap_animal")]

    c3, c4 = st.columns(2)
    with c3:
        tipo_id = st.selectbox(
            "Função do identificador", ["manejo", "oficial", "eletronico"],
            key="brap_tipo",
            format_func=lambda t: {"manejo": "Manejo (uso diário)",
                                   "oficial": "Oficial (§4.1)",
                                   "eletronico": "Eletrônico"}[t])
    with c4:
        lido = st.text_input(
            "Código lido no leitor", key="brap_lido",
            help="Opcional. Confere visual × eletrônico na hora (§5.3). "
                 "Divergência é registrada, não impede — pode ser erro de "
                 "leitura, e recusar travaria o trabalho no curral.").strip().upper()

    if lido and lido != dispositivo["codigo_visual"]:
        st.warning(f"⚠️ O leitor devolveu **{lido}**, e o brinco é "
                   f"**{dispositivo['codigo_visual']}**. A aplicação segue, e a "
                   "divergência fica registrada para conferência.")

    # Já existe identificador vigente desta função? Então isto é TROCA, e o
    # §4.2.3 exige motivo. Perguntar aqui evita o erro voltar depois de o
    # operador achar que gravou.
    vigente = db.identificadores.get_ativo(animal["uuid"], tipo_id)
    troca = bool(vigente and vigente["valor"] != dispositivo["codigo_visual"])
    motivo_sub = ""
    if troca:
        st.info(f"O animal já tem {tipo_id} **{vigente['valor']}**. Aplicar outro "
                "é **substituição**: o anterior é encerrado, não apagado (§4.2.3).")
        motivo_sub = st.text_input("Motivo da substituição *",
                                   key="brap_motivo").strip()

    pode = not troca or bool(motivo_sub)
    if st.button("✅ Aplicar dispositivo", type="primary", disabled=not pode,
                 key="brap_salvar"):
        r = db.dispositivos.aplicar(
            dispositivo["id"], animal["uuid"],
            aplicador=st.session_state.user["name"],
            tipo_identificador=tipo_id, eletronico_lido=lido,
            motivo_substituicao=motivo_sub)
        if r.get("ok"):
            db.clear_cache()
            if r.get("divergencia"):
                st.warning(f"✅ Aplicado com divergência registrada: {r['divergencia']}")
            else:
                st.success(f"✅ {dispositivo['codigo_visual']} aplicado em "
                           f"{animal['id']}.")
            st.rerun()
        else:
            st.error(f"🚫 {r.get('erro', 'Não foi possível aplicar.')}")


def _brincos_importar():
    st.caption("§5.3: brincos chegam em faixa numérica contínua. Reimportar a "
               "mesma faixa **pula** os números já cadastrados — não duplica nem "
               "apaga histórico de aplicação.")

    c1, c2, c3 = st.columns(3)
    with c1:
        inicio = st.text_input("Do número *", placeholder="BR0001",
                               key="brimp_ini").strip().upper()
    with c2:
        fim = st.text_input("Até o número *", placeholder="BR0500",
                            key="brimp_fim").strip().upper()
    with c3:
        lote = st.text_input("Lote de compra *", placeholder="NF 1234",
                             key="brimp_lote").strip()

    c4, c5, c6 = st.columns(3)
    with c4:
        tipo = st.selectbox("Tipo", list(_TIPO_BRINCO),
                            format_func=lambda t: _TIPO_BRINCO[t], key="brimp_tipo")
    with c5:
        fabricante = st.text_input("Fabricante", key="brimp_fab").strip()
    with c6:
        aquisicao = st.date_input("Data de aquisição", value=date.today(),
                                  key="brimp_data")

    pode = bool(inicio and fim and lote)
    if st.button("📥 Importar faixa", type="primary", disabled=not pode,
                 key="brimp_salvar"):
        r = db.dispositivos.importar_lote(
            inicio, fim, lote=lote, tipo=tipo, fabricante=fabricante,
            data_aquisicao=aquisicao.isoformat(),
            usuario=st.session_state.user["name"])
        if r.get("ok"):
            db.clear_cache()
            msg = f"✅ {r['criados']} de {r['total_na_faixa']} importados."
            if r["pulados"]:
                msg += f" {r['pulados']} já existiam e foram mantidos como estão."
            st.success(msg)
            st.rerun()
        else:
            st.error(f"🚫 {r['erro']}")


def _brincos_importar_arquivo():
    """Importa um lote de códigos arbitrários vindo de arquivo do fornecedor.

    Diferente da aba "Importar lote" (faixa numérica contígua): aqui o
    arquivo pode trazer qualquer conjunto de códigos, então a garantia de
    não duplicar precisa reconciliar contra **todo** o estoque, em qualquer
    situação (spec 0033) — não só a faixa que está sendo importada.
    """
    st.caption("Arquivo do fornecedor (CSV), uma linha por dispositivo — "
               "`codigo_visual` obrigatório; `codigo_eletronico`, `tipo`, "
               "`fabricante`, `modelo` e `data_fabricacao` quando existirem no "
               "arquivo. Aceita `;` ou `,`, com ou sem cabeçalho.")

    arquivo = st.file_uploader("Arquivo do fornecedor", type=["csv", "txt"],
                               key="brarq_upload")
    if arquivo is None:
        return

    bruto = arquivo.getvalue()
    try:
        texto = bruto.decode("utf-8-sig")
    except UnicodeDecodeError:
        texto = bruto.decode("latin-1")

    lido = arquivo_dispositivos_ler(texto)
    aceitos, rejeitados = lido["aceitos"], lido["rejeitados"]

    m1, m2, m3 = st.columns(3)
    m1.metric("Linhas lidas", lido["total_linhas"])
    m2.metric("Aceitas", len(aceitos))
    m3.metric("Rejeitadas", len(rejeitados))

    if rejeitados:
        st.error(f"{len(rejeitados)} linha(s) não serão importadas:")
        st.dataframe(pd.DataFrame([
            {"Linha": r["linha"], "Motivo": r["motivo"], "Conteúdo": r["conteudo"]}
            for r in rejeitados
        ]), use_container_width=True, hide_index=True)

    if not aceitos:
        st.info("Nada a reconciliar.")
        return

    codigos = db.dispositivos.codigos_em_estoque()
    reconciliado = dispositivos_reconciliar(aceitos, codigos)
    para_gravar = reconciliado["para_gravar"]
    ja_existentes = reconciliado["ja_existentes"]

    r1, r2 = st.columns(2)
    r1.metric("Novos (a gravar)", len(para_gravar))
    r2.metric("Já em estoque", len(ja_existentes))

    if ja_existentes:
        with st.expander(
            f"Ver os {len(ja_existentes)} já cadastrados (não serão duplicados)"):
            st.dataframe(pd.DataFrame([
                {"Código": i["codigo_visual"],
                 "Situação atual": _ESTADO_BRINCO.get(
                     i["status_atual"], i["status_atual"])}
                for i in ja_existentes
            ]), use_container_width=True, hide_index=True)

    if not para_gravar:
        st.info("Todos os códigos do arquivo já estão cadastrados — nada novo a gravar.")
        return

    divergencias = arquivo_dispositivos_conferir_pareamento(para_gravar)
    if divergencias:
        st.warning(
            f"⚠️ {_plural(len(divergencias), 'dispositivo', 'dispositivos')} com "
            "possível divergência entre código visual e eletrônico no arquivo — "
            "não bloqueia a importação (§5.3), mas vale conferir antes.")

    st.markdown(f"**Prévia dos {len(para_gravar)} novos**")
    st.dataframe(pd.DataFrame([
        {"Código": i["codigo_visual"],
         "Eletrônico": i.get("codigo_eletronico") or "—",
         "Tipo": _TIPO_BRINCO.get(i.get("tipo"), i.get("tipo") or "—"),
         "Fabricante": i.get("fabricante") or "—"}
        for i in para_gravar
    ]), use_container_width=True, hide_index=True)

    lote = st.text_input("Lote de compra *", placeholder="NF 1234",
                         key="brarq_lote").strip()
    if st.button(f"📥 Importar {len(para_gravar)} novo(s)", type="primary",
                 disabled=not lote, key="brarq_salvar"):
        r = db.dispositivos.importar_arquivo(
            para_gravar, lote=lote, usuario=st.session_state.user["name"])
        db.clear_cache()
        st.success(f"✅ {r['criados']} dispositivo(s) importado(s) de {arquivo.name}.")
        st.rerun()

_TIPO_MOV = {
    "entre_propriedades_mesmo_titular": "Entre propriedades (mesmo titular)",
    "entre_titulares_diferentes": "Entre titulares diferentes",
    "venda": "Venda", "compra": "Compra", "remate": "Remate",
    "emprestimo": "Empréstimo", "parceria": "Parceria", "exposicao": "Exposição",
    "evento_agropecuario": "Evento agropecuário", "retorno": "Retorno",
    "frigorifico": "Frigorífico (abate)", "confinamento": "Confinamento",
    "temporaria": "Temporária", "sanitaria": "Sanitária", "outra": "Outra",
}
_STATUS_MOV = {"rascunho": "📝 Rascunho", "liberada": "🚚 Liberada",
               "em_transito": "🛣️ Em trânsito", "concluida": "✅ Concluída",
               "divergente": "⚠️ Concluída com divergência",
               "cancelada": "🚫 Cancelada"}


def page_movimentacao():
    """Movimentação entre propriedades, com GTA (PNIB §8).

    Não se confunde com o trânsito piquete→piquete do Modo Campo, que continua
    sendo manejo interno. Esta tem GTA, titular, transportador e confirmação de
    chegada — e é a que o órgão enxerga.
    """
    st.markdown('<div class="page-title">🚚 Movimentação entre Propriedades</div>',
                unsafe_allow_html=True)

    props = db.propriedades.listar()
    if len(props) < 2:
        st.info("É preciso ter **duas ou mais propriedades cadastradas** para "
                "movimentar entre elas. Cadastre em ⚙️ Admin → Propriedades. "
                "Para trocar de piquete dentro da mesma fazenda, use o 📱 Modo Campo.")
        return

    abertas = db.movimentacoes.abertas()
    t_fila, t_nova = st.tabs(
        [f"📋 Em andamento ({len(abertas)})", "➕ Nova movimentação"])

    with t_fila:
        _mov_fila(abertas, props)
    with t_nova:
        _mov_nova(props)


def _rotulo_prop(props, pid):
    for p in props:
        if p["id"] == pid:
            return p["nome"]
    return "—"


def _mov_fila(abertas, props):
    if not abertas:
        st.info("Nenhuma movimentação em andamento. As concluídas saem desta fila.")
        return

    rotulos = {}
    for m in abertas:
        rotulos[f"{_STATUS_MOV.get(m['status'], m['status'])} · "
                f"{_TIPO_MOV.get(m['tipo'], m['tipo'])} · "
                f"{_rotulo_prop(props, m['propriedade_origem_id'])} → "
                f"{_rotulo_prop(props, m['propriedade_destino_id'])} · "
                f"{m.get('data_prevista') or 'sem data'}"] = m

    escolha = st.selectbox("Movimentação", list(rotulos), key="mov_sel")
    mov = db.movimentacoes.get(rotulos[escolha]["id"])
    if mov is None:
        st.error("Movimentação não encontrada."); return

    st.markdown(f"**GTA:** {mov.get('gta_numero') or '— não informada —'} · "
                f"**Transportador:** {mov.get('transportador') or '—'} · "
                f"**Animais:** {len(mov['animais'])}")

    _mov_conferir_gta(mov, props)

    if mov["status"] == "rascunho":
        _mov_liberar(mov)
    else:
        _mov_confirmar_chegada(mov)


def _mov_conferir_gta(mov, props):
    """Confere o documento físico da GTA contra o que o sistema já sabe (spec 0038).

    `services.gta.validar` checa emissão/validade/quantidade declarada/UF —
    dados que só existem no PAPEL da GTA; `movimentacoes` não guarda isso, e
    a spec 0038 deixou a decisão de onde coletar para o mantenedor (R31).
    Optou-se por **não persistir**: a conferência é ad-hoc, feita na hora,
    porque nem toda movimentação tem o papel em mãos no momento do cadastro.

    Não bloqueia liberar/confirmar — é apoio à decisão, não substitui a
    pré-validação do §8.3 (`movimentacoes.pre_validar`, que já bloqueia).
    """
    with st.expander("📄 Conferir a GTA física (§8)"):
        st.caption("Digite o que está escrito no papel da GTA na hora de conferir — "
                   "não fica salvo, é conferência pontual, não um campo do cadastro.")

        c1, c2, c3 = st.columns(3)
        with c1:
            tem_emissao = st.checkbox("Tenho a data de emissão",
                                      key=f"gta_hasE_{mov['id']}")
            emissao = st.date_input("Emissão", value=date.today(),
                                    key=f"gta_emissao_{mov['id']}",
                                    disabled=not tem_emissao)
        with c2:
            tem_validade = st.checkbox("Tenho a data de validade",
                                       key=f"gta_hasV_{mov['id']}")
            validade = st.date_input("Validade", value=date.today(),
                                     key=f"gta_validade_{mov['id']}",
                                     disabled=not tem_validade)
        with c3:
            qtd = st.number_input("Quantidade declarada no papel", min_value=0,
                                  step=1, value=len(mov["animais"]),
                                  key=f"gta_qtd_{mov['id']}")

        opcoes_embarque = {(a["brinco"] or a["animal_uuid"][:8]): a["animal_uuid"]
                           for a in mov["animais"]}
        sel_embarque = st.multiselect(
            "Animais que realmente subiram no caminhão", list(opcoes_embarque),
            default=list(opcoes_embarque), key=f"gta_embarque_{mov['id']}",
            help="Pré-marcado com todos os animais da movimentação — desmarque "
                 "quem não embarcou de fato.")
        animais_no_embarque_uuids = [opcoes_embarque[s] for s in sel_embarque]

        brincos = [a["brinco"] for a in mov["animais"] if a["brinco"]]
        wd_batch = db.get_withdrawal_end_batch(brincos)
        hoje = date.today()
        animais_em_carencia_uuids = [
            a["animal_uuid"] for a in mov["animais"]
            if a["brinco"] and wd_batch.get(a["brinco"])
            and wd_batch[a["brinco"]] > hoje
        ]

        movimentacao_ctx = {
            "gta_numero": mov.get("gta_numero"),
            "propriedade_origem_nome": _rotulo_prop(props, mov.get("propriedade_origem_id")),
            "propriedade_destino_nome": _rotulo_prop(props, mov.get("propriedade_destino_id")),
            "finalidade": mov.get("finalidade"),
            "animais_uuids": [a["animal_uuid"] for a in mov["animais"]],
        }
        dados_do_documento = {
            "emissao": emissao.isoformat() if tem_emissao else None,
            "validade": validade.isoformat() if tem_validade else None,
            "quantidade_declarada": int(qtd),
        }

        gta, contexto = gta_montar_contexto(
            movimentacao_ctx, dados_do_documento, animais_no_embarque_uuids,
            animais_em_carencia_uuids, hoje.isoformat())
        problemas = gta_validar(gta, contexto)

        if problemas:
            for p in problemas:
                texto = f"{_GRAVIDADE_ICONE.get(p['gravidade'], '•')} {p['mensagem']}"
                if p["gravidade"] == "bloqueio":
                    st.error(texto)
                elif p["gravidade"] == "alerta":
                    st.warning(texto)
                else:
                    st.info(texto)
        else:
            st.success("✅ Nenhum problema encontrado na conferência da GTA.")


def _mov_liberar(mov):
    v = db.movimentacoes.pre_validar(mov["id"])
    if not v.get("ok"):
        st.error(f"🚫 {v['erro']}"); return

    problemas = v["problemas"]
    if problemas:
        st.markdown("**Pré-validação da saída (§8.3)**")
        for p in problemas:
            texto = f"{_GRAVIDADE_ICONE.get(p['gravidade'], '•')} {p['mensagem']}"
            if p["gravidade"] == "bloqueio":
                st.error(texto)
            elif p["gravidade"] == "alerta":
                st.warning(texto)
            else:
                st.info(texto)
    else:
        st.success("✅ Nenhum problema na pré-validação.")

    if not v["pode_liberar"]:
        st.error("🚫 Há bloqueio. Corrija a movimentação — não existe liberar "
                 "assim mesmo, e é isso que separa bloqueio de alerta.")
        return

    justificativa = ""
    if v["exige_confirmacao"]:
        # §8.4 pede "confirmação e justificativa". Uma caixa de seleção não
        # serve aqui: o que fica no registro é o texto de quem avaliou.
        justificativa = st.text_input(
            "Justificativa dos alertas *", key="mov_just",
            help="§8.4: a justificativa fica no evento e na auditoria. É o que "
                 "distingue quem avaliou de quem apenas clicou.").strip()

    pode = not v["exige_confirmacao"] or bool(justificativa)
    if st.button("🚚 Liberar saída", type="primary", disabled=not pode,
                 key="mov_liberar"):
        r = db.movimentacoes.liberar(mov["id"],
                                     usuario=st.session_state.user["name"],
                                     justificativa=justificativa)
        if r.get("ok"):
            db.clear_cache()
            st.success("✅ Saída liberada. Os animais receberam evento de saída.")
            st.rerun()
        else:
            st.error(f"🚫 {r.get('erro', 'Não foi possível liberar.')}")


def _mov_confirmar_chegada(mov):
    st.markdown("**Confirmação de chegada (§8.2)**")
    st.caption("Desmarque quem **não** chegou. Animal declarado e não recebido "
               "fica registrado como divergência, em vez de a movimentação "
               "inteira ser dada como concluída sem ressalva.")

    data = st.date_input("Data da chegada", value=date.today(), key="mov_chegada")

    recebidos = []
    for a in mov["animais"]:
        if st.checkbox(f"{a['brinco']}", value=True, key=f"mov_rec_{a['animal_uuid']}"):
            recebidos.append(a["animal_uuid"])

    faltantes = len(mov["animais"]) - len(recebidos)
    obs = ""
    if faltantes:
        st.warning(f"⚠️ {_plural(faltantes, 'animal', 'animais')} não recebido(s). "
                   "A movimentação será concluída **com divergência**.")
        obs = st.text_input("Observação da divergência", key="mov_div").strip()

    if st.button("✅ Confirmar chegada", type="primary", key="mov_confirmar"):
        r = db.movimentacoes.confirmar_chegada(
            mov["id"], data=data.isoformat(),
            usuario=st.session_state.user["name"],
            recebidos=recebidos, divergencias=obs)
        if r.get("ok"):
            db.clear_cache()
            if r["status"] == "divergente":
                st.warning(f"⚠️ Concluída com {len(r['nao_recebidos'])} "
                           "não recebido(s), registrado.")
            else:
                st.success("✅ Chegada confirmada. Os animais mudaram de propriedade.")
            st.rerun()
        else:
            st.error(f"🚫 {r.get('erro', 'Não foi possível confirmar.')}")


def _mov_nova(props):
    st.caption("A movimentação nasce em **rascunho** e não é validada agora — "
               "rascunho é onde se monta. A conferência do §8.3 acontece na aba "
               "**Em andamento**, antes de liberar a saída.")

    c1, c2 = st.columns(2)
    with c1:
        tipo = st.selectbox("Tipo *", list(_TIPO_MOV),
                            format_func=lambda t: _TIPO_MOV[t], key="movn_tipo")
    with c2:
        finalidade = st.text_input(
            "Finalidade", key="movn_fin",
            help="Escreva 'abate' ou 'frigorifico' quando for o caso: é o que "
                 "faz a carência virar bloqueio em vez de aviso.").strip()

    rot_p = {f"{p['nome']} — {p['produtor_nome']}": p["id"] for p in props}
    c3, c4 = st.columns(2)
    with c3:
        origem = rot_p[st.selectbox("Propriedade de origem *", list(rot_p),
                                    key="movn_origem")]
    with c4:
        destinos = [k for k, v in rot_p.items() if v != origem]
        destino = rot_p[st.selectbox("Propriedade de destino *", destinos,
                                     key="movn_destino")]

    c5, c6, c7 = st.columns(3)
    with c5:
        prevista = st.date_input("Data prevista", value=date.today(),
                                 key="movn_data")
    with c6:
        gta = st.text_input("Nº da GTA", key="movn_gta",
                            help="Pode ficar em branco no rascunho: sem GTA é "
                                 "alerta na liberação, não impedimento — a guia "
                                 "costuma sair depois de o lote estar montado."
                            ).strip()
    with c7:
        transportador = st.text_input("Transportador", key="movn_transp").strip()

    # Só animais da origem: levar um animal de outra propriedade é BLOQUEIO na
    # pré-validação, e oferecer o que a regra vai recusar é armadilha.
    candidatos = [a for a in db.get_all_animals(status="ativo")
                  if a.get("property_id") == origem]
    if not candidatos:
        st.warning("Nenhum animal ativo nesta propriedade de origem.")
        return

    rot_a = {f"{a['id']} — {a['breed']}": a["uuid"] for a in candidatos}
    escolhidos = st.multiselect(f"Animais * ({len(candidatos)} disponíveis)",
                                list(rot_a), key="movn_animais")

    if st.button("📝 Criar rascunho", type="primary",
                 disabled=not escolhidos, key="movn_salvar"):
        r = db.movimentacoes.criar(
            tipo, propriedade_origem_id=origem, propriedade_destino_id=destino,
            finalidade=finalidade, data_prevista=prevista.isoformat(),
            transportador=transportador, gta_numero=gta,
            animais=[rot_a[e] for e in escolhidos],
            usuario=st.session_state.user["name"])
        if r.get("ok"):
            db.clear_cache()
            st.success(f"✅ Rascunho criado com {len(escolhidos)} animal(is). "
                       "Vá em **Em andamento** para conferir e liberar.")
            st.rerun()
        else:
            st.error(f"🚫 {r['erro']}")

def _ler_poligono(texto):
    """Lê o perímetro digitado como `lon,lat` por linha.

    Aceita também `lat,lon`? **Não.** O GeoJSON e o EPSG:4326 usam
    (longitude, latitude), e aceitar as duas ordens obrigaria a adivinhar qual
    é qual — no Brasil ambos são negativos e a troca passa despercebida até a
    área sair errada.
    """
    anel = []
    for linha in (texto or "").strip().splitlines():
        linha = linha.strip().replace(";", ",")
        if not linha:
            continue
        partes = [p.strip() for p in linha.split(",")]
        if len(partes) != 2:
            raise ValueError(f"Linha inválida: {linha!r}. Use `longitude, latitude`.")
        anel.append((float(partes[0]), float(partes[1])))
    return anel


def _poligono_para_texto(geojson_txt):
    if not geojson_txt:
        return ""
    try:
        g = json.loads(geojson_txt)
        anel = g["coordinates"][0]
    except Exception:
        return ""
    return "\n".join(f"{lon}, {lat}" for lon, lat in anel)


def page_propriedades():
    """Hierarquia Organização → Produtor → Propriedade (PNIB §3).

    A hierarquia existe desde o B4 e nunca teve tela. É ela que responde "de
    quem é este animal e onde ele está" — a pergunta que a rastreabilidade
    persegue.
    """
    if st.session_state.user["role"] != "admin":
        st.error("🔒 Acesso restrito ao Administrador."); return
    st.markdown('<div class="page-title">🏞️ Propriedades</div>',
                unsafe_allow_html=True)

    props = db.propriedades.listar(apenas_ativas=False)
    t_lista, t_nova = st.tabs([f"📋 Cadastradas ({len(props)})", "➕ Nova propriedade"])
    with t_lista:
        _propriedades_editar(props)
    with t_nova:
        _propriedade_nova()


def _propriedades_editar(props):
    if not props:
        st.info("Nenhuma propriedade cadastrada.")
        return

    rot = {f"{p['nome']} — {p['produtor_nome']} "
           f"({'ativa' if p['situacao'] == 'ativa' else p['situacao']})": p
           for p in props}
    p = rot[st.selectbox("Propriedade", list(rot), key="prop_sel")]

    # O titular NÃO é editável: trocá-lo é transferência de titularidade, que é
    # evento do §8, com GTA e data. Oferecer aqui como campo de cadastro faria
    # uma mudança regulatória parecer correção de digitação.
    st.caption(f"Titular: **{p['produtor_nome']}** · Organização: "
               f"**{p['organizacao_nome']}**. O titular é definido na criação e "
               "só muda por transferência (§8), não por edição.")

    c1, c2, c3 = st.columns(3)
    with c1:
        nome = st.text_input("Nome *", value=p["nome"], key="prop_nome").strip()
    with c2:
        codigo = st.text_input("Código oficial", value=p.get("codigo_oficial") or "",
                               key="prop_codigo",
                               help="Cadastro da propriedade no órgão estadual (§3.2).").strip()
    with c3:
        situacao = st.selectbox(
            "Situação", ["ativa", "inativa", "encerrada"],
            index=["ativa", "inativa", "encerrada"].index(
                p["situacao"] if p["situacao"] in ("ativa", "inativa", "encerrada")
                else "ativa"),
            key="prop_situacao")

    c4, c5 = st.columns([2, 1])
    with c4:
        municipio = st.text_input("Município", value=p.get("municipio") or "",
                                  key="prop_municipio").strip()
    with c5:
        uf = st.text_input("UF", value=p.get("uf") or "", max_chars=2,
                           key="prop_uf").strip().upper()

    encerramento = ""
    if situacao == "encerrada":
        # Propriedade encerrada sem data não conta história nenhuma: o §3 pede
        # saber quando o vínculo terminou, não só que terminou.
        encerramento = st.text_input(
            "Data de encerramento * (AAAA-MM-DD)",
            value=p.get("encerramento") or "", key="prop_encerramento").strip()

    st.markdown("---")
    st.markdown("**Perímetro da propriedade**")
    st.caption("Um vértice por linha, no formato `longitude, latitude` — a ordem "
               "do GeoJSON. A área é **calculada**, não digitada: área digitada "
               "e perímetro desenhado divergem com o tempo, e aí ninguém sabe "
               "qual dos dois vale.")

    texto = st.text_area("Vértices", value=_poligono_para_texto(p.get("poligono")),
                         height=140, key="prop_poligono",
                         placeholder="-51.2300, -30.0300\n-51.2280, -30.0300\n"
                                     "-51.2280, -30.0320")

    anel, erro_leitura, problemas = [], "", []
    if texto.strip():
        try:
            anel = _ler_poligono(texto)
        except ValueError as e:
            erro_leitura = str(e)
        else:
            problemas = geometria_validar(anel)

    if erro_leitura:
        st.error(f"🚫 {erro_leitura}")
    for prob in problemas:
        st.error(f"🚫 {prob}")

    if anel and not problemas:
        lon, lat = geometria_centroide(anel)
        g1, g2, g3 = st.columns(3)
        g1.metric("Área", f"{_num_br(geometria_area_ha(anel), 2)} ha")
        g2.metric("Perímetro", f"{_num_br(geometria_perimetro_m(anel), 0)} m")
        g3.metric("Centro", f"{lat:.4f}, {lon:.4f}")

    pode = (bool(nome) and not erro_leitura and not problemas
            and (situacao != "encerrada" or bool(encerramento)))
    if st.button("💾 Salvar propriedade", type="primary", disabled=not pode,
                 key="prop_salvar"):
        campos = {"nome": nome, "codigo_oficial": codigo or None,
                  "municipio": municipio or None, "uf": uf or None,
                  "situacao": situacao,
                  "encerramento": encerramento or None}
        if anel:
            campos["poligono"] = json.dumps(
                {"type": "Polygon", "coordinates": [[list(v) for v in anel]]})
            lon, lat = geometria_centroide(anel)
            campos["longitude"], campos["latitude"] = lon, lat
        elif not texto.strip():
            campos["poligono"] = None

        if db.propriedades.atualizar(p["id"], **campos):
            db.clear_cache()
            st.success("✅ Propriedade atualizada.")
            st.rerun()
        else:
            st.error("🚫 Nada foi alterado.")


def _propriedade_nova():
    produtores = db.propriedades.listar_produtores()
    if not produtores:
        st.warning("Nenhum produtor cadastrado. A hierarquia do §3 é "
                   "Organização → Produtor → Propriedade, nessa ordem.")
        return

    rot = {f"{p['nome']} — {p['organizacao_nome']}": p["id"] for p in produtores}
    c1, c2 = st.columns(2)
    with c1:
        produtor = rot[st.selectbox("Titular (produtor) *", list(rot),
                                    key="propn_produtor",
                                    help="Escolhido agora e **imutável**: mudar "
                                         "depois é transferência (§8).")]
    with c2:
        nome = st.text_input("Nome da propriedade *", key="propn_nome").strip()

    c3, c4, c5 = st.columns(3)
    with c3:
        codigo = st.text_input("Código oficial", key="propn_codigo").strip()
    with c4:
        municipio = st.text_input("Município", key="propn_municipio").strip()
    with c5:
        uf = st.text_input("UF", max_chars=2, key="propn_uf").strip().upper()

    if st.button("➕ Cadastrar propriedade", type="primary", disabled=not nome,
                 key="propn_salvar"):
        db.propriedades.criar_propriedade(
            produtor, nome, codigo_oficial=codigo, municipio=municipio, uf=uf)
        db.clear_cache()
        st.success(f"✅ Propriedade **{nome}** cadastrada. "
                   "O perímetro pode ser desenhado na aba **Cadastradas**.")
        st.rerun()

_NIVEL_REGRA = {"informativo": "🔵 Informativo", "alerta": "🟡 Alerta",
                "bloqueio": "🔴 Bloqueio"}
_ESFERA_REGRA = {"federal": "Federal", "estadual": "Estadual",
                 "protocolo": "Protocolo (frigorífico, certificadora)",
                 "interna": "Interna da fazenda"}


def _casos_do_rebanho():
    """O rebanho atual no formato que `simular` espera.

    Simular contra o rebanho real, e não contra exemplos inventados, é o que
    torna o número da §11.3 uma resposta em vez de um exercício.
    """
    casos = []
    for a in db.get_all_animals(status="ativo"):
        casos.append({
            "id": a["id"],
            "especie": "bovino",
            "categoria": db.get_age_category(a.get("birth_date"), a.get("sex")),
            "sexo": a.get("sex"),
            "idade_meses": db.get_age_months(a.get("birth_date")),
            "peso": a.get("last_weight"),
            "raca": a.get("breed"),
        })
    return casos


def _mostrar_simulacao(regra_dict, casos):
    """§11.3: quantos animais esta regra atinge, hoje."""
    s = simular_regra_pura(regra_dict, casos)
    if not s["vigente_na_data"]:
        st.info("A regra não está vigente na data de hoje — a simulação mede o "
                "alcance dela quando entrar em vigor.")
    c1, c2, c3 = st.columns(3)
    c1.metric("Rebanho avaliado", s["total_avaliado"])
    c2.metric("Animais atingidos", s["atingidos"])
    c3.metric("Alcance", f"{_num_br(s['percentual'], 1)}%")
    if s["ids"]:
        with st.expander(f"Ver os {len(s['ids'])} atingidos"):
            st.dataframe(pd.DataFrame({"Brinco": s["ids"]}),
                         use_container_width=True, hide_index=True)
    return s


def page_regras():
    """Motor de regras regulatórias (PNIB §11).

    O §11 abre dizendo que as regras "não devem ficar fixadas no código-fonte".
    A tabela existe desde o B5 e nunca teve tela — o que significa que, na
    prática, elas continuavam fixadas: mudá-las exigia SQL.
    """
    if st.session_state.user["role"] != "admin":
        st.error("🔒 Acesso restrito ao Administrador."); return
    st.markdown('<div class="page-title">📜 Regras Regulatórias</div>',
                unsafe_allow_html=True)
    st.caption("§11: a portaria muda, cada UF acrescenta a sua e um frigorífico "
               "impõe protocolo próprio. Regra é **dado**, com vigência e "
               "versão — não `if` no código.")

    regras = db.regras.listar(apenas_ativas=False)
    t_lista, t_nova = st.tabs([f"📋 Cadastradas ({len(regras)})", "➕ Nova regra"])
    with t_lista:
        _regras_lista(regras)
    with t_nova:
        _regra_nova()


def _regras_lista(regras):
    if not regras:
        st.info("Nenhuma regra cadastrada.")
        return

    st.dataframe(
        pd.DataFrame([{
            "Nome": r["nome"], "Nível": _NIVEL_REGRA.get(r["nivel"], r["nivel"]),
            "Esfera": r["esfera"], "UF": r.get("uf") or "todas",
            "Versão": r.get("versao"),
            "Vigência": f"{r.get('data_inicial') or '—'} → "
                        f"{r.get('data_final') or 'em aberto'}",
            "Aprovada por": r.get("aprovado_por") or "— rascunho —",
        } for r in regras]),
        use_container_width=True, hide_index=True)

    rot = {f"{r['nome']} · v{r.get('versao')} · "
           f"{_NIVEL_REGRA.get(r['nivel'], r['nivel'])}": r for r in regras}
    r = rot[st.selectbox("Regra", list(rot), key="reg_sel")]

    if not r.get("aprovado_por"):
        st.warning("📝 **Rascunho.** Nasceu inativa porque não tem responsável "
                   "pela aprovação — o §11.1 pede o responsável justamente para "
                   "que exista a quem perguntar depois. Publique com uma nova "
                   "versão aprovada.")

    if r.get("descricao"):
        st.markdown(r["descricao"])
    if r.get("fundamento"):
        st.caption(f"Fundamento: {r['fundamento']}")

    st.markdown("**Alcance no rebanho de hoje (§11.3)**")
    _mostrar_simulacao(r, _casos_do_rebanho())

    st.markdown("---")
    # Não existe "editar". O §11.2 quer o histórico: uma movimentação julgada em
    # 2027 precisa continuar explicada pelo texto que valia então. Editar no
    # lugar reescreveria o passado.
    st.markdown("**Nova versão**")
    st.caption("Regra não se edita: cria-se outra versão. A anterior é encerrada "
               "com `data_final` de ontem e continua explicando o que foi julgado "
               "sob ela.")

    c1, c2 = st.columns(2)
    with c1:
        nivel = st.selectbox("Nível", list(_NIVEL_REGRA),
                             index=list(_NIVEL_REGRA).index(r["nivel"]),
                             format_func=lambda n: _NIVEL_REGRA[n],
                             key="reg_nivel")
    with c2:
        aprovador = st.text_input(
            "Responsável pela aprovação *", key="reg_aprovador",
            help="§11.1. Sem ele não há nova versão — decisão regulatória sem "
                 "responsável é decisão de ninguém.").strip()

    mensagem = st.text_input("Mensagem ao usuário",
                             value=r.get("mensagem") or "", key="reg_msg").strip()

    if st.button("📄 Publicar nova versão", type="primary",
                 disabled=not aprovador, key="reg_versao"):
        res = db.regras.nova_versao(
            r["id"], aprovado_por=aprovador,
            usuario=st.session_state.user["name"],
            nivel=nivel, mensagem=mensagem)
        if res.get("ok"):
            db.clear_cache()
            st.success(f"✅ Versão {res.get('versao', '')} publicada. "
                       "A anterior continua no histórico.")
            st.rerun()
        else:
            st.error(f"🚫 {res.get('erro', 'Não foi possível versionar.')}")


def _regra_nova():
    c1, c2, c3 = st.columns(3)
    with c1:
        nome = st.text_input("Nome *", key="regn_nome").strip()
    with c2:
        nivel = st.selectbox("Nível *", list(_NIVEL_REGRA),
                             format_func=lambda n: _NIVEL_REGRA[n], key="regn_nivel")
    with c3:
        esfera = st.selectbox("Esfera *", list(_ESFERA_REGRA),
                              format_func=lambda e: _ESFERA_REGRA[e],
                              key="regn_esfera")

    descricao = st.text_input("Descrição", key="regn_desc").strip()
    c4, c5 = st.columns(2)
    with c4:
        fundamento = st.text_input("Fundamento", key="regn_fund",
                                   placeholder="Portaria SDA/MAPA 1.331/2025, §7").strip()
    with c5:
        mensagem = st.text_input("Mensagem ao usuário", key="regn_msg").strip()

    st.markdown("**Escopo** — campo em branco significa *qualquer*, e é o que "
                "permite escrever uma regra federal sem enumerar as 27 UFs.")
    c6, c7, c8 = st.columns(3)
    with c6:
        uf = st.text_input("UF", max_chars=2, key="regn_uf").strip().upper()
    with c7:
        sexo = st.selectbox("Sexo", ["", "M", "F"], key="regn_sexo",
                            format_func=lambda s: "qualquer" if not s else s)
    with c8:
        finalidade = st.text_input("Finalidade", key="regn_fin").strip()

    c9, c10 = st.columns(2)
    with c9:
        idade_min = st.number_input("Idade mínima (meses)", min_value=0, max_value=300,
                                    value=0, key="regn_idade_min")
    with c10:
        idade_max = st.number_input("Idade máxima (meses)", min_value=0, max_value=300,
                                    value=0, key="regn_idade_max",
                                    help="0 = sem limite.")

    c11, c12 = st.columns(2)
    with c11:
        inicio = st.text_input("Vigência a partir de (AAAA-MM-DD)",
                               value=date.today().isoformat(), key="regn_inicio").strip()
    with c12:
        aprovador = st.text_input(
            "Responsável pela aprovação", key="regn_aprovador",
            help="§11.1. **Em branco, a regra nasce como rascunho inativo** — "
                 "não é recusa, é o estado correto de uma decisão sem "
                 "responsável.").strip()

    regra = {"nome": nome or "(sem nome)", "nivel": nivel, "esfera": esfera,
             "uf": uf or None, "sexo": sexo or None,
             "finalidade": finalidade or None,
             "idade_min_meses": int(idade_min) or None,
             "idade_max_meses": int(idade_max) or None,
             "data_inicial": inicio or None, "data_final": None,
             "condicao": None}

    # A simulação vem ANTES do botão, e não depois de salvar. Ativar bloqueio
    # sem saber o alcance é descobri-lo no dia em que o caminhão está no curral.
    st.markdown("---")
    st.markdown("**Alcance desta regra no rebanho de hoje (§11.3)**")
    s = _mostrar_simulacao(regra, _casos_do_rebanho())
    if nivel == "bloqueio" and s["atingidos"]:
        st.warning(f"⚠️ Como **bloqueio**, esta regra impediria a operação de "
                   f"{s['atingidos']} de {s['total_avaliado']} animais ativos.")

    if st.button("➕ Cadastrar regra", type="primary", disabled=not nome,
                 key="regn_salvar"):
        res = db.regras.criar(
            nome, nivel=nivel, esfera=esfera, descricao=descricao,
            fundamento=fundamento, mensagem=mensagem, uf=uf or None,
            sexo=sexo or None, finalidade=finalidade or None,
            idade_min_meses=int(idade_min) or None,
            idade_max_meses=int(idade_max) or None,
            data_inicial=inicio or None, aprovado_por=aprovador,
            usuario=st.session_state.user["name"])
        if res.get("ok"):
            db.clear_cache()
            if res.get("ativa"):
                st.success("✅ Regra cadastrada e **ativa**.")
            else:
                st.info("📝 Regra cadastrada como **rascunho inativo**: sem "
                        "responsável pela aprovação ela não vale (§11.1).")
            st.rerun()
        else:
            st.error(f"🚫 {res['erro']}")

def page_sincronizacao():
    """Painel de sincronização com o sistema oficial (PNIB §10.4).

    O §10.2 pede "log técnico e número de tentativas" — vem do histórico de
    `evento_sincronizacao`, não de contador guardado, que poderia mentir. O
    §10.4 pede "registro manual de protocolo" e "dupla conferência" **enquanto
    não houver API** (§23) — hoje. Esta tela é esse "enquanto".
    """
    if st.session_state.user["role"] != "admin":
        st.error("🔒 Acesso restrito ao Administrador."); return
    st.markdown('<div class="page-title">📡 Sincronização com o Sistema Oficial</div>',
                unsafe_allow_html=True)
    st.caption("A fila é o que separa **registrei** de **comuniquei** — e o PNIB cobra "
               "a segunda (§10). Um evento só some daqui quando alguém confirma que o "
               "sistema oficial aceitou, cancelou ou não se aplica a ele.")

    pendentes = db.eventos.pendentes_de_sincronizacao(limite=200)
    c1, c2 = st.columns(2)
    c1.metric("Eventos pendentes", len(pendentes))

    if not pendentes:
        st.success("✅ Nenhum evento pendente de comunicação.")
        return

    resumo = sincronizacao_resumo([e["situacao_sincronizacao"] for e in pendentes])
    c2.metric("Situações distintas", len(resumo))
    st.markdown("**Por situação atual**")
    st.dataframe(
        pd.DataFrame([{"Situação": sincronizacao_rotulo(s), "Quantidade": n}
                     for s, n in resumo.items()]),
        use_container_width=True, hide_index=True)

    t_lista, t_fechar = st.tabs(
        ["📋 Acompanhar (por evento)", "✅ Marcar como sincronizado (em lote)"])

    with t_lista:
        _sinc_acompanhar(pendentes)
    with t_fechar:
        _sinc_fechar_em_lote(pendentes)


def _sinc_acompanhar(pendentes: list[dict]):
    st.caption("Registra uma transição de qualquer situação do §10.3 — inclusive "
               "as que NÃO fecham a pendência (enviado, rejeitado, erro técnico). "
               "É o que dá o histórico técnico que o §10.2 pede.")

    rot = {f"#{e['id']} · {_TIPO_EVENTO_ROTULO.get(e['tipo'], e['tipo'])} · "
           f"{e['ocorrido_em'][:10]} · atual: {sincronizacao_rotulo(e['situacao_sincronizacao'])}": e
           for e in pendentes}
    evento = rot[st.selectbox("Evento", list(rot), key="sinc_ac_evento")]

    c1, c2 = st.columns(2)
    with c1:
        situacao = st.selectbox(
            "Nova situação", list(SITUACOES_SINCRONIZACAO),
            format_func=sincronizacao_rotulo, key="sinc_ac_situacao")
    with c2:
        sistema = st.selectbox("Sistema", list(SISTEMAS_SINCRONIZACAO),
                               key="sinc_ac_sistema")

    c3, c4 = st.columns(2)
    with c3:
        protocolo = st.text_input("Protocolo (se houver)",
                                  key="sinc_ac_protocolo").strip()
    with c4:
        mensagem = st.text_input(
            "Mensagem do sistema (erro, ressalva)", key="sinc_ac_mensagem").strip()

    if st.button("📝 Registrar transição", key="sinc_ac_salvar"):
        r = db.eventos.registrar_situacao(
            evento["id"], situacao, sistema=sistema,
            protocolo=protocolo or None, mensagem=mensagem,
            usuario=st.session_state.user["name"])
        if r.get("ok"):
            db.clear_cache()
            st.success("✅ Transição registrada.")
            st.rerun()
        else:
            st.error(f"🚫 {r.get('erro')}")

    with st.expander(f"Histórico técnico do evento #{evento['id']}"):
        hist = db.eventos.historico_de_sincronizacao(evento["id"])
        if hist:
            st.dataframe(
                pd.DataFrame([{"Quando": h["registrado_em"][:16].replace("T", " "),
                              "Sistema": h["sistema"], "Situação": sincronizacao_rotulo(h["situacao"]),
                              "Protocolo": h.get("protocolo") or "—",
                              "Conferido por": h.get("conferido_por") or "—"}
                             for h in hist]),
                use_container_width=True, hide_index=True)
        else:
            st.caption("Nenhuma transição registrada ainda para este evento.")


def _sinc_fechar_em_lote(pendentes: list[dict]):
    st.caption("Só oferece as situações que **encerram** a pendência (§10.3) — "
               "registrar aqui uma situação como 'rejeitado' esconderia uma "
               "obrigação de comunicar que continua de pé; use a aba ao lado.")

    rot = {f"#{e['id']} · {_TIPO_EVENTO_ROTULO.get(e['tipo'], e['tipo'])} · "
           f"{e['ocorrido_em'][:10]}": e for e in pendentes}
    escolhidos = st.multiselect("Eventos", list(rot), key="sinc_lote_eventos")

    c1, c2 = st.columns(2)
    with c1:
        situacao = st.selectbox(
            "Situação final", list(SITUACOES_RESOLVIDAS),
            format_func=sincronizacao_rotulo, key="sinc_lote_situacao")
    with c2:
        protocolo = st.text_input("Protocolo (se houver)",
                                  key="sinc_lote_protocolo").strip()

    conferido_por = st.text_input(
        "Conferido por", key="sinc_lote_conferido",
        help="§10.4 pede dupla conferência. Fica em branco se ninguém conferiu — "
             "e a ausência é auditável, não é erro de preenchimento.").strip()

    if not conferido_por:
        st.caption("⚠️ Sem conferente informado. Fica registrado assim mesmo.")

    if st.button("✅ Marcar selecionados", type="primary",
                 disabled=not escolhidos, key="sinc_lote_salvar"):
        ids = [rot[r]["id"] for r in escolhidos]
        r = db.eventos.marcar_sincronizado(
            ids, situacao=situacao, protocolo=protocolo or None,
            conferido_por=conferido_por, usuario=st.session_state.user["name"])
        if r.get("ok"):
            db.clear_cache()
            st.success(f"✅ {r['registrados']} evento(s) marcado(s) como "
                       f"{sincronizacao_rotulo(situacao).lower()}.")
            st.rerun()
        else:
            st.error(f"🚫 {r.get('erro')} (parou no evento #{r.get('parou_em')})")

def page_admin():
    if st.session_state.user["role"]!="admin":
        st.error("🔒 Acesso restrito ao Administrador."); return
    st.markdown('<div class="page-title">⚙️ Administração</div>', unsafe_allow_html=True)
    at_user,at1,at2,at3=st.tabs(["👥 Usuários","📋 Dados","🔧 Status Animais","🗄️ Banco"])

    with at_user:
        _admin_users()

    with at1:
        st.subheader("✏️ Edição Direta de Dados")
        st.caption("Corrija qualquer registro: edite células, adicione linhas (+) ou remova (🗑). "
                   "Clique em **Salvar alterações** para gravar no banco.")
        st.warning("⚠️ Área técnica: alterações são gravadas diretamente no banco. "
                   "Edite com cuidado — não há desfazer.")

        tab=st.selectbox("Tabela", db.ADMIN_TABLES, key="admin_tab_edit")
        cols, pk = db.admin_table_info(tab)
        orig_rows = db.admin_get_rows(tab)
        orig_df = pd.DataFrame(orig_rows, columns=cols)

        st.caption(f"Tabela **{tab}** · chave primária: **{pk}** · {len(orig_rows)} registro(s)")

        edited = st.data_editor(
            orig_df, num_rows="dynamic", use_container_width=True,
            hide_index=True, key=f"editor_{tab}",
            column_config={pk: st.column_config.Column(f"{pk} (chave)", help="Chave primária")},
        )

        cbtn1, cbtn2 = st.columns([1,3])
        with cbtn1:
            do_save = st.button("💾 Salvar alterações", type="primary",
                                use_container_width=True, key=f"savebtn_{tab}")
        with cbtn2:
            st.caption("Para inserir: use a linha em branco no fim da tabela. "
                       "Em tabelas com ID automático, deixe a chave vazia.")

        if do_save:
            def _pyval(v):
                if v is None: return None
                try:
                    if pd.isna(v): return None
                except (TypeError, ValueError): pass
                if hasattr(v, "item"):
                    try: v = v.item()
                    except Exception: pass
                if isinstance(v, float) and v.is_integer(): return int(v)
                return v

            def _is_empty_pk(v):
                pv = _pyval(v)
                return pv is None or (isinstance(pv, str) and pv.strip() == "")

            orig_by_pk = {_pyval(r[pk]): r for r in orig_rows}
            updates, inserts, seen = [], [], set()

            for rec in edited.to_dict("records"):
                clean = {col: _pyval(rec.get(col)) for col in cols}
                pkv = clean.get(pk)
                if _is_empty_pk(pkv) or pkv not in orig_by_pk:
                    row_ins = dict(clean)
                    if _is_empty_pk(pkv):
                        row_ins.pop(pk, None)   # deixa o banco gerar o ID
                    # ignora linhas totalmente vazias
                    if any(v is not None and str(v) != "" for v in row_ins.values()):
                        inserts.append(row_ins)
                else:
                    seen.add(pkv)
                    orig = orig_by_pk[pkv]
                    if any(str(_pyval(orig.get(col))) != str(clean.get(col)) for col in cols):
                        updates.append(clean)

            delete_pks = [k for k in orig_by_pk if k not in seen]

            try:
                res = db.admin_apply_changes(tab, updates, inserts, delete_pks)
                st.success(f"✅ Salvo em **{tab}**: {res['updated']} atualizada(s), "
                           f"{res['inserted']} inserida(s), {res['deleted']} excluída(s).")
                st.rerun()
            except Exception as e:
                st.error(f"❌ Erro ao salvar: {e}")

    with at2:
        st.subheader("Alterar Status de Animal")
        all_a=db.get_all_animals(status=None)
        por_id={a["id"]:a for a in all_a}
        ac1,ac2=st.columns(2)
        with ac1: sel_a=st.selectbox("Animal",[a["id"] for a in all_a])
        with ac2: new_st=st.selectbox("Novo Status",["ativo","vendido","morto","carencia"])

        atual = por_id[sel_a]["status"] if sel_a in por_id else ""
        # A autorização vem do PAPEL, e não de ter chegado nesta página. O
        # dispatch já restringe operador a OPERATOR_PAGES, mas quem decide uma
        # transição sensível é a permissão do usuário (§14.2) — não a rota.
        eh_admin = st.session_state.user["role"] == "admin"
        veredito = db.transicao_permitida(atual, new_st, tem_autorizacao=eh_admin)

        st.caption(f"Status atual: **{atual or '—'}** → **{new_st}**")

        justificativa = ""
        if veredito["exige_justificativa"]:
            st.warning(f"⚠️ {veredito['motivo']}")
            justificativa = st.text_area(
                "Justificativa (obrigatória)",
                placeholder="Por que este animal está saindo de um estado final?",
                key=f"just_{sel_a}_{new_st}")
        elif not veredito["permitida"]:
            st.error(f"🚫 {veredito['motivo']}")

        pode = veredito["permitida"] and (
            not veredito["exige_justificativa"] or justificativa.strip())
        if st.button("✅ Atualizar",type="primary",disabled=not pode):
            r = db.update_animal_status(
                sel_a, new_st,
                tem_autorizacao=eh_admin,
                justificativa=justificativa,
                operador=st.session_state.user["name"])
            if r.get("ok"):
                st.success(f"{sel_a}: {r['de']} → {r['para']}")
                st.rerun()
            else:
                st.error(f"🚫 {r.get('motivo','Transição recusada.')}")

    with at3:
        import os
        if db.USE_PG:
            st.markdown("**Banco de Dados:** PostgreSQL / Supabase (nuvem) ☁️")
            st.caption("Os dados ficam no Supabase e são acessíveis de qualquer lugar.")
            for t in db.ADMIN_TABLES:
                try:
                    n = db.admin_get_rows(t)
                    st.write(f"• `{t}`: {len(n)} registro(s)")
                except Exception:
                    pass
        else:
            st.markdown("**Banco de Dados:** SQLite local `agrotop.db` — funciona offline")
            if os.path.exists(db.DB_PATH):
                st.metric("Tamanho",f"{os.path.getsize(db.DB_PATH)/1024:.1f} KB")

        st.markdown("---")
        st.subheader("💾 Backup dos Dados")
        st.caption("Baixe uma cópia completa de todos os dados em Excel (uma aba por tabela). "
                   "Guarde em local seguro (pen drive, nuvem). O Supabase também mantém "
                   "backups automáticos diários no servidor.")
        if st.button("📥 Gerar backup completo (Excel)", type="primary"):
            data = _backup_xlsx()
            st.download_button(
                "⬇️ Baixar backup agora", data,
                f"agrotop_backup_{date.today().isoformat()}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)

# ══════════════════════════════════════════════════════════════════════════════
# ROTEADOR
# ══════════════════════════════════════════════════════════════════════════════
# Páginas que o operador pode acessar (as demais são exclusivas do admin)
# Aplicar brinco é trabalho de curral, não de escritório — por isso o
# operador entra em "brincos".
OPERATOR_PAGES = {"campo", "cadastrar", "estoque", "brincos"}

_COOKIE_NAME = "agrotop_sid"

def _init_cookies():
    """Cria a instância de cookies para ESTA execução. Precisa ser recriada a
    cada run para reler os cookies do navegador (o CookieManager lê no __init__)."""
    try:
        import extra_streamlit_components as stx
        st.session_state["_cm"] = stx.CookieManager(key="agrotop_cookie_mgr")
    except Exception:
        st.session_state["_cm"] = None

def _cookie_manager():
    """Retorna a instância de cookies desta execução (ou None se indisponível)."""
    return st.session_state.get("_cm")

def _try_restore_session():
    """Restaura o login a partir do token guardado em cookie (mantém login ao recarregar)."""
    if st.session_state.authenticated:
        return
    cm = _cookie_manager()
    if cm is None:
        return
    try:
        token = cm.get(_COOKIE_NAME)
    except Exception:
        token = None
    if token:
        u = db.get_session_user(token)
        if u:
            st.session_state.authenticated = True
            st.session_state.user = u
            st.session_state.page = "dashboard" if u["role"] == "admin" else "campo"

def main():
    tema = cores("escuro")
    st.markdown(
        f"""
        <link rel="manifest" href="/app/static/manifest.json">
        <link rel="apple-touch-icon" href="/app/static/icon-192.png">
        <meta name="theme-color" content="{tema['primaria']}">
        """,
        unsafe_allow_html=True,
    )
    _init_cookies()          # instância nova a cada run (relê o cookie)
    _try_restore_session()
    if not st.session_state.authenticated:
        page_login(); return

    user = st.session_state.user
    # Controle de acesso: operador só acessa páginas permitidas
    if user["role"] != "admin" and st.session_state.page not in OPERATOR_PAGES:
        st.session_state.page = "campo"

    _sidebar()
    {
        "dashboard": page_dashboard,
        "campo":     page_campo,
        "rebanho":   page_rebanho,
        "animal":    page_animal,
        "lotes":     page_lotes,
        "desempenho":page_desempenho,
        "financeiro":page_financeiro,
        "estoque":   page_estoque,
        "brincos":   page_brincos,
        "movimentacao": page_movimentacao,
        "propriedades": page_propriedades,
        "regras":    page_regras,
        "sincronizacao": page_sincronizacao,
        "nutricao":  page_nutricao,
        "sanitario": page_sanitario,
        "clima":     page_clima,
        "alertas":   page_alertas,
        "relatorios":page_relatorios,
        "cadastrar": page_cadastrar,
        "admin":     page_admin,
    }.get(st.session_state.page, page_campo if user["role"]!="admin" else page_dashboard)()

if __name__ == "__main__":
    main()